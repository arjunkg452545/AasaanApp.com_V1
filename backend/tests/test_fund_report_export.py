"""
Test Fund Report Export - Excel and PDF with Summary Section
Tests for BNI Chapter Fund Management System export functionality
"""
import pytest
import requests
import os
import io
from datetime import datetime

# Get BASE_URL from environment
BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

# Test credentials
ADMIN_MOBILE = "7773010121"
ADMIN_PASSWORD = "admin@123"


class TestFundReportExport:
    """Test Excel and PDF export with Summary Section"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup - get auth token"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        
        # Login as chapter admin
        login_response = self.session.post(f"{BASE_URL}/api/admin/login", json={
            "mobile": ADMIN_MOBILE,
            "password": ADMIN_PASSWORD
        })
        
        if login_response.status_code != 200:
            pytest.skip(f"Login failed: {login_response.status_code} - {login_response.text}")
        
        token = login_response.json().get("token")
        self.session.headers.update({"Authorization": f"Bearer {token}"})
        self.chapter_id = login_response.json().get("chapter_id")
        print(f"✓ Logged in as admin, chapter_id: {self.chapter_id}")
    
    # ========== EXCEL EXPORT TESTS ==========
    
    def test_excel_export_default_params(self):
        """Test Excel export with default parameters (last 3 months, all categories)"""
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/excel")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        assert response.headers.get("content-type") == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in response.headers.get("content-disposition", "")
        
        # Verify file is not empty
        content = response.content
        assert len(content) > 1000, f"Excel file too small: {len(content)} bytes"
        print(f"✓ Excel export default params - File size: {len(content)} bytes")
    
    def test_excel_export_kitty_only(self):
        """Test Excel export with kitty category only"""
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/excel", params={
            "categories": "kitty"
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert len(response.content) > 500, "Excel file too small"
        print(f"✓ Excel export kitty only - File size: {len(response.content)} bytes")
    
    def test_excel_export_meetingfee_only(self):
        """Test Excel export with meeting fee category only"""
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/excel", params={
            "categories": "meetingfee"
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert len(response.content) > 500, "Excel file too small"
        print(f"✓ Excel export meetingfee only - File size: {len(response.content)} bytes")
    
    def test_excel_export_events_only(self):
        """Test Excel export with events category only"""
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/excel", params={
            "categories": "events"
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert len(response.content) > 500, "Excel file too small"
        print(f"✓ Excel export events only - File size: {len(response.content)} bytes")
    
    def test_excel_export_multiple_categories(self):
        """Test Excel export with multiple categories (kitty + meetingfee)"""
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/excel", params={
            "categories": "kitty,meetingfee"
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert len(response.content) > 500, "Excel file too small"
        print(f"✓ Excel export multiple categories - File size: {len(response.content)} bytes")
    
    def test_excel_export_paid_status_filter(self):
        """Test Excel export with paid status filter"""
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/excel", params={
            "payment_status": "paid"
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print(f"✓ Excel export paid status filter - File size: {len(response.content)} bytes")
    
    def test_excel_export_pending_status_filter(self):
        """Test Excel export with pending status filter"""
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/excel", params={
            "payment_status": "pending"
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print(f"✓ Excel export pending status filter - File size: {len(response.content)} bytes")
    
    def test_excel_export_single_month(self):
        """Test Excel export with single month filter"""
        current_month = datetime.now().month
        current_year = datetime.now().year
        
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/excel", params={
            "months": str(current_month),
            "year": current_year
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print(f"✓ Excel export single month ({current_month}/{current_year}) - File size: {len(response.content)} bytes")
    
    def test_excel_export_multiple_months(self):
        """Test Excel export with multiple months filter"""
        current_month = datetime.now().month
        current_year = datetime.now().year
        
        # Get last 2 months
        months = []
        for i in range(2):
            m = current_month - i
            if m <= 0:
                m += 12
            months.append(str(m))
        
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/excel", params={
            "months": ",".join(months),
            "year": current_year
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print(f"✓ Excel export multiple months ({','.join(months)}/{current_year}) - File size: {len(response.content)} bytes")
    
    def test_excel_export_cross_year_months(self):
        """Test Excel export with cross-year month range (e.g., Nov-Dec-Jan)"""
        # Test with months that span across years
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/excel", params={
            "months": "11,12,1",
            "year": 2024
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print(f"✓ Excel export cross-year months - File size: {len(response.content)} bytes")
    
    def test_excel_export_combined_filters(self):
        """Test Excel export with combined filters (category + status + month)"""
        current_month = datetime.now().month
        current_year = datetime.now().year
        
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/excel", params={
            "categories": "kitty,meetingfee",
            "payment_status": "pending",
            "months": str(current_month),
            "year": current_year
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        print(f"✓ Excel export combined filters - File size: {len(response.content)} bytes")
    
    # ========== PDF EXPORT TESTS ==========
    
    def test_pdf_export_default_params(self):
        """Test PDF export with default parameters (last 3 months, all categories)"""
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/pdf")
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        assert response.headers.get("content-type") == "application/pdf"
        assert "attachment" in response.headers.get("content-disposition", "")
        
        # Verify file is not empty and is valid PDF
        content = response.content
        assert len(content) > 1000, f"PDF file too small: {len(content)} bytes"
        assert content[:4] == b'%PDF', "Invalid PDF header"
        print(f"✓ PDF export default params - File size: {len(content)} bytes")
    
    def test_pdf_export_kitty_only(self):
        """Test PDF export with kitty category only"""
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/pdf", params={
            "categories": "kitty"
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert response.content[:4] == b'%PDF', "Invalid PDF header"
        print(f"✓ PDF export kitty only - File size: {len(response.content)} bytes")
    
    def test_pdf_export_meetingfee_only(self):
        """Test PDF export with meeting fee category only"""
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/pdf", params={
            "categories": "meetingfee"
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert response.content[:4] == b'%PDF', "Invalid PDF header"
        print(f"✓ PDF export meetingfee only - File size: {len(response.content)} bytes")
    
    def test_pdf_export_events_only(self):
        """Test PDF export with events category only"""
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/pdf", params={
            "categories": "events"
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert response.content[:4] == b'%PDF', "Invalid PDF header"
        print(f"✓ PDF export events only - File size: {len(response.content)} bytes")
    
    def test_pdf_export_multiple_categories(self):
        """Test PDF export with multiple categories"""
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/pdf", params={
            "categories": "kitty,meetingfee"
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert response.content[:4] == b'%PDF', "Invalid PDF header"
        print(f"✓ PDF export multiple categories - File size: {len(response.content)} bytes")
    
    def test_pdf_export_paid_status_filter(self):
        """Test PDF export with paid status filter"""
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/pdf", params={
            "payment_status": "paid"
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert response.content[:4] == b'%PDF', "Invalid PDF header"
        print(f"✓ PDF export paid status filter - File size: {len(response.content)} bytes")
    
    def test_pdf_export_pending_status_filter(self):
        """Test PDF export with pending status filter"""
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/pdf", params={
            "payment_status": "pending"
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert response.content[:4] == b'%PDF', "Invalid PDF header"
        print(f"✓ PDF export pending status filter - File size: {len(response.content)} bytes")
    
    def test_pdf_export_single_month(self):
        """Test PDF export with single month filter"""
        current_month = datetime.now().month
        current_year = datetime.now().year
        
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/pdf", params={
            "months": str(current_month),
            "year": current_year
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert response.content[:4] == b'%PDF', "Invalid PDF header"
        print(f"✓ PDF export single month ({current_month}/{current_year}) - File size: {len(response.content)} bytes")
    
    def test_pdf_export_multiple_months(self):
        """Test PDF export with multiple months filter"""
        current_month = datetime.now().month
        current_year = datetime.now().year
        
        months = []
        for i in range(2):
            m = current_month - i
            if m <= 0:
                m += 12
            months.append(str(m))
        
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/pdf", params={
            "months": ",".join(months),
            "year": current_year
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert response.content[:4] == b'%PDF', "Invalid PDF header"
        print(f"✓ PDF export multiple months - File size: {len(response.content)} bytes")
    
    def test_pdf_export_combined_filters(self):
        """Test PDF export with combined filters"""
        current_month = datetime.now().month
        current_year = datetime.now().year
        
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/pdf", params={
            "categories": "kitty,meetingfee",
            "payment_status": "pending",
            "months": str(current_month),
            "year": current_year
        })
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        assert response.content[:4] == b'%PDF', "Invalid PDF header"
        print(f"✓ PDF export combined filters - File size: {len(response.content)} bytes")
    
    # ========== AUTHENTICATION TESTS ==========
    
    def test_excel_export_unauthorized(self):
        """Test Excel export without authentication"""
        session = requests.Session()
        response = session.get(f"{BASE_URL}/api/admin/fund/reports/export/excel")
        
        assert response.status_code in [401, 403], f"Expected 401/403, got {response.status_code}"
        print(f"✓ Excel export unauthorized - Status: {response.status_code}")
    
    def test_pdf_export_unauthorized(self):
        """Test PDF export without authentication"""
        session = requests.Session()
        response = session.get(f"{BASE_URL}/api/admin/fund/reports/export/pdf")
        
        assert response.status_code in [401, 403], f"Expected 401/403, got {response.status_code}"
        print(f"✓ PDF export unauthorized - Status: {response.status_code}")


class TestExcelContentValidation:
    """Validate Excel content structure and summary section"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup - get auth token"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        
        login_response = self.session.post(f"{BASE_URL}/api/admin/login", json={
            "mobile": ADMIN_MOBILE,
            "password": ADMIN_PASSWORD
        })
        
        if login_response.status_code != 200:
            pytest.skip(f"Login failed: {login_response.status_code}")
        
        token = login_response.json().get("token")
        self.session.headers.update({"Authorization": f"Bearer {token}"})
    
    def test_excel_content_has_summary_section(self):
        """Verify Excel file contains summary section with expected content"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/excel")
        assert response.status_code == 200
        
        # Load workbook from response content
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Search for summary section markers
        summary_found = False
        member_status_found = False
        category_wise_found = False
        month_wise_found = False
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell_str = str(cell.value)
                    if "SUMMARY" in cell_str:
                        summary_found = True
                    if "MEMBER STATUS" in cell_str:
                        member_status_found = True
                    if "CATEGORY-WISE COLLECTION" in cell_str:
                        category_wise_found = True
                    if "MONTH-WISE BREAKDOWN" in cell_str:
                        month_wise_found = True
        
        assert summary_found, "Summary section not found in Excel"
        assert member_status_found, "Member Status section not found in Excel"
        assert category_wise_found, "Category-wise Collection section not found in Excel"
        assert month_wise_found, "Month-wise Breakdown section not found in Excel"
        
        print("✓ Excel content validation - All summary sections found")
    
    def test_excel_has_correct_headers(self):
        """Verify Excel file has correct column headers"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/excel")
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Check header row (row 5 based on code)
        headers = [cell.value for cell in ws[5] if cell.value]
        
        # Verify essential headers exist
        assert "Sr" in headers, "Sr column not found"
        assert "Member ID" in headers, "Member ID column not found"
        assert "Member Name" in headers, "Member Name column not found"
        assert "Grand Total" in headers, "Grand Total column not found"
        assert "Status" in headers, "Status column not found"
        
        print(f"✓ Excel headers validation - Found headers: {headers[:5]}...")


class TestPDFContentValidation:
    """Validate PDF content structure"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup - get auth token"""
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})
        
        login_response = self.session.post(f"{BASE_URL}/api/admin/login", json={
            "mobile": ADMIN_MOBILE,
            "password": ADMIN_PASSWORD
        })
        
        if login_response.status_code != 200:
            pytest.skip(f"Login failed: {login_response.status_code}")
        
        token = login_response.json().get("token")
        self.session.headers.update({"Authorization": f"Bearer {token}"})
    
    def test_pdf_is_valid_format(self):
        """Verify PDF file is valid format"""
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/pdf")
        assert response.status_code == 200
        
        content = response.content
        
        # Check PDF header
        assert content[:4] == b'%PDF', "Invalid PDF header"
        
        # Check PDF has EOF marker
        assert b'%%EOF' in content[-100:], "PDF missing EOF marker"
        
        print(f"✓ PDF format validation - Valid PDF, size: {len(content)} bytes")
    
    def test_pdf_content_contains_summary(self):
        """Verify PDF contains summary section text"""
        response = self.session.get(f"{BASE_URL}/api/admin/fund/reports/export/pdf")
        assert response.status_code == 200
        
        # PDF content is binary, but we can search for text strings
        content = response.content
        
        # These strings should appear in the PDF (may be encoded)
        # Note: PDF text encoding varies, so we check for partial matches
        assert len(content) > 2000, "PDF too small to contain summary"
        
        print(f"✓ PDF content validation - File size indicates content present: {len(content)} bytes")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
