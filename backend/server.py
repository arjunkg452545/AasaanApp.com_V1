from dotenv import load_dotenv
from pathlib import Path
import os

# Load environment variables FIRST before any other imports
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, File, UploadFile, Request
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import logging
from datetime import datetime, timezone, timedelta
import io
import pytz

from models import *
from auth import create_access_token, verify_token, hash_password, verify_password
from qr_generator import generate_qr_token, create_qr_image, verify_qr_token
from report_generator import generate_excel_report, generate_pdf_report

# Shared database connection (also used by route modules via database.py)
from database import db, client

# IST Timezone
IST = pytz.timezone('Asia/Kolkata')

app = FastAPI()
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Health check endpoint
@api_router.get("/")
async def health_check():
    return {"status": "ok", "message": "BNI Attendance API is running"}

# Dependency for auth
async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    payload = verify_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid token")
    return payload

# Role-based access helpers
def require_role(*allowed_roles):
    """Dependency factory: returns a dependency that checks the user's role."""
    async def _check(user=Depends(get_current_user)):
        if user.get("role") not in allowed_roles:
            raise HTTPException(status_code=403, detail="Forbidden")
        return user
    return _check

# ===== SUPER ADMIN ENDPOINTS =====
@api_router.post("/superadmin/login", response_model=LoginResponse)
async def superadmin_login(data: LoginRequest):
    admin = await db.superadmins.find_one({"mobile": data.mobile}, {"_id": 0})
    if not admin or not verify_password(data.password, admin["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # Check if super admin is active
    if admin.get("is_active") is False:
        raise HTTPException(status_code=403, detail="Account is deactivated. Contact developer admin.")

    token = create_access_token({"mobile": data.mobile, "role": "superadmin"})
    return LoginResponse(token=token, role="superadmin", mobile=data.mobile)

@api_router.post("/superadmin/chapters")
async def create_chapter(chapter: ChapterCreateEnhanced, user=Depends(get_current_user)):
    if user["role"] not in ("superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    # --- Subscription enforcement (skip for developer role) ---
    if user["role"] == "superadmin":
        mobile = user.get("mobile", "")
        sa = await db.superadmins.find_one({"mobile": mobile}, {"_id": 0})
        sa_id = sa.get("superadmin_id", mobile) if sa else mobile

        sub = await db.subscriptions.find_one(
            {"superadmin_id": sa_id, "status": "active"},
            {"_id": 0},
            sort=[("end_date", -1)]
        )
        if not sub:
            raise HTTPException(status_code=403, detail="Please recharge to create chapters. No active subscription found.")

        # Check expiry
        end_dt = datetime.fromisoformat(sub["end_date"]) if isinstance(sub["end_date"], str) else sub["end_date"]
        if end_dt < datetime.now(timezone.utc):
            await db.subscriptions.update_one({"subscription_id": sub["subscription_id"]}, {"$set": {"status": "expired"}})
            raise HTTPException(status_code=403, detail="Your subscription has expired. Please recharge to create chapters.")

        # Check chapter limit
        chapters_used = await db.chapters.count_documents({"created_by": mobile})
        if chapters_used >= sub.get("chapters_allowed", 0):
            raise HTTPException(status_code=403, detail="Chapter limit reached. Please upgrade your subscription.")

    admin_password_hash = hash_password(chapter.admin_password)
    chapter_data = {
        "chapter_id": f"CH{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}",
        "name": chapter.name,
        "created_by": user.get("mobile", user.get("email", "")),
        "admin_mobile": chapter.admin_mobile,
        "admin_password_hash": admin_password_hash,
        "region": chapter.region,
        "state": chapter.state,
        "city": chapter.city,
        "status": "active",
        "audit_logs": [],
        "created_at": datetime.now(timezone.utc).isoformat()
    }

    await db.chapters.insert_one(chapter_data)
    return {k: v for k, v in chapter_data.items() if k not in ("admin_password_hash", "_id")}

@api_router.get("/superadmin/chapters", response_model=List[ChapterResponse])
async def get_all_chapters(user=Depends(get_current_user)):
    if user["role"] not in ("superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    chapters = await db.chapters.find({}, {"_id": 0, "admin_password_hash": 0}).to_list(1000)
    return chapters

@api_router.put("/superadmin/chapters/{chapter_id}/credentials")
async def update_chapter_credentials(chapter_id: str, data: UpdateCredentials, user=Depends(get_current_user)):
    if user["role"] not in ("superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    new_password_hash = hash_password(data.new_password)
    audit_log = {
        "changed_by": user["mobile"],
        "changed_at": datetime.now(timezone.utc).isoformat(),
        "new_mobile": data.new_mobile
    }
    
    result = await db.chapters.update_one(
        {"chapter_id": chapter_id},
        {"$set": {"admin_mobile": data.new_mobile, "admin_password_hash": new_password_hash},
         "$push": {"audit_logs": audit_log}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Chapter not found")
    
    return {"message": "Credentials updated successfully"}

@api_router.delete("/superadmin/chapters/{chapter_id}")
async def delete_chapter(chapter_id: str, user=Depends(get_current_user)):
    if user["role"] not in ("superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    # Delete chapter
    result = await db.chapters.delete_one({"chapter_id": chapter_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Chapter not found")
    
    # Delete all related data
    await db.members.delete_many({"chapter_id": chapter_id})
    await db.meetings.delete_many({"chapter_id": chapter_id})
    
    # Delete attendance records for all meetings of this chapter
    meetings = await db.meetings.find({"chapter_id": chapter_id}, {"_id": 0, "meeting_id": 1}).to_list(1000)
    meeting_ids = [m["meeting_id"] for m in meetings]
    if meeting_ids:
        await db.attendance.delete_many({"meeting_id": {"$in": meeting_ids}})
    
    return {"message": "Chapter and all related data deleted successfully"}

@api_router.get("/superadmin/chapters/{chapter_id}/audit-logs")
async def get_audit_logs(chapter_id: str, user=Depends(get_current_user)):
    if user["role"] not in ("superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    chapter = await db.chapters.find_one({"chapter_id": chapter_id}, {"_id": 0, "audit_logs": 1})
    if not chapter:
        raise HTTPException(status_code=404, detail="Chapter not found")
    
    return {"audit_logs": chapter.get("audit_logs", [])}

# ===== CHAPTER ADMIN ENDPOINTS =====
@api_router.post("/admin/login", response_model=LoginResponse)
async def admin_login(data: LoginRequest):
    chapter = await db.chapters.find_one({"admin_mobile": data.mobile}, {"_id": 0})
    if not chapter or not verify_password(data.password, chapter["admin_password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_access_token({"mobile": data.mobile, "role": "admin", "chapter_id": chapter["chapter_id"], "chapter_name": chapter.get("name", "")})
    return LoginResponse(token=token, role="admin", mobile=data.mobile, chapter_id=chapter["chapter_id"], chapter_name=chapter.get("name", ""))

@api_router.get("/admin/profile")
async def get_admin_profile(user=Depends(get_current_user)):
    """Get admin profile with chapter name"""
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    chapter = await db.chapters.find_one({"chapter_id": user.get("chapter_id")}, {"_id": 0, "admin_password_hash": 0})
    return {
        "mobile": user.get("mobile"),
        "chapter_id": user.get("chapter_id"),
        "chapter_name": chapter.get("name", "") if chapter else user.get("chapter_name", ""),
        "role": user.get("role")
    }

@api_router.get("/admin/members")
async def get_members(
    search: str = "",
    status_filter: str = "",
    category: str = "",
    sort_by: str = "unique_member_id",
    user=Depends(get_current_user)
):
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    chapter_id = user.get("chapter_id")
    if not chapter_id:
        raise HTTPException(status_code=400, detail="Chapter ID required")

    query = {"chapter_id": chapter_id, "archived": {"$ne": True}}

    if status_filter:
        query["membership_status"] = status_filter
    if category:
        query["business_category"] = {"$regex": category, "$options": "i"}
    if search:
        query["$or"] = [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"primary_mobile": {"$regex": search, "$options": "i"}},
            {"business_name": {"$regex": search, "$options": "i"}},
            {"unique_member_id": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
        ]

    sort_field = sort_by if sort_by in ("full_name", "unique_member_id", "created_at", "renewal_date") else "unique_member_id"
    members = await db.members.find(query, {"_id": 0}).sort(sort_field, 1).to_list(2000)
    return members

@api_router.post("/admin/members", response_model=MemberResponse)
async def add_member(member: MemberCreate, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    chapter_id = user.get("chapter_id")

    # SuperAdmin: resolve chapter_id from their chapters
    if user["role"] == "superadmin" and not chapter_id:
        sa = await db.superadmins.find_one({"mobile": user.get("mobile")}, {"_id": 0})
        sa_id = sa.get("superadmin_id", user.get("mobile")) if sa else user.get("mobile")
        first_ch = await db.chapters.find_one({"created_by": sa_id}, {"_id": 0, "chapter_id": 1})
        if first_ch:
            chapter_id = first_ch["chapter_id"]
        else:
            raise HTTPException(status_code=400, detail="No chapter found for this SuperAdmin")

    if not chapter_id:
        raise HTTPException(status_code=400, detail="Chapter ID required")

    # Check if member with same unique_member_id already exists in this chapter
    existing_member = await db.members.find_one({
        "chapter_id": chapter_id,
        "unique_member_id": member.unique_member_id
    })

    if existing_member:
        raise HTTPException(status_code=400, detail=f"Member with ID {member.unique_member_id} already exists in this chapter")

    # Check if mobile number already exists in this chapter
    existing_mobile = await db.members.find_one({
        "chapter_id": chapter_id,
        "primary_mobile": member.primary_mobile
    })

    if existing_mobile:
        raise HTTPException(status_code=400, detail="Member with this mobile number already exists")

    from uuid import uuid4

    # Approval workflow: admin → pending, superadmin/developer → active
    if user["role"] == "admin":
        membership_status = "pending"
        sync_status = "Inactive"  # pending members should not appear in fund/attendance
    else:
        membership_status = "active"
        sync_status = "Active"

    member_data = {
        "member_id": str(uuid4()),
        "chapter_id": chapter_id,
        "unique_member_id": member.unique_member_id,
        "full_name": member.full_name,
        "primary_mobile": member.primary_mobile,
        "secondary_mobile": member.secondary_mobile,
        "email": member.email,
        "business_name": member.business_name,
        "business_category": member.business_category,
        "joining_date": member.joining_date,
        "renewal_date": member.renewal_date,
        "induction_fee": member.induction_fee,
        "membership_status": membership_status,
        "status": sync_status,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "bni_member_id": member.unique_member_id,
        "organization_id": chapter_id,
        "status_history": [{
            "action": "created",
            "from_status": None,
            "to_status": membership_status,
            "reason": "Member created",
            "changed_by": user.get("mobile", user.get("email", "system")),
            "timestamp": datetime.now(timezone.utc).isoformat()
        }],
        "archived": False,
        "transfer_from_chapter": None,
        "transfer_date": None,
    }

    try:
        await db.members.insert_one(member_data)
        return MemberResponse(**member_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to add member: {str(e)}")

@api_router.post("/admin/members/bulk")
async def bulk_add_members(members: List[MemberCreate], user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    from uuid import uuid4
    
    # Get existing members to check for duplicates
    existing_members = await db.members.find(
        {"chapter_id": user["chapter_id"]},
        {"unique_member_id": 1, "primary_mobile": 1}
    ).to_list(1000)
    
    existing_ids = {m.get("unique_member_id") for m in existing_members}
    existing_mobiles = {m.get("primary_mobile") for m in existing_members}
    
    inserted = 0
    skipped = []
    errors = []
    duplicate_mobiles = []
    
    for idx, member in enumerate(members):
        # Check for duplicates in DB
        if member.unique_member_id in existing_ids:
            skipped.append(f"Row {idx+1}: Member ID {member.unique_member_id} already exists")
            continue
        
        if member.primary_mobile in existing_mobiles:
            skipped.append(f"Row {idx+1}: Mobile {member.primary_mobile} already exists")
            duplicate_mobiles.append(member.primary_mobile)
            continue
        
        # Duplicates already checked above - no need for additional DB query
        member_data = {
            "member_id": str(uuid4()),  # Use UUID for guaranteed uniqueness
            "chapter_id": user["chapter_id"],
            "unique_member_id": member.unique_member_id,
            "full_name": member.full_name,
            "primary_mobile": member.primary_mobile,
            "secondary_mobile": member.secondary_mobile,
            "status": member.status,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "bni_member_id": member.unique_member_id,
            "organization_id": user["chapter_id"],
        }
        try:
            await db.members.insert_one(member_data)
            inserted += 1
            existing_ids.add(member.unique_member_id)
            existing_mobiles.add(member.primary_mobile)
        except Exception as e:
            errors.append(f"Row {idx+1}: Database error: {str(e)}")
    
    return {
        "message": "Bulk upload completed",
        "total_processed": len(members),
        "successfully_added": inserted,
        "duplicate_mobiles": duplicate_mobiles,
        "skipped": len(skipped),
        "errors": len(errors),
        "skipped_details": skipped if skipped else None,
        "error_details": errors if errors else None
    }

@api_router.get("/admin/members/template")
async def download_member_template(user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Members Template"
    
    # Headers
    headers = ["Member ID", "Full Name", "Primary Mobile", "Secondary Mobile", "Status"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="CF2030", end_color="CF2030", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Add sample data
    ws.append(["01", "Sample Member 1", "9876543210", "9123456789", "Active"])
    ws.append(["02", "Sample Member 2", "9999888877", "", "Active"])
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=members_template.xlsx"}
    )

@api_router.post("/admin/members/upload")
async def upload_members_excel(file: UploadFile = File(...), user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    from openpyxl import load_workbook
    
    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active
    
    members_data = []
    errors = []
    skipped = []
    duplicate_mobiles = []
    
    # Load existing members to avoid duplicate key errors
    existing_members = await db.members.find(
        {"chapter_id": user["chapter_id"]},
        {"unique_member_id": 1, "primary_mobile": 1}
    ).to_list(1000)
    existing_ids = {m.get("unique_member_id") for m in existing_members}
    existing_mobiles = {m.get("primary_mobile") for m in existing_members}
    
    from uuid import uuid4
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0] or not row[1] or not row[2]:  # Skip empty rows
            continue
        
        unique_member_id = str(row[0])
        full_name = str(row[1])
        primary_mobile = str(row[2])
        secondary_mobile = str(row[3]) if row[3] else None
        status = str(row[4]) if row[4] else "Active"
        
        # Check duplicates in DB
        if unique_member_id in existing_ids:
            skipped.append(f"Row {idx}: Member ID {unique_member_id} already exists")
            continue
        if primary_mobile in existing_mobiles:
            skipped.append(f"Row {idx}: Mobile {primary_mobile} already exists")
            duplicate_mobiles.append(primary_mobile)
            continue
        
        # Check duplicates within this upload batch
        batch_ids = {m["unique_member_id"] for m in members_data}
        batch_mobiles = {m["primary_mobile"] for m in members_data}
        if unique_member_id in batch_ids:
            skipped.append(f"Row {idx}: Duplicate Member ID {unique_member_id} in Excel file")
            continue
        if primary_mobile in batch_mobiles:
            skipped.append(f"Row {idx}: Duplicate Mobile {primary_mobile} in Excel file")
            continue
        
        try:
            member_data = {
                "member_id": str(uuid4()),
                "chapter_id": user["chapter_id"],
                "unique_member_id": unique_member_id,
                "full_name": full_name,
                "primary_mobile": primary_mobile,
                "secondary_mobile": secondary_mobile,
                "status": status,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "bni_member_id": unique_member_id,
                "organization_id": user["chapter_id"],
            }
            members_data.append(member_data)
            existing_ids.add(unique_member_id)
            existing_mobiles.add(primary_mobile)
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")
    
    inserted_count = 0
    for idx, member_data in enumerate(members_data):
        try:
            await db.members.insert_one(member_data)
            inserted_count += 1
        except Exception as e:
            errors.append(f"Insert error for row {idx+2}: {str(e)}")
    
    return {
        "message": f"{inserted_count} members uploaded successfully",
        "total": len(members_data),
        "duplicate_mobiles": duplicate_mobiles,
        "skipped": len(skipped),
        "errors": errors,
        "skipped_details": skipped if skipped else None
    }

@api_router.put("/admin/members/{member_id}", response_model=MemberResponse)
async def update_member(member_id: str, member: MemberUpdate, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    chapter_id = user.get("chapter_id")
    query = {"member_id": member_id}
    if chapter_id:
        query["chapter_id"] = chapter_id

    update_data = {k: v for k, v in member.dict(exclude_unset=True).items()}
    result = await db.members.update_one(query, {"$set": update_data})

    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Member not found")

    updated_member = await db.members.find_one({"member_id": member_id}, {"_id": 0})
    return MemberResponse(**updated_member)

@api_router.delete("/admin/members/{member_id}")
async def delete_member(member_id: str, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    result = await db.members.delete_one({"member_id": member_id, "chapter_id": user["chapter_id"]})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Member not found")
    
    return {"message": "Member deleted successfully"}

# ===== ENHANCED MEMBER MANAGEMENT ENDPOINTS =====

# Helper: sync dual status fields
def _sync_status(membership_status: str) -> str:
    """Return legacy status field value based on membership_status."""
    return "Active" if membership_status == "active" else "Inactive"

@api_router.get("/admin/members/stats")
async def get_member_stats(user=Depends(get_current_user)):
    """Return member counts by status for the chapter."""
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    chapter_id = user.get("chapter_id")
    if not chapter_id and user["role"] == "superadmin":
        sa = await db.superadmins.find_one({"mobile": user.get("mobile")}, {"_id": 0})
        sa_id = sa.get("superadmin_id", user.get("mobile")) if sa else user.get("mobile")
        chapters = await db.chapters.find({"created_by": sa_id}, {"chapter_id": 1}).to_list(100)
        chapter_ids = [c["chapter_id"] for c in chapters]
        base_filter = {"chapter_id": {"$in": chapter_ids}}
    elif chapter_id:
        base_filter = {"chapter_id": chapter_id}
    else:
        base_filter = {}

    total = await db.members.count_documents({**base_filter, "archived": {"$ne": True}})
    active = await db.members.count_documents({**base_filter, "membership_status": "active", "archived": {"$ne": True}})
    pending = await db.members.count_documents({**base_filter, "membership_status": "pending"})
    inactive = await db.members.count_documents({**base_filter, "membership_status": "inactive", "archived": {"$ne": True}})
    suspended = await db.members.count_documents({**base_filter, "membership_status": "suspended"})

    # Expiring soon: renewal_date within 30 days
    from datetime import date
    today_str = date.today().isoformat()
    thirty_days = (date.today() + timedelta(days=30)).isoformat()
    expiring_soon = await db.members.count_documents({
        **base_filter,
        "membership_status": "active",
        "renewal_date": {"$ne": None, "$lte": thirty_days, "$gte": today_str}
    })

    return {
        "total": total,
        "active": active,
        "pending": pending,
        "inactive": inactive,
        "suspended": suspended,
        "expiring_soon": expiring_soon
    }

@api_router.get("/admin/members/expiring")
async def get_expiring_members(user=Depends(get_current_user)):
    """Members with renewal_date within next 30 days."""
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    chapter_id = user.get("chapter_id")
    if not chapter_id:
        raise HTTPException(status_code=400, detail="Chapter ID required")

    from datetime import date
    today_str = date.today().isoformat()
    thirty_days = (date.today() + timedelta(days=30)).isoformat()

    members = await db.members.find({
        "chapter_id": chapter_id,
        "membership_status": "active",
        "renewal_date": {"$ne": None, "$lte": thirty_days, "$gte": today_str}
    }, {"_id": 0}).to_list(500)
    return members

@api_router.get("/admin/members/export")
async def export_members(user=Depends(get_current_user)):
    """Export member list as Excel."""
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    chapter_id = user.get("chapter_id")
    if not chapter_id:
        raise HTTPException(status_code=400, detail="Chapter ID required")

    members = await db.members.find(
        {"chapter_id": chapter_id, "archived": {"$ne": True}},
        {"_id": 0}
    ).sort("unique_member_id", 1).to_list(2000)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Members"

    headers = [
        "Member ID", "Full Name", "Primary Mobile", "Secondary Mobile",
        "Email", "Business Name", "Business Category",
        "Joining Date", "Renewal Date", "Induction Fee",
        "Membership Status", "Status"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="CF2030", end_color="CF2030", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for m in members:
        ws.append([
            m.get("unique_member_id", ""),
            m.get("full_name", ""),
            m.get("primary_mobile", ""),
            m.get("secondary_mobile", ""),
            m.get("email", ""),
            m.get("business_name", ""),
            m.get("business_category", ""),
            m.get("joining_date", ""),
            m.get("renewal_date", ""),
            m.get("induction_fee", ""),
            m.get("membership_status", "active"),
            m.get("status", "Active"),
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=members_export.xlsx"}
    )

@api_router.get("/admin/members/{member_id}/profile")
async def get_member_profile(member_id: str, user=Depends(get_current_user)):
    """Full member profile with attendance & payment history."""
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    member = await db.members.find_one({"member_id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Attendance history (last 20)
    attendance = await db.attendance.find(
        {"unique_member_id": member.get("unique_member_id"), "meeting_id": {"$exists": True}},
        {"_id": 0}
    ).sort("timestamp", -1).to_list(20)

    # Enrich attendance with meeting dates
    for att in attendance:
        meeting = await db.meetings.find_one({"meeting_id": att["meeting_id"]}, {"_id": 0, "date": 1})
        att["meeting_date"] = meeting.get("date") if meeting else None

    # Attendance stats
    all_att = await db.attendance.find(
        {"unique_member_id": member.get("unique_member_id"), "type": "member"},
        {"_id": 0, "status": 1, "late_type": 1}
    ).to_list(5000)
    total_meetings_attended = len(all_att)
    present_count = sum(1 for a in all_att if a.get("status") == "present")
    late_count = sum(1 for a in all_att if a.get("late_type") in ("late", "very_late"))

    # Payment history: kitty, meeting fee, misc, events (last 20 each)
    kitty_payments = await db.kitty_payments.find(
        {"member_id": member["member_id"]}, {"_id": 0}
    ).sort("year", -1).to_list(20)

    meetingfee_payments = await db.meetingfee_payments.find(
        {"member_id": member["member_id"]}, {"_id": 0}
    ).sort("year", -1).to_list(20)

    misc_records = await db.misc_payment_records.find(
        {"member_id": member["member_id"]}, {"_id": 0}
    ).to_list(20)

    event_payments = await db.event_payments.find(
        {"member_id": member["member_id"]}, {"_id": 0}
    ).to_list(20)

    # Chapter name
    chapter = await db.chapters.find_one({"chapter_id": member.get("chapter_id")}, {"_id": 0, "name": 1})

    return {
        "member": member,
        "chapter_name": chapter.get("name") if chapter else "",
        "attendance": {
            "recent": attendance,
            "total_attended": total_meetings_attended,
            "present_count": present_count,
            "late_count": late_count,
        },
        "payments": {
            "kitty": kitty_payments,
            "meeting_fee": meetingfee_payments,
            "misc": misc_records,
            "events": event_payments,
        }
    }

@api_router.post("/admin/members/{member_id}/status")
async def change_member_status(member_id: str, data: MemberStatusChange, user=Depends(get_current_user)):
    """Change member status: deactivate, suspend, reactivate. Requires reason."""
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    member = await db.members.find_one({"member_id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    action_map = {
        "deactivate": "inactive",
        "suspend": "suspended",
        "reactivate": "active",
    }
    new_ms = action_map.get(data.action)
    if not new_ms:
        raise HTTPException(status_code=400, detail=f"Invalid action: {data.action}. Must be deactivate, suspend, or reactivate")

    old_ms = member.get("membership_status", "active")

    history_entry = {
        "action": data.action,
        "from_status": old_ms,
        "to_status": new_ms,
        "reason": data.reason,
        "changed_by": user.get("mobile", user.get("email", "system")),
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

    await db.members.update_one(
        {"member_id": member_id},
        {
            "$set": {
                "membership_status": new_ms,
                "status": _sync_status(new_ms),
            },
            "$push": {"status_history": history_entry}
        }
    )

    return {"message": f"Member {data.action}d successfully", "new_status": new_ms}

@api_router.post("/admin/members/bulk-status")
async def bulk_change_status(data: BulkMemberStatus, user=Depends(get_current_user)):
    """Bulk activate/deactivate members."""
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    action_map = {"activate": "active", "deactivate": "inactive"}
    new_ms = action_map.get(data.action)
    if not new_ms:
        raise HTTPException(status_code=400, detail="Action must be activate or deactivate")

    history_entry = {
        "action": data.action,
        "from_status": "bulk",
        "to_status": new_ms,
        "reason": data.reason,
        "changed_by": user.get("mobile", user.get("email", "system")),
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

    result = await db.members.update_many(
        {"member_id": {"$in": data.member_ids}},
        {
            "$set": {
                "membership_status": new_ms,
                "status": _sync_status(new_ms),
            },
            "$push": {"status_history": history_entry}
        }
    )

    return {"message": f"{result.modified_count} members updated", "new_status": new_ms}

@api_router.post("/admin/members/auto-archive")
async def auto_archive_members(user=Depends(get_current_user)):
    """Archive members inactive for 6+ months."""
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    chapter_id = user.get("chapter_id")
    if not chapter_id:
        raise HTTPException(status_code=400, detail="Chapter ID required")

    six_months_ago = (datetime.now(timezone.utc) - timedelta(days=180)).isoformat()

    # Find inactive members whose last status change was 6+ months ago
    inactive_members = await db.members.find({
        "chapter_id": chapter_id,
        "membership_status": "inactive",
        "archived": {"$ne": True},
    }, {"_id": 0, "member_id": 1, "status_history": 1}).to_list(1000)

    archived_ids = []
    for m in inactive_members:
        history = m.get("status_history", [])
        if history:
            last_entry = history[-1]
            if last_entry.get("timestamp", "") < six_months_ago:
                archived_ids.append(m["member_id"])
        else:
            # No history — check created_at or archive anyway
            archived_ids.append(m["member_id"])

    if archived_ids:
        await db.members.update_many(
            {"member_id": {"$in": archived_ids}},
            {"$set": {"archived": True}}
        )

    return {"message": f"{len(archived_ids)} members archived"}

# ===== SUPERADMIN MEMBER APPROVAL ENDPOINTS =====

@api_router.get("/superadmin/members/pending")
async def get_pending_members(user=Depends(get_current_user)):
    """List all pending members across ED's chapters."""
    if user["role"] not in ("superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    if user["role"] == "developer":
        # Developer sees all pending
        members = await db.members.find(
            {"membership_status": "pending"}, {"_id": 0}
        ).sort("created_at", -1).to_list(500)
    else:
        # SuperAdmin sees only their chapters' pending
        sa = await db.superadmins.find_one({"mobile": user.get("mobile")}, {"_id": 0})
        sa_id = sa.get("superadmin_id", user.get("mobile")) if sa else user.get("mobile")
        chapters = await db.chapters.find({"created_by": sa_id}, {"chapter_id": 1}).to_list(100)
        chapter_ids = [c["chapter_id"] for c in chapters]
        members = await db.members.find(
            {"chapter_id": {"$in": chapter_ids}, "membership_status": "pending"},
            {"_id": 0}
        ).sort("created_at", -1).to_list(500)

    # Enrich with chapter name
    for m in members:
        ch = await db.chapters.find_one({"chapter_id": m.get("chapter_id")}, {"_id": 0, "name": 1})
        m["chapter_name"] = ch.get("name") if ch else ""

    return members

@api_router.post("/superadmin/members/{member_id}/approve")
async def approve_member(member_id: str, user=Depends(get_current_user)):
    """Approve a pending member."""
    if user["role"] not in ("superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    member = await db.members.find_one({"member_id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    if member.get("membership_status") != "pending":
        raise HTTPException(status_code=400, detail="Member is not in pending status")

    history_entry = {
        "action": "approved",
        "from_status": "pending",
        "to_status": "active",
        "reason": "Approved by ED",
        "changed_by": user.get("mobile", user.get("email", "system")),
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

    await db.members.update_one(
        {"member_id": member_id},
        {
            "$set": {
                "membership_status": "active",
                "status": "Active",
            },
            "$push": {"status_history": history_entry}
        }
    )

    return {"message": "Member approved successfully"}

@api_router.post("/superadmin/members/{member_id}/reject")
async def reject_member(member_id: str, data: MemberApprovalAction, user=Depends(get_current_user)):
    """Reject a pending member."""
    if user["role"] not in ("superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    member = await db.members.find_one({"member_id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    if member.get("membership_status") != "pending":
        raise HTTPException(status_code=400, detail="Member is not in pending status")

    if not data.reason:
        raise HTTPException(status_code=400, detail="Reason is required for rejection")

    history_entry = {
        "action": "rejected",
        "from_status": "pending",
        "to_status": "rejected",
        "reason": data.reason,
        "changed_by": user.get("mobile", user.get("email", "system")),
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

    await db.members.update_one(
        {"member_id": member_id},
        {
            "$set": {
                "membership_status": "rejected",
                "status": "Inactive",
            },
            "$push": {"status_history": history_entry}
        }
    )

    return {"message": "Member rejected"}

@api_router.post("/superadmin/members/{member_id}/transfer")
async def transfer_member(member_id: str, data: MemberTransfer, user=Depends(get_current_user)):
    """Transfer member to a different chapter."""
    if user["role"] not in ("superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    member = await db.members.find_one({"member_id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Verify target chapter exists
    target_ch = await db.chapters.find_one({"chapter_id": data.target_chapter_id}, {"_id": 0})
    if not target_ch:
        raise HTTPException(status_code=404, detail="Target chapter not found")

    old_chapter = member.get("chapter_id")

    history_entry = {
        "action": "transferred",
        "from_status": member.get("membership_status", "active"),
        "to_status": member.get("membership_status", "active"),
        "reason": data.reason or f"Transferred from chapter {old_chapter} to {data.target_chapter_id}",
        "changed_by": user.get("mobile", user.get("email", "system")),
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

    await db.members.update_one(
        {"member_id": member_id},
        {
            "$set": {
                "chapter_id": data.target_chapter_id,
                "organization_id": data.target_chapter_id,
                "transfer_from_chapter": old_chapter,
                "transfer_date": datetime.now(timezone.utc).isoformat(),
            },
            "$push": {"status_history": history_entry}
        }
    )

    return {"message": f"Member transferred to {target_ch.get('name', data.target_chapter_id)}"}

@api_router.get("/superadmin/members/all")
async def get_all_ed_members(
    search: str = "",
    status_filter: str = "",
    chapter_filter: str = "",
    user=Depends(get_current_user)
):
    """Cross-chapter member view for SuperAdmin/Developer."""
    if user["role"] not in ("superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    if user["role"] == "developer":
        base_filter = {}
    else:
        sa = await db.superadmins.find_one({"mobile": user.get("mobile")}, {"_id": 0})
        sa_id = sa.get("superadmin_id", user.get("mobile")) if sa else user.get("mobile")
        chapters = await db.chapters.find({"created_by": sa_id}, {"chapter_id": 1}).to_list(100)
        chapter_ids = [c["chapter_id"] for c in chapters]
        base_filter = {"chapter_id": {"$in": chapter_ids}}

    query = {**base_filter, "archived": {"$ne": True}}

    if status_filter:
        query["membership_status"] = status_filter
    if chapter_filter:
        query["chapter_id"] = chapter_filter
    if search:
        query["$or"] = [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"primary_mobile": {"$regex": search, "$options": "i"}},
            {"business_name": {"$regex": search, "$options": "i"}},
            {"unique_member_id": {"$regex": search, "$options": "i"}},
        ]

    members = await db.members.find(query, {"_id": 0}).sort("full_name", 1).to_list(2000)

    # Enrich with chapter names
    chapter_cache = {}
    for m in members:
        cid = m.get("chapter_id")
        if cid not in chapter_cache:
            ch = await db.chapters.find_one({"chapter_id": cid}, {"_id": 0, "name": 1})
            chapter_cache[cid] = ch.get("name") if ch else ""
        m["chapter_name"] = chapter_cache[cid]

    return members

@api_router.post("/admin/meetings", response_model=MeetingResponse)
async def create_meeting(meeting: MeetingCreate, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    meeting_id = f"MT{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"
    
    # Generate single QR token (type will be selected manually in form)
    qr_token = generate_qr_token(meeting_id, user["chapter_id"], "all")
    
    meeting_data = {
        "meeting_id": meeting_id,
        "chapter_id": user["chapter_id"],
        "date": meeting.date,
        "start_time": meeting.start_time,
        "late_cutoff_time": meeting.late_cutoff_time,
        "end_time": meeting.end_time,
        "qr_token": qr_token,
        "qr_expires_at": (datetime.now(timezone.utc) + timedelta(seconds=10)).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.meetings.insert_one(meeting_data)
    return MeetingResponse(**meeting_data)

@api_router.get("/admin/meetings", response_model=List[MeetingResponse])
async def get_meetings(user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    meetings = await db.meetings.find({"chapter_id": user["chapter_id"]}, {"_id": 0}).to_list(1000)
    return meetings

@api_router.delete("/admin/meetings/{meeting_id}")
async def delete_meeting(meeting_id: str, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    result = await db.meetings.delete_one({"meeting_id": meeting_id, "chapter_id": user["chapter_id"]})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Meeting not found")
    
    # Also delete associated attendance records
    await db.attendance.delete_many({"meeting_id": meeting_id})
    
    return {"message": "Meeting deleted successfully"}

@api_router.get("/admin/meetings/{meeting_id}/qr")
async def get_qr_code(meeting_id: str, request: Request, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    meeting = await db.meetings.find_one({"meeting_id": meeting_id, "chapter_id": user["chapter_id"]}, {"_id": 0})
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")
    
    # Check if QR expired, regenerate if needed
    qr_expires = datetime.fromisoformat(meeting["qr_expires_at"])
    if datetime.now(timezone.utc) >= qr_expires:
        new_token = generate_qr_token(meeting_id, user["chapter_id"], "all")
        new_expiry = (datetime.now(timezone.utc) + timedelta(seconds=10)).isoformat()
        await db.meetings.update_one(
            {"meeting_id": meeting_id},
            {"$set": {"qr_token": new_token, "qr_expires_at": new_expiry}}
        )
        meeting["qr_token"] = new_token
    
    # Always use FRONTEND_URL for QR code generation
    # This ensures QR codes always point to the custom domain (e.g., aasaanapp.com)
    # regardless of which domain the admin panel is accessed from
    qr_image = create_qr_image(meeting["qr_token"], request_host=None)
    return StreamingResponse(io.BytesIO(qr_image), media_type="image/png")

@api_router.get("/admin/meetings/{meeting_id}/attendance", response_model=List[AttendanceResponse])
async def get_attendance(meeting_id: str, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    # Only show approved attendance
    attendance = await db.attendance.find({
        "meeting_id": meeting_id,
        "approval_status": "approved"
    }, {"_id": 0}).to_list(1000)
    return attendance

@api_router.get("/admin/meetings/{meeting_id}/summary")
async def get_meeting_summary(meeting_id: str, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    # Get meeting details
    meeting = await db.meetings.find_one({"meeting_id": meeting_id, "chapter_id": user["chapter_id"]}, {"_id": 0})
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")
    
    # Get all ACTIVE members for this chapter
    members = await db.members.find({"chapter_id": user["chapter_id"], "status": "Active"}, {"_id": 0}).to_list(1000)
    total_members = len(members)

    # Get all attendance for this meeting
    attendance = await db.attendance.find({"meeting_id": meeting_id}, {"_id": 0}).to_list(1000)
    
    # Count members and substitutes
    member_attendance = [a for a in attendance if a["type"] == "member"]
    substitute_attendance = [a for a in attendance if a["type"] == "substitute"]
    visitor_attendance = [a for a in attendance if a["type"] == "visitor"]
    
    present_count = len(member_attendance)
    substitute_count = len(substitute_attendance)
    visitor_count = len(visitor_attendance)
    
    # Check if meeting has ended
    end_time = datetime.fromisoformat(meeting["end_time"])
    if end_time.tzinfo is None:
        end_time = end_time.replace(tzinfo=timezone.utc).astimezone(IST)
    else:
        end_time = end_time.astimezone(IST)
    now_ist = datetime.now(IST)
    meeting_ended = now_ist > end_time
    
    # Build detailed lists for Quick View
    # 1. All members list
    all_members = []
    for member in members:
        all_members.append({
            "unique_member_id": member["unique_member_id"],
            "full_name": member["full_name"],
            "primary_mobile": member.get("primary_mobile", ""),
            "status": "Active"
        })
    
    # 2. Present members list (with timestamp)
    present_members = []
    for att in member_attendance:
        member_info = next((m for m in members if m["unique_member_id"] == att.get("unique_member_id")), None)
        present_members.append({
            "unique_member_id": att.get("unique_member_id"),
            "full_name": member_info["full_name"] if member_info else att.get("member_name", "Unknown"),
            "timestamp": att.get("timestamp"),
            "status": "Present"
        })
    
    # 3. Substitute members list (with substitute name)
    substitute_members = []
    for att in substitute_attendance:
        member_info = next((m for m in members if m["unique_member_id"] == att.get("unique_member_id")), None)
        substitute_members.append({
            "unique_member_id": att.get("unique_member_id"),
            "full_name": member_info["full_name"] if member_info else att.get("member_name", "Unknown"),
            "substitute_name": att.get("substitute_name", ""),
            "substitute_mobile": att.get("substitute_mobile", ""),
            "timestamp": att.get("timestamp"),
            "status": "Substitute"
        })
    
    # 4. Visitors list
    visitors = []
    for att in visitor_attendance:
        visitors.append({
            "visitor_name": att.get("visitor_name", ""),
            "company": att.get("company_name", att.get("visitor_company", "")),
            "mobile": att.get("visitor_mobile", ""),
            "invited_by": att.get("invited_by", ""),
            "timestamp": att.get("timestamp"),
            "status": "Visitor"
        })
    
    # 5. Pending/Absent members
    attended_member_ids = {a.get("unique_member_id") for a in attendance if a["type"] in ["member", "substitute"] and a.get("unique_member_id")}
    pending_members = []
    
    for member in members:
        if member["unique_member_id"] not in attended_member_ids:
            pending_members.append({
                "unique_member_id": member["unique_member_id"],
                "full_name": member["full_name"],
                "primary_mobile": member.get("primary_mobile", ""),
                "status": "Absent" if meeting_ended else "Pending"
            })
    
    absent_count = len(pending_members) if meeting_ended else 0
    pending_count = len(pending_members) if not meeting_ended else 0
    
    return {
        "total_members": total_members,
        "present_count": present_count,
        "substitute_count": substitute_count,
        "visitor_count": visitor_count,
        "absent_count": absent_count,
        "pending_count": pending_count,
        "meeting_ended": meeting_ended,
        "pending_members": pending_members,
        # Detailed lists for Quick View
        "all_members": all_members,
        "present_members": present_members,
        "substitute_members": substitute_members,
        "visitors": visitors
    }

@api_router.get("/admin/attendance/pending", response_model=List[AttendanceResponse])
async def get_pending_attendance(user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    # Get all pending attendance for this chapter
    pending = await db.attendance.find({
        "approval_status": "pending"
    }, {"_id": 0}).to_list(1000)
    
    # Filter by chapter - get meeting IDs for this chapter
    meetings = await db.meetings.find({"chapter_id": user["chapter_id"]}, {"_id": 0, "meeting_id": 1}).to_list(1000)
    meeting_ids = [m["meeting_id"] for m in meetings]
    
    # Filter pending attendance by chapter meetings
    chapter_pending = [att for att in pending if att["meeting_id"] in meeting_ids]
    
    return chapter_pending

@api_router.post("/admin/attendance/{attendance_id}/approve")
async def approve_attendance(attendance_id: str, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    result = await db.attendance.update_one(
        {"attendance_id": attendance_id},
        {"$set": {"approval_status": "approved"}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Attendance not found")
    
    return {"message": "Attendance approved successfully"}

@api_router.post("/admin/attendance/{attendance_id}/reject")
async def reject_attendance(attendance_id: str, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    result = await db.attendance.update_one(
        {"attendance_id": attendance_id},
        {"$set": {"approval_status": "rejected"}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Attendance not found")
    
    return {"message": "Attendance rejected successfully"}

@api_router.get("/admin/meetings/{meeting_id}/report/excel")
async def download_excel_report(meeting_id: str, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    meeting = await db.meetings.find_one({"meeting_id": meeting_id}, {"_id": 0})
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")
    
    chapter = await db.chapters.find_one({"chapter_id": user["chapter_id"]}, {"_id": 0, "name": 1})
    chapter_name = chapter.get("name", "") if chapter else ""
    
    members = await db.members.find({"chapter_id": user["chapter_id"]}, {"_id": 0}).to_list(1000)
    attendance = await db.attendance.find({"meeting_id": meeting_id}, {"_id": 0}).to_list(1000)
    
    excel_file = generate_excel_report(meeting, members, attendance, chapter_name)
    
    return StreamingResponse(
        io.BytesIO(excel_file),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=attendance_{meeting_id}.xlsx"}
    )

@api_router.get("/admin/meetings/{meeting_id}/report/pdf")
async def download_pdf_report(meeting_id: str, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    
    meeting = await db.meetings.find_one({"meeting_id": meeting_id}, {"_id": 0})
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")
    
    chapter = await db.chapters.find_one({"chapter_id": user["chapter_id"]}, {"_id": 0, "name": 1})
    chapter_name = chapter.get("name", "") if chapter else ""
    
    members = await db.members.find({"chapter_id": user["chapter_id"]}, {"_id": 0}).to_list(1000)
    attendance = await db.attendance.find({"meeting_id": meeting_id}, {"_id": 0}).to_list(1000)
    
    pdf_file = generate_pdf_report(meeting, members, attendance, chapter_name)
    
    return StreamingResponse(
        io.BytesIO(pdf_file),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=attendance_{meeting_id}.pdf"}
    )

# ===== PUBLIC ENDPOINTS =====
@api_router.get("/qr/verify/{token}")
async def verify_qr(token: str):
    payload = verify_qr_token(token)
    if not payload:
        raise HTTPException(status_code=400, detail="Invalid or expired QR code")
    
    meeting = await db.meetings.find_one({"meeting_id": payload["meeting_id"]}, {"_id": 0})
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")
    
    # Check if meeting time window is valid
    now_ist = datetime.now(IST)
    start_time = datetime.fromisoformat(meeting["start_time"]).replace(tzinfo=timezone.utc).astimezone(IST)
    end_time = datetime.fromisoformat(meeting["end_time"]).replace(tzinfo=timezone.utc).astimezone(IST)
    
    # Meeting hasn't started yet
    if now_ist < start_time:
        raise HTTPException(
            status_code=400,
            detail=f"Meeting hasn't started yet. Opens at {start_time.strftime('%I:%M %p on %d %B %Y')}"
        )
    
    # Meeting has ended
    if now_ist > end_time:
        raise HTTPException(status_code=400, detail="Meeting has ended. QR code expired.")
    
    return {
        "meeting_id": meeting["meeting_id"],
        "chapter_id": meeting["chapter_id"],
        "date": meeting["date"],
        "start_time": meeting["start_time"],
        "end_time": meeting["end_time"],
        "attendance_type": payload.get("attendance_type", "member")  # Return type from token
    }

@api_router.get("/members/{chapter_id}")
async def get_chapter_members(chapter_id: str):
    """Public endpoint to get members list for attendance form - sorted alphabetically by name (A-Z)"""
    members = await db.members.find(
        {"chapter_id": chapter_id, "status": "Active"}, 
        {"_id": 0, "member_id": 1, "unique_member_id": 1, "full_name": 1, "primary_mobile": 1}
    ).to_list(1000)
    
    # Sort alphabetically by full_name (case-insensitive A-Z)
    members.sort(key=lambda x: (x.get("full_name") or "").lower())
    return members

@api_router.post("/attendance/mark", response_model=AttendanceResponse)
async def mark_attendance(attendance: AttendanceCreate):
    meeting = await db.meetings.find_one({"meeting_id": attendance.meeting_id}, {"_id": 0})
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")
    
    # Use IST timezone
    now_ist = datetime.now(IST)
    
    meeting_date = datetime.fromisoformat(meeting["date"]).replace(tzinfo=timezone.utc).astimezone(IST)
    start_time = datetime.fromisoformat(meeting["start_time"]).replace(tzinfo=timezone.utc).astimezone(IST)
    late_cutoff = datetime.fromisoformat(meeting["late_cutoff_time"]).replace(tzinfo=timezone.utc).astimezone(IST)
    end_time = datetime.fromisoformat(meeting["end_time"]).replace(tzinfo=timezone.utc).astimezone(IST)
    
    # Check if attendance window is valid - must be between start and end time
    if now_ist < start_time:
        raise HTTPException(
            status_code=400, 
            detail=f"Attendance window not yet open. Meeting starts at {start_time.strftime('%I:%M %p on %d %B %Y')}"
        )
    
    if now_ist > end_time:
        raise HTTPException(status_code=400, detail="Attendance window closed")
    
    # Member validation and duplicate check
    member = None
    approval_status = "approved"  # All attendance auto-approved now
    
    if attendance.type == "member":
        member = await db.members.find_one({"unique_member_id": attendance.unique_member_id}, {"_id": 0})
        if not member:
            raise HTTPException(status_code=404, detail="Member not found")
        
        # Check if member already marked attendance (device_fingerprint removed from duplicate check to fix iPhone false positives)
        existing = await db.attendance.find_one({
            "meeting_id": attendance.meeting_id,
            "$or": [
                {"primary_mobile": attendance.primary_mobile},
                {"unique_member_id": attendance.unique_member_id}
            ]
        })
        if existing:
            raise HTTPException(status_code=400, detail="Attendance already marked for this member")
    
    # For substitute - check if member already attended
    if attendance.type == "substitute":
        member = await db.members.find_one({"unique_member_id": attendance.unique_member_id}, {"_id": 0})
        if not member:
            raise HTTPException(status_code=404, detail="Member ID not found")
        
        # Check if member already marked attendance
        existing = await db.attendance.find_one({
            "meeting_id": attendance.meeting_id,
            "unique_member_id": attendance.unique_member_id
        })
        if existing:
            raise HTTPException(status_code=400, detail=f"Member {attendance.unique_member_id} has already marked attendance")
        
        # Also check if member's primary mobile was used
        existing_mobile = await db.attendance.find_one({
            "meeting_id": attendance.meeting_id,
            "primary_mobile": member["primary_mobile"]
        })
        if existing_mobile:
            raise HTTPException(status_code=400, detail=f"Member {attendance.unique_member_id}'s mobile number already used for attendance")
    
    # Determine status and late type based on cutoff time
    status = "Present"
    late_type = None
    
    if now_ist <= late_cutoff:
        late_type = "On time"
    else:
        late_type = "Late"
    
    # Get invited by member name for visitors - CRITICAL: Filter by chapter_id to avoid cross-chapter data
    invited_by_member_name = None
    if attendance.type == "visitor" and attendance.invited_by_member_id:
        invited_member = await db.members.find_one(
            {
                "unique_member_id": attendance.invited_by_member_id,
                "chapter_id": chapter_id  # CRITICAL: Must filter by chapter_id
            }, 
            {"_id": 0, "full_name": 1}
        )
        if invited_member:
            invited_by_member_name = invited_member["full_name"]
    
    attendance_data = {
        "attendance_id": f"A{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S%f')}",
        "meeting_id": attendance.meeting_id,
        "unique_member_id": attendance.unique_member_id if attendance.type in ["member", "substitute"] else None,
        "type": attendance.type,
        "status": status,
        "timestamp": now_ist.isoformat(),
        "late_type": late_type,
        "member_name": member.get("full_name") if member else None,
        "primary_mobile": attendance.primary_mobile if attendance.type == "member" else None,
        "substitute_name": attendance.substitute_name,
        "substitute_mobile": attendance.substitute_mobile,
        "visitor_name": attendance.visitor_name,
        "visitor_mobile": attendance.visitor_mobile,
        "visitor_company": attendance.visitor_company,
        "invited_by_member_id": attendance.invited_by_member_id,
        "invited_by_member_name": invited_by_member_name,
        "device_fingerprint": attendance.device_fingerprint,
        "ip_address": attendance.ip_address,
        "approval_status": approval_status
    }
    
    await db.attendance.insert_one(attendance_data)
    return AttendanceResponse(**attendance_data)

# ===== FUND MANAGEMENT ENDPOINTS =====

# Kitty Settings
@api_router.get("/admin/fund/kitty/settings")
async def get_kitty_settings(user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    settings = await db.kitty_settings.find({"chapter_id": chapter_id}, {"_id": 0}).to_list(100)
    return settings

@api_router.post("/admin/fund/kitty/settings")
async def set_kitty_amount(data: KittySettingCreate, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    # Check if setting already exists for this month/year
    existing = await db.kitty_settings.find_one({
        "chapter_id": chapter_id,
        "month": data.month,
        "year": data.year
    })
    
    if existing:
        # Update existing
        await db.kitty_settings.update_one(
            {"chapter_id": chapter_id, "month": data.month, "year": data.year},
            {"$set": {"amount": data.amount}}
        )
        return {"message": "Kitty amount updated", "setting_id": existing["setting_id"]}
    
    # Create new setting
    setting_id = f"KS{datetime.now().strftime('%Y%m%d%H%M%S')}"
    setting_data = {
        "setting_id": setting_id,
        "chapter_id": chapter_id,
        "month": data.month,
        "year": data.year,
        "amount": data.amount,
        "created_at": datetime.now(IST).isoformat()
    }
    await db.kitty_settings.insert_one(setting_data)
    return {"message": "Kitty amount set", "setting_id": setting_id}

# Kitty Payments
@api_router.get("/admin/fund/kitty/payments")
async def get_kitty_payments(month: int = None, year: int = None, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    # Get all active members
    members = await db.members.find(
        {"chapter_id": chapter_id, "status": "Active"},
        {"_id": 0, "member_id": 1, "unique_member_id": 1, "full_name": 1}
    ).to_list(500)
    
    # Get kitty setting for the month (default/bulk amount)
    query = {"chapter_id": chapter_id}
    if month and year:
        query["month"] = month
        query["year"] = year
    
    setting = await db.kitty_settings.find_one(query, {"_id": 0})
    default_amount = setting["amount"] if setting else 0
    
    # Get individual member amounts (custom amounts)
    member_amounts = await db.member_amounts.find({
        "chapter_id": chapter_id,
        "month": month,
        "year": year,
        "type": "kitty"
    }, {"_id": 0}).to_list(500)
    member_amount_map = {ma["member_id"]: ma["amount"] for ma in member_amounts}
    
    # Get payments for the month
    payment_query = {"chapter_id": chapter_id}
    if month and year:
        payment_query["month"] = month
        payment_query["year"] = year
    
    payments = await db.kitty_payments.find(payment_query, {"_id": 0}).to_list(500)
    payment_map = {p["member_id"]: p for p in payments}
    
    # Build response with all members
    result = []
    for member in members:
        payment = payment_map.get(member["member_id"])
        # Priority: payment amount > individual amount > default amount
        if payment:
            amount = payment.get("amount", default_amount)
        else:
            amount = member_amount_map.get(member["member_id"], default_amount)
        
        result.append({
            "member_id": member["member_id"],
            "unique_member_id": member["unique_member_id"],
            "member_name": member["full_name"],
            "amount": amount,
            "status": payment["status"] if payment else "pending",
            "paid_date": payment.get("paid_date") if payment else None,
            "payment_id": payment.get("payment_id") if payment else None
        })
    
    # Sort: pending first, then by name
    result.sort(key=lambda x: (0 if x["status"] == "pending" else 1, x["member_name"].lower()))
    return result

@api_router.get("/admin/fund/kitty/payments/all")
async def get_all_kitty_payments(user = Depends(get_current_user)):
    """Get all kitty payments with month/year info for reports"""
    chapter_id = user.get("chapter_id")
    payments = await db.kitty_payments.find(
        {"chapter_id": chapter_id, "status": "paid"},
        {"_id": 0}
    ).to_list(5000)
    return payments

@api_router.post("/admin/fund/kitty/payments/mark")
async def mark_kitty_payment(data: KittyPaymentCreate, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    # Get kitty amount for the month
    setting = await db.kitty_settings.find_one({
        "chapter_id": chapter_id,
        "month": data.month,
        "year": data.year
    })
    
    if not setting:
        raise HTTPException(status_code=400, detail="Kitty amount not set for this month")
    
    # Check if already paid
    existing = await db.kitty_payments.find_one({
        "chapter_id": chapter_id,
        "member_id": data.member_id,
        "month": data.month,
        "year": data.year
    })
    
    if existing and existing.get("status") == "paid":
        raise HTTPException(status_code=400, detail="Already paid")
    
    payment_id = f"KP{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    payment_data = {
        "payment_id": payment_id,
        "chapter_id": chapter_id,
        "member_id": data.member_id,
        "month": data.month,
        "year": data.year,
        "amount": setting["amount"],
        "status": "paid",
        "paid_date": datetime.now(IST).isoformat(),
        "received_by": user.get("mobile")
    }
    
    if existing:
        await db.kitty_payments.update_one(
            {"_id": existing["_id"]},
            {"$set": payment_data}
        )
    else:
        await db.kitty_payments.insert_one(payment_data)
    
    return {"message": "Payment marked", "payment_id": payment_id}

# Kitty Unmark Payment
@api_router.post("/admin/fund/kitty/payments/unmark")
async def unmark_kitty_payment(data: KittyPaymentCreate, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    result = await db.kitty_payments.delete_one({
        "chapter_id": chapter_id,
        "member_id": data.member_id,
        "month": data.month,
        "year": data.year
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    return {"message": "Payment unmarked"}

# Kitty Bulk Mark
@api_router.post("/admin/fund/kitty/payments/bulk-mark")
async def bulk_mark_kitty(data: BulkMarkPayment, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    setting = await db.kitty_settings.find_one({
        "chapter_id": chapter_id,
        "month": data.month,
        "year": data.year
    })
    
    if not setting:
        raise HTTPException(status_code=400, detail="Kitty amount not set for this month")
    
    marked_count = 0
    for member_id in data.member_ids:
        existing = await db.kitty_payments.find_one({
            "chapter_id": chapter_id,
            "member_id": member_id,
            "month": data.month,
            "year": data.year,
            "status": "paid"
        })
        
        if not existing:
            payment_data = {
                "payment_id": f"KP{datetime.now().strftime('%Y%m%d%H%M%S%f')}",
                "chapter_id": chapter_id,
                "member_id": member_id,
                "month": data.month,
                "year": data.year,
                "amount": setting["amount"],
                "status": "paid",
                "paid_date": datetime.now(IST).isoformat(),
                "received_by": user.get("mobile")
            }
            await db.kitty_payments.insert_one(payment_data)
            marked_count += 1
    
    return {"message": f"{marked_count} payments marked"}

# Kitty Update Payment Amount
@api_router.put("/admin/fund/kitty/payments/{payment_id}")
async def update_kitty_payment_amount(payment_id: str, data: dict, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    new_amount = data.get("amount")
    
    if not new_amount:
        raise HTTPException(status_code=400, detail="Amount required")
    
    result = await db.kitty_payments.update_one(
        {"payment_id": payment_id, "chapter_id": chapter_id},
        {"$set": {"amount": float(new_amount)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    return {"message": "Amount updated"}

# Update individual member amount (for both pending and paid)
@api_router.put("/admin/fund/kitty/member-amount")
async def update_kitty_member_amount(data: dict, user = Depends(get_current_user)):
    """Update individual member's kitty amount for a specific month"""
    chapter_id = user.get("chapter_id")
    member_id = data.get("member_id")
    month = data.get("month")
    year = data.get("year")
    new_amount = data.get("amount")
    
    if not all([member_id, month, year, new_amount]):
        raise HTTPException(status_code=400, detail="member_id, month, year, and amount required")
    
    # Store/Update individual member amount
    await db.member_amounts.update_one(
        {
            "chapter_id": chapter_id,
            "member_id": member_id,
            "month": int(month),
            "year": int(year),
            "type": "kitty"
        },
        {
            "$set": {
                "amount": float(new_amount),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        },
        upsert=True
    )
    
    # Also update if payment already exists
    await db.kitty_payments.update_one(
        {
            "chapter_id": chapter_id,
            "member_id": member_id,
            "month": int(month),
            "year": int(year)
        },
        {"$set": {"amount": float(new_amount)}}
    )
    
    return {"message": "Amount updated"}

# Kitty Bulk Unmark
@api_router.post("/admin/fund/kitty/payments/bulk-unmark")
async def bulk_unmark_kitty(data: BulkUnmarkPayment, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    result = await db.kitty_payments.delete_many({
        "chapter_id": chapter_id,
        "member_id": {"$in": data.member_ids},
        "month": data.month,
        "year": data.year
    })
    
    return {"message": f"{result.deleted_count} payments unmarked"}

# ==================== MEETING FEES (Monthly like Kitty) ====================

# Meeting Fee Settings
@api_router.get("/admin/fund/meetingfee/settings")
async def get_meetingfee_settings(user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    settings = await db.meetingfee_settings.find({"chapter_id": chapter_id}, {"_id": 0}).to_list(100)
    return settings

@api_router.post("/admin/fund/meetingfee/settings")
async def set_meetingfee_amount(data: dict, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    month = data.get("month")
    year = data.get("year")
    amount = data.get("amount")
    
    existing = await db.meetingfee_settings.find_one({
        "chapter_id": chapter_id, "month": month, "year": year
    })
    
    if existing:
        await db.meetingfee_settings.update_one(
            {"chapter_id": chapter_id, "month": month, "year": year},
            {"$set": {"amount": amount}}
        )
        return {"message": "Meeting fee amount updated", "setting_id": existing["setting_id"]}
    
    setting_id = f"MFS{datetime.now().strftime('%Y%m%d%H%M%S')}"
    setting_data = {
        "setting_id": setting_id,
        "chapter_id": chapter_id,
        "month": month,
        "year": year,
        "amount": amount,
        "created_at": datetime.now(IST).isoformat()
    }
    await db.meetingfee_settings.insert_one(setting_data)
    return {"message": "Meeting fee amount set", "setting_id": setting_id}

# Meeting Fee Payments List
@api_router.get("/admin/fund/meetingfee/payments")
async def get_meetingfee_payments(month: int = None, year: int = None, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    members = await db.members.find(
        {"chapter_id": chapter_id, "status": "Active"},
        {"_id": 0, "member_id": 1, "unique_member_id": 1, "full_name": 1}
    ).to_list(500)
    
    query = {"chapter_id": chapter_id}
    if month: query["month"] = month
    if year: query["year"] = year
    
    setting = await db.meetingfee_settings.find_one(query, {"_id": 0})
    default_amount = setting["amount"] if setting else 0
    
    # Get individual member amounts (custom amounts)
    member_amounts = await db.member_amounts.find({
        "chapter_id": chapter_id,
        "month": month,
        "year": year,
        "type": "meetingfee"
    }, {"_id": 0}).to_list(500)
    member_amount_map = {ma["member_id"]: ma["amount"] for ma in member_amounts}
    
    result = []
    for member in members:
        payment = await db.meetingfee_payments.find_one({
            "chapter_id": chapter_id,
            "member_id": member["member_id"],
            "month": month,
            "year": year
        }, {"_id": 0})
        
        # Priority: payment amount > individual amount > default amount
        if payment:
            amount = payment.get("amount", default_amount)
        else:
            amount = member_amount_map.get(member["member_id"], default_amount)
        
        result.append({
            "member_id": member["member_id"],
            "member_name": member.get("full_name", ""),
            "unique_id": member.get("unique_member_id", ""),
            "month": month,
            "year": year,
            "amount": amount,
            "status": "paid" if payment else "pending",
            "paid_date": payment.get("paid_date") if payment else None,
            "payment_id": payment.get("payment_id") if payment else None,
            "payment_mode": payment.get("payment_mode") if payment else None
        })
    
    # Sort: pending first, then by name
    result.sort(key=lambda x: (0 if x["status"] == "pending" else 1, x["member_name"].lower()))
    return result

# Update individual member amount (for both pending and paid)
@api_router.put("/admin/fund/meetingfee/member-amount")
async def update_meetingfee_member_amount(data: dict, user = Depends(get_current_user)):
    """Update individual member's meeting fee amount for a specific month"""
    chapter_id = user.get("chapter_id")
    member_id = data.get("member_id")
    month = data.get("month")
    year = data.get("year")
    new_amount = data.get("amount")
    
    if not all([member_id, month, year, new_amount]):
        raise HTTPException(status_code=400, detail="member_id, month, year, and amount required")
    
    # Store/Update individual member amount
    await db.member_amounts.update_one(
        {
            "chapter_id": chapter_id,
            "member_id": member_id,
            "month": int(month),
            "year": int(year),
            "type": "meetingfee"
        },
        {
            "$set": {
                "amount": float(new_amount),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        },
        upsert=True
    )
    
    # Also update if payment already exists
    await db.meetingfee_payments.update_one(
        {
            "chapter_id": chapter_id,
            "member_id": member_id,
            "month": int(month),
            "year": int(year)
        },
        {"$set": {"amount": float(new_amount)}}
    )
    
    return {"message": "Amount updated"}

@api_router.get("/admin/fund/meetingfee/payments/all")
async def get_all_meetingfee_payments(user = Depends(get_current_user)):
    """Get all meeting fee payments with month/year info for reports"""
    chapter_id = user.get("chapter_id")
    payments = await db.meetingfee_payments.find(
        {"chapter_id": chapter_id, "status": "paid"},
        {"_id": 0}
    ).to_list(5000)
    return payments

# Meeting Fee Mark Payment
@api_router.post("/admin/fund/meetingfee/payments/mark")
async def mark_meetingfee_payment(data: dict, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    member_id = data.get("member_id")
    month = data.get("month")
    year = data.get("year")
    custom_amount = data.get("amount")
    payment_mode = data.get("payment_mode", "Cash")  # Point 5: Payment mode
    
    existing = await db.meetingfee_payments.find_one({
        "chapter_id": chapter_id, "member_id": member_id, "month": month, "year": year
    })
    
    if existing:
        raise HTTPException(status_code=400, detail="Payment already marked")
    
    setting = await db.meetingfee_settings.find_one({"chapter_id": chapter_id, "month": month, "year": year})
    amount = custom_amount if custom_amount else (setting["amount"] if setting else 0)
    
    payment_id = f"MFP{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    payment_data = {
        "payment_id": payment_id,
        "chapter_id": chapter_id,
        "member_id": member_id,
        "month": month,
        "year": year,
        "amount": amount,
        "status": "paid",
        "payment_mode": payment_mode,  # Point 5: Payment mode
        "paid_date": datetime.now(IST).isoformat(),
        "received_by": user.get("mobile")
    }
    await db.meetingfee_payments.insert_one(payment_data)
    return {"message": "Payment marked", "payment_id": payment_id}

# Meeting Fee Unmark
@api_router.post("/admin/fund/meetingfee/payments/unmark")
async def unmark_meetingfee_payment(data: dict, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    result = await db.meetingfee_payments.delete_one({
        "chapter_id": chapter_id,
        "member_id": data.get("member_id"),
        "month": data.get("month"),
        "year": data.get("year")
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    return {"message": "Payment unmarked"}

# Meeting Fee Update Amount
@api_router.put("/admin/fund/meetingfee/payments/{payment_id}")
async def update_meetingfee_amount(payment_id: str, data: dict, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    new_amount = data.get("amount")
    
    result = await db.meetingfee_payments.update_one(
        {"payment_id": payment_id, "chapter_id": chapter_id},
        {"$set": {"amount": new_amount}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    return {"message": "Amount updated"}

# Meeting Fee Bulk Mark
@api_router.post("/admin/fund/meetingfee/payments/bulk-mark")
async def bulk_mark_meetingfee(data: dict, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    member_ids = data.get("member_ids", [])
    month = data.get("month")
    year = data.get("year")
    
    setting = await db.meetingfee_settings.find_one({"chapter_id": chapter_id, "month": month, "year": year})
    amount = setting["amount"] if setting else 0
    
    marked_count = 0
    for member_id in member_ids:
        existing = await db.meetingfee_payments.find_one({
            "chapter_id": chapter_id, "member_id": member_id, "month": month, "year": year
        })
        if not existing:
            payment_id = f"MFP{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
            await db.meetingfee_payments.insert_one({
                "payment_id": payment_id,
                "chapter_id": chapter_id,
                "member_id": member_id,
                "month": month,
                "year": year,
                "amount": amount,
                "status": "paid",
                "paid_date": datetime.now(IST).isoformat(),
                "received_by": user.get("mobile")
            })
            marked_count += 1
    
    return {"message": f"{marked_count} payments marked"}

# Meeting Fee Bulk Unmark
@api_router.post("/admin/fund/meetingfee/payments/bulk-unmark")
async def bulk_unmark_meetingfee(data: dict, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    result = await db.meetingfee_payments.delete_many({
        "chapter_id": chapter_id,
        "member_id": {"$in": data.get("member_ids", [])},
        "month": data.get("month"),
        "year": data.get("year")
    })
    return {"message": f"{result.deleted_count} payments unmarked"}

# ==================== END MEETING FEES ====================

# Miscellaneous Payments
@api_router.get("/admin/fund/misc")
async def get_misc_payments(user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    payments = await db.misc_payments.find({"chapter_id": chapter_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Calculate stats for each payment
    for payment in payments:
        records = await db.misc_payment_records.find({
            "misc_payment_id": payment["misc_payment_id"]
        }, {"_id": 0}).to_list(500)
        
        paid_count = len([r for r in records if r["status"] == "paid"])
        payment["paid_count"] = paid_count
        payment["total_collected"] = paid_count * payment["amount"]
    
    return payments

@api_router.post("/admin/fund/misc")
async def create_misc_payment(data: MiscPaymentCreate, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    misc_payment_id = f"MP{datetime.now().strftime('%Y%m%d%H%M%S')}"
    payment_data = {
        "misc_payment_id": misc_payment_id,
        "chapter_id": chapter_id,
        "payment_name": data.payment_name,
        "amount": data.amount,
        "due_date": data.due_date,
        "description": data.description,
        "created_at": datetime.now(IST).isoformat()
    }
    
    await db.misc_payments.insert_one(payment_data)
    return {"message": "Payment created", "misc_payment_id": misc_payment_id}

@api_router.get("/admin/fund/misc/{misc_payment_id}/members")
async def get_misc_payment_members(misc_payment_id: str, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    # Get payment details
    payment = await db.misc_payments.find_one({"misc_payment_id": misc_payment_id, "chapter_id": chapter_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    # Get all members
    members = await db.members.find(
        {"chapter_id": chapter_id, "status": "Active"},
        {"_id": 0, "member_id": 1, "unique_member_id": 1, "full_name": 1}
    ).to_list(500)
    
    # Get payment records
    records = await db.misc_payment_records.find({"misc_payment_id": misc_payment_id}, {"_id": 0}).to_list(500)
    record_map = {r["member_id"]: r for r in records}
    
    result = []
    for member in members:
        record = record_map.get(member["member_id"])
        result.append({
            "member_id": member["member_id"],
            "unique_member_id": member["unique_member_id"],
            "member_name": member["full_name"],
            "amount": payment["amount"],
            "status": record["status"] if record else "pending",
            "payment_mode": record.get("payment_mode") if record else None,
            "paid_date": record.get("paid_date") if record else None,
            "transaction_id": record.get("transaction_id") if record else None,
            "cheque_no": record.get("cheque_no") if record else None,
            "bank_name": record.get("bank_name") if record else None
        })
    
    result.sort(key=lambda x: (0 if x["status"] == "pending" else 1, x["member_name"].lower()))
    return {"payment": payment, "members": result}

@api_router.post("/admin/fund/misc/record")
async def mark_misc_payment(data: MiscPaymentRecordCreate, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    # Verify payment exists
    payment = await db.misc_payments.find_one({"misc_payment_id": data.misc_payment_id, "chapter_id": chapter_id})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    record_id = f"MR{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    record_data = {
        "record_id": record_id,
        "misc_payment_id": data.misc_payment_id,
        "member_id": data.member_id,
        "payment_mode": data.payment_mode,
        "status": "paid",
        "paid_date": datetime.now(IST).isoformat(),
        "transaction_id": data.transaction_id,
        "cheque_no": data.cheque_no,
        "bank_name": data.bank_name
    }
    
    # Check if already exists
    existing = await db.misc_payment_records.find_one({
        "misc_payment_id": data.misc_payment_id,
        "member_id": data.member_id
    })
    
    if existing:
        await db.misc_payment_records.update_one(
            {"_id": existing["_id"]},
            {"$set": record_data}
        )
    else:
        await db.misc_payment_records.insert_one(record_data)
    
    return {"message": "Payment recorded", "record_id": record_id}

# Misc Unmark Payment
class MiscUnmarkPayment(BaseModel):
    misc_payment_id: str
    member_id: str

@api_router.post("/admin/fund/misc/unmark")
async def unmark_misc_payment(data: MiscUnmarkPayment, user = Depends(get_current_user)):
    result = await db.misc_payment_records.delete_one({
        "misc_payment_id": data.misc_payment_id,
        "member_id": data.member_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    return {"message": "Payment unmarked"}

# Misc Update Payment
@api_router.put("/admin/fund/misc/{misc_payment_id}")
async def update_misc_payment(misc_payment_id: str, data: dict, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    update_fields = {}
    if "payment_name" in data:
        update_fields["payment_name"] = data["payment_name"]
    if "amount" in data:
        update_fields["amount"] = float(data["amount"])
    if "due_date" in data:
        update_fields["due_date"] = data["due_date"]
    if "description" in data:
        update_fields["description"] = data["description"]
    
    if not update_fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    result = await db.misc_payments.update_one(
        {"misc_payment_id": misc_payment_id, "chapter_id": chapter_id},
        {"$set": update_fields}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    return {"message": "Payment updated"}

# Misc Delete Payment
@api_router.delete("/admin/fund/misc/{misc_payment_id}")
async def delete_misc_payment(misc_payment_id: str, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    # Delete payment
    result = await db.misc_payments.delete_one({"misc_payment_id": misc_payment_id, "chapter_id": chapter_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    # Delete all related records
    await db.misc_payment_records.delete_many({"misc_payment_id": misc_payment_id})
    
    return {"message": "Payment deleted"}

# Misc Bulk Mark
@api_router.post("/admin/fund/misc/bulk-mark")
async def bulk_mark_misc(data: BulkMarkPayment, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    payment = await db.misc_payments.find_one({"misc_payment_id": data.payment_id, "chapter_id": chapter_id})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    marked_count = 0
    for member_id in data.member_ids:
        existing = await db.misc_payment_records.find_one({
            "misc_payment_id": data.payment_id,
            "member_id": member_id
        })
        
        if not existing:
            record_data = {
                "record_id": f"MR{datetime.now().strftime('%Y%m%d%H%M%S%f')}",
                "misc_payment_id": data.payment_id,
                "member_id": member_id,
                "payment_mode": data.payment_mode,
                "status": "paid",
                "paid_date": datetime.now(IST).isoformat()
            }
            await db.misc_payment_records.insert_one(record_data)
            marked_count += 1
    
    return {"message": f"{marked_count} payments marked"}

# Misc Bulk Unmark
@api_router.post("/admin/fund/misc/bulk-unmark")
async def bulk_unmark_misc(data: BulkUnmarkPayment, user = Depends(get_current_user)):
    result = await db.misc_payment_records.delete_many({
        "misc_payment_id": data.payment_id,
        "member_id": {"$in": data.member_ids}
    })
    
    return {"message": f"{result.deleted_count} payments unmarked"}

# Event-based Payments
@api_router.get("/admin/fund/events")
async def get_events(user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    events = await db.fund_events.find({"chapter_id": chapter_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    for event in events:
        # Get member count and payment stats
        if event["event_type"] == "compulsory":
            members_count = await db.members.count_documents({"chapter_id": chapter_id, "status": "Active"})
        else:
            members_count = await db.event_members.count_documents({"event_id": event["event_id"]})
        
        paid_count = await db.event_payments.count_documents({"event_id": event["event_id"], "status": "paid"})
        
        event["total_members"] = members_count
        event["paid_count"] = paid_count
        event["pending_count"] = members_count - paid_count
        event["total_collected"] = paid_count * event["amount"]
    
    return events

@api_router.post("/admin/fund/events")
async def create_event(data: EventCreate, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    event_id = f"EV{datetime.now().strftime('%Y%m%d%H%M%S')}"
    event_data = {
        "event_id": event_id,
        "chapter_id": chapter_id,
        "event_name": data.event_name,
        "amount": data.amount,
        "event_date": data.event_date,
        "event_type": data.event_type,
        "description": data.description,
        "created_at": datetime.now(IST).isoformat()
    }
    
    await db.fund_events.insert_one(event_data)
    
    # If optional, add selected members
    if data.event_type == "optional" and data.selected_members:
        for member_id in data.selected_members:
            await db.event_members.insert_one({
                "event_id": event_id,
                "member_id": member_id,
                "added_at": datetime.now(IST).isoformat()
            })
    
    return {"message": "Event created", "event_id": event_id}

# Event Update
@api_router.put("/admin/fund/events/{event_id}")
async def update_event(event_id: str, data: dict, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    update_fields = {}
    if "event_name" in data:
        update_fields["event_name"] = data["event_name"]
    if "amount" in data:
        update_fields["amount"] = float(data["amount"])
    if "event_date" in data:
        update_fields["event_date"] = data["event_date"]
    if "description" in data:
        update_fields["description"] = data["description"]
    
    if not update_fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    result = await db.fund_events.update_one(
        {"event_id": event_id, "chapter_id": chapter_id},
        {"$set": update_fields}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    
    return {"message": "Event updated"}

# Event Delete
@api_router.delete("/admin/fund/events/{event_id}")
async def delete_event(event_id: str, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    # Delete event
    result = await db.fund_events.delete_one({"event_id": event_id, "chapter_id": chapter_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Event not found")
    
    # Delete all related records
    await db.event_payments.delete_many({"event_id": event_id})
    await db.event_members.delete_many({"event_id": event_id})
    
    return {"message": "Event deleted"}

@api_router.get("/admin/fund/events/{event_id}/members")
async def get_event_members(event_id: str, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    # Get event details
    event = await db.fund_events.find_one({"event_id": event_id, "chapter_id": chapter_id}, {"_id": 0})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    # Get members based on event type
    if event["event_type"] == "compulsory":
        members = await db.members.find(
            {"chapter_id": chapter_id, "status": "Active"},
            {"_id": 0, "member_id": 1, "unique_member_id": 1, "full_name": 1}
        ).to_list(500)
    else:
        # Get selected members
        event_members = await db.event_members.find({"event_id": event_id}, {"_id": 0}).to_list(500)
        member_ids = [em["member_id"] for em in event_members]
        members = await db.members.find(
            {"chapter_id": chapter_id, "member_id": {"$in": member_ids}},
            {"_id": 0, "member_id": 1, "unique_member_id": 1, "full_name": 1}
        ).to_list(500)
    
    # Get payment records
    payments = await db.event_payments.find({"event_id": event_id}, {"_id": 0}).to_list(500)
    payment_map = {p["member_id"]: p for p in payments}
    
    result = []
    for member in members:
        payment = payment_map.get(member["member_id"])
        result.append({
            "member_id": member["member_id"],
            "unique_member_id": member["unique_member_id"],
            "member_name": member["full_name"],
            "amount": event["amount"],
            "status": payment["status"] if payment else "pending",
            "payment_mode": payment.get("payment_mode") if payment else None,
            "paid_date": payment.get("paid_date") if payment else None,
            "transaction_id": payment.get("transaction_id") if payment else None,
            "cheque_no": payment.get("cheque_no") if payment else None,
            "bank_name": payment.get("bank_name") if payment else None
        })
    
    result.sort(key=lambda x: (0 if x["status"] == "pending" else 1, x["member_name"].lower()))
    return {"event": event, "members": result}

@api_router.get("/admin/fund/events/payments")
async def get_all_event_payments(user = Depends(get_current_user)):
    """Get all event payments for the chapter"""
    chapter_id = user.get("chapter_id")
    
    # Get all events for this chapter
    events = await db.fund_events.find({"chapter_id": chapter_id}, {"_id": 0}).to_list(100)
    event_ids = [e["event_id"] for e in events]
    
    # Get all payments for these events
    payments = await db.event_payments.find(
        {"event_id": {"$in": event_ids}},
        {"_id": 0}
    ).to_list(5000)
    
    return payments

@api_router.post("/admin/fund/events/payment")
async def mark_event_payment(data: EventPaymentCreate, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    # Verify event exists
    event = await db.fund_events.find_one({"event_id": data.event_id, "chapter_id": chapter_id})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    payment_id = f"EP{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
    payment_data = {
        "payment_id": payment_id,
        "event_id": data.event_id,
        "member_id": data.member_id,
        "payment_mode": data.payment_mode,
        "status": "paid",
        "paid_date": datetime.now(IST).isoformat(),
        "transaction_id": data.transaction_id,
        "cheque_no": data.cheque_no,
        "bank_name": data.bank_name
    }
    
    # Check if already exists
    existing = await db.event_payments.find_one({
        "event_id": data.event_id,
        "member_id": data.member_id
    })
    
    if existing:
        await db.event_payments.update_one(
            {"_id": existing["_id"]},
            {"$set": payment_data}
        )
    else:
        await db.event_payments.insert_one(payment_data)
    
    return {"message": "Payment recorded", "payment_id": payment_id}

class EventUnmarkPayment(BaseModel):
    event_id: str
    member_id: str

# Event Unmark Payment
@api_router.post("/admin/fund/events/unmark")
async def unmark_event_payment(data: EventUnmarkPayment, user = Depends(get_current_user)):
    result = await db.event_payments.delete_one({
        "event_id": data.event_id,
        "member_id": data.member_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    return {"message": "Payment unmarked"}

# Event Bulk Mark
@api_router.post("/admin/fund/events/bulk-mark")
async def bulk_mark_event(data: BulkMarkPayment, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    event = await db.fund_events.find_one({"event_id": data.event_id, "chapter_id": chapter_id})
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    marked_count = 0
    for member_id in data.member_ids:
        existing = await db.event_payments.find_one({
            "event_id": data.event_id,
            "member_id": member_id
        })
        
        if not existing:
            payment_data = {
                "payment_id": f"EP{datetime.now().strftime('%Y%m%d%H%M%S%f')}",
                "event_id": data.event_id,
                "member_id": member_id,
                "payment_mode": data.payment_mode,
                "status": "paid",
                "paid_date": datetime.now(IST).isoformat()
            }
            await db.event_payments.insert_one(payment_data)
            marked_count += 1
    
    return {"message": f"{marked_count} payments marked"}

# Event Bulk Unmark
@api_router.post("/admin/fund/events/bulk-unmark")
async def bulk_unmark_event(data: BulkUnmarkPayment, user = Depends(get_current_user)):
    result = await db.event_payments.delete_many({
        "event_id": data.event_id,
        "member_id": {"$in": data.member_ids}
    })
    
    return {"message": f"{result.deleted_count} payments unmarked"}

# Fund Reports
@api_router.get("/admin/fund/reports/summary")
async def get_fund_summary(current_month: bool = True, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    # Get current month/year for filtering
    now = datetime.now(IST)
    current_month_num = now.month
    current_year = now.year
    
    # Build date filters for misc and events
    month_start = datetime(current_year, current_month_num, 1)
    if current_month_num == 12:
        month_end = datetime(current_year + 1, 1, 1)
    else:
        month_end = datetime(current_year, current_month_num + 1, 1)
    
    # Kitty collection - filter by current month if enabled
    if current_month:
        kitty_query = {"chapter_id": chapter_id, "status": "paid", "month": current_month_num, "year": current_year}
    else:
        kitty_query = {"chapter_id": chapter_id, "status": "paid"}
    
    kitty_payments = await db.kitty_payments.find(kitty_query, {"_id": 0, "amount": 1}).to_list(1000)
    kitty_total = sum(p["amount"] for p in kitty_payments)
    
    # Meeting Fees collection - filter by current month if enabled
    if current_month:
        meetingfee_query = {"chapter_id": chapter_id, "status": "paid", "month": current_month_num, "year": current_year}
    else:
        meetingfee_query = {"chapter_id": chapter_id, "status": "paid"}
    
    meetingfee_payments = await db.meetingfee_payments.find(meetingfee_query, {"_id": 0, "amount": 1}).to_list(1000)
    meetingfee_total = sum(p["amount"] for p in meetingfee_payments)
    
    # Event collection - filter by current month if enabled
    event_total = 0
    if current_month:
        event_payments = await db.event_payments.find({
            "status": "paid",
            "paid_date": {"$gte": month_start.isoformat(), "$lt": month_end.isoformat()}
        }, {"_id": 0, "event_id": 1}).to_list(1000)
    else:
        event_payments = await db.event_payments.find({"status": "paid"}, {"_id": 0, "event_id": 1}).to_list(1000)
    
    for payment in event_payments:
        event = await db.fund_events.find_one(
            {"event_id": payment["event_id"], "chapter_id": chapter_id},
            {"_id": 0, "amount": 1}
        )
        if event:
            event_total += event["amount"]
    
    return {
        "kitty_total": kitty_total,
        "meetingfee_total": meetingfee_total,
        "event_total": event_total,
        "grand_total": kitty_total + meetingfee_total + event_total,
        "month": current_month_num,
        "year": current_year,
        "is_current_month": current_month
    }

@api_router.get("/admin/fund/reports/member/{member_id}")
async def get_member_fund_report(member_id: str, user = Depends(get_current_user)):
    chapter_id = user.get("chapter_id")
    
    # Get member details
    member = await db.members.find_one({"member_id": member_id, "chapter_id": chapter_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    # Kitty payments
    kitty_payments = await db.kitty_payments.find(
        {"chapter_id": chapter_id, "member_id": member_id},
        {"_id": 0}
    ).to_list(100)
    
    # Misc payments
    misc_records = await db.misc_payment_records.find({"member_id": member_id}, {"_id": 0}).to_list(100)
    misc_payments = []
    for record in misc_records:
        payment = await db.misc_payments.find_one(
            {"misc_payment_id": record["misc_payment_id"], "chapter_id": chapter_id},
            {"_id": 0}
        )
        if payment:
            misc_payments.append({**record, "payment_name": payment["payment_name"], "amount": payment["amount"]})
    
    # Event payments
    event_records = await db.event_payments.find({"member_id": member_id}, {"_id": 0}).to_list(100)
    event_payments = []
    for record in event_records:
        event = await db.fund_events.find_one(
            {"event_id": record["event_id"], "chapter_id": chapter_id},
            {"_id": 0}
        )
        if event:
            event_payments.append({**record, "event_name": event["event_name"], "amount": event["amount"]})
    
    # Calculate totals
    kitty_total = sum(p["amount"] for p in kitty_payments if p["status"] == "paid")
    misc_total = sum(p["amount"] for p in misc_payments if p["status"] == "paid")
    event_total = sum(p["amount"] for p in event_payments if p["status"] == "paid")
    
    return {
        "member": member,
        "kitty_payments": kitty_payments,
        "misc_payments": misc_payments,
        "event_payments": event_payments,
        "totals": {
            "kitty": kitty_total,
            "misc": misc_total,
            "event": event_total,
            "total": kitty_total + misc_total + event_total
        }
    }

# Fund Report Export - Excel with Date Filter
@api_router.get("/admin/fund/reports/export/excel")
async def export_fund_report_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    months: Optional[str] = None,
    year: Optional[int] = None,
    category: Optional[str] = None,
    categories: Optional[str] = None,
    payment_status: Optional[str] = None,  # all, paid, pending
    event_id: Optional[str] = None,  # Filter by specific event
    user = Depends(get_current_user)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    chapter_id = user.get("chapter_id")
    chapter_name = user.get("chapter_name", "Chapter")
    
    # Parse date filters
    current_year = datetime.now().year
    current_month = datetime.now().month
    filter_year = year or current_year
    
    # Month-year list: Store tuples of (month, year) for proper cross-year handling
    month_year_list = []
    
    if months:
        # User specified months - use with filter_year, validate range 1-12
        month_list = [int(m.strip()) for m in months.split(",") if m.strip().isdigit()]
        month_year_list = [(m, filter_year) for m in month_list if 1 <= m <= 12]
    else:
        # Default to last 3 months with proper year handling
        for i in range(3):
            m = current_month - i
            y = current_year
            if m <= 0:
                m += 12
                y -= 1
            month_year_list.append((m, y))
        month_year_list.reverse()  # Oldest first
    
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    # Build filter description with year for each month
    filter_parts = []
    for m, y in month_year_list:
        filter_parts.append(f"{month_names[m-1]}'{str(y)[-2:]}")
    filter_desc = ', '.join(filter_parts)
    
    # Categories
    cat_list = []
    if categories:
        cat_list = [c.strip() for c in categories.split(',')]
    elif category and category != 'all':
        cat_list = [category]
    
    include_kitty = len(cat_list) == 0 or 'kitty' in cat_list
    include_meetingfee = len(cat_list) == 0 or 'meetingfee' in cat_list
    include_events = len(cat_list) == 0 or 'events' in cat_list
    
    category_desc = ""
    if cat_list:
        cat_names = {'kitty': 'Kitty', 'meetingfee': 'Meeting Fee', 'events': 'Events'}
        category_desc = f" - {', '.join([cat_names.get(c, c) for c in cat_list])}"
    
    status_desc = ""
    if payment_status and payment_status != 'all':
        status_desc = f" ({payment_status.capitalize()} Only)"
    
    # Get all members
    members = await db.members.find(
        {"chapter_id": chapter_id, "status": "Active"},
        {"_id": 0}
    ).to_list(500)
    
    # Pre-fetch all payments
    all_kitty = await db.kitty_payments.find(
        {"chapter_id": chapter_id, "status": "paid"},
        {"_id": 0}
    ).to_list(5000)
    
    all_kitty_settings = await db.kitty_settings.find(
        {"chapter_id": chapter_id},
        {"_id": 0}
    ).to_list(100)
    kitty_setting_map = {(s["month"], s["year"]): s.get("amount", 0) for s in all_kitty_settings}
    
    all_meetingfee = await db.meetingfee_payments.find(
        {"chapter_id": chapter_id, "status": "paid"},
        {"_id": 0}
    ).to_list(5000)
    
    all_mf_settings = await db.meetingfee_settings.find(
        {"chapter_id": chapter_id},
        {"_id": 0}
    ).to_list(100)
    mf_setting_map = {(s["month"], s["year"]): s.get("amount", 0) for s in all_mf_settings}
    
    # Get events
    all_events = await db.fund_events.find(
        {"chapter_id": chapter_id},
        {"_id": 0}
    ).to_list(100)
    events_map = {e["event_id"]: e for e in all_events}
    
    all_event_payments = await db.event_payments.find(
        {"chapter_id": chapter_id, "status": "paid"} if chapter_id else {"status": "paid"},
        {"_id": 0}
    ).to_list(5000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Fund Report"
    
    # Styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    
    # Title
    ws['A1'] = f"FUND REPORT - {chapter_name}{category_desc}{status_desc}"
    ws['A1'].font = title_font
    ws['A1'].fill = yellow_fill
    ws['A2'] = f"Period: {filter_desc}"
    ws['A3'] = f"Generated: {datetime.now(IST).strftime('%d-%b-%Y %H:%M')}"
    
    # Build dynamic headers (Point 2: Month-wise columns)
    headers = ["Sr", "Member ID", "Member Name"]
    
    # Add month-wise columns for each category with year
    if include_kitty:
        for m, y in month_year_list:
            headers.append(f"K-{month_names[m-1]}'{str(y)[-2:]}")
        headers.append("Kitty Total")
    
    if include_meetingfee:
        for m, y in month_year_list:
            headers.append(f"MF-{month_names[m-1]}'{str(y)[-2:]}")
        headers.append("M.Fee Total")
        headers.append("Pay Mode")
    
    if include_events:
        # Add event columns
        for event in all_events:
            if not event_id or event["event_id"] == event_id:
                headers.append(f"Event:{event.get('name', 'Unknown')[:10]}")
        headers.append("Event Total")
    
    headers.append("Grand Total")
    headers.append("Status")  # Paid/Pending status
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    
    # Data rows
    row = 6
    grand_totals = {
        "kitty_months": {(m, y): 0 for m, y in month_year_list},
        "kitty_total": 0,
        "mf_months": {(m, y): 0 for m, y in month_year_list},
        "mf_total": 0,
        "event_total": 0,
        "grand": 0
    }
    
    # Summary tracking
    summary_data = {
        "total_members": len(members),
        "paid_members": 0,
        "pending_members": 0,
        "kitty_pending_months": {(m, y): 0 for m, y in month_year_list},
        "kitty_pending_total": 0,
        "mf_pending_months": {(m, y): 0 for m, y in month_year_list},
        "mf_pending_total": 0,
        "event_pending_total": 0,
        "total_pending": 0
    }
    
    for idx, member in enumerate(members, 1):
        member_id = member["member_id"]
        member_data = {"kitty": {}, "meetingfee": {}, "events": {}}
        member_total = 0
        member_pending = 0
        
        # Calculate Kitty month-wise with proper year
        if include_kitty:
            for m, y in month_year_list:
                kitty_paid = [k for k in all_kitty if k.get("member_id") == member_id and k.get("month") == m and k.get("year") == y]
                paid_amt = sum(k.get("amount", 0) for k in kitty_paid)
                expected_amt = kitty_setting_map.get((m, y), 0)
                member_data["kitty"][(m, y)] = {"paid": paid_amt, "expected": expected_amt}
                if paid_amt == 0 and expected_amt > 0:
                    member_pending += expected_amt
                    summary_data["kitty_pending_months"][(m, y)] += expected_amt
                    summary_data["kitty_pending_total"] += expected_amt
        
        # Calculate Meeting Fee month-wise with proper year
        mf_payment_mode = "-"
        if include_meetingfee:
            for m, y in month_year_list:
                mf_paid = [mf for mf in all_meetingfee if mf.get("member_id") == member_id and mf.get("month") == m and mf.get("year") == y]
                paid_amt = sum(mf.get("amount", 0) for mf in mf_paid)
                expected_amt = mf_setting_map.get((m, y), 0)
                member_data["meetingfee"][(m, y)] = {"paid": paid_amt, "expected": expected_amt}
                if mf_paid and mf_paid[0].get("payment_mode"):
                    mf_payment_mode = mf_paid[0]["payment_mode"]
                if paid_amt == 0 and expected_amt > 0:
                    member_pending += expected_amt
                    summary_data["mf_pending_months"][(m, y)] += expected_amt
                    summary_data["mf_pending_total"] += expected_amt
        
        # Calculate Events
        if include_events:
            for event in all_events:
                if not event_id or event["event_id"] == event_id:
                    ev_paid = [ep for ep in all_event_payments if ep.get("member_id") == member_id and ep.get("event_id") == event["event_id"]]
                    paid_amt = event.get("amount", 0) if ev_paid else 0
                    member_data["events"][event["event_id"]] = {"paid": paid_amt, "expected": event.get("amount", 0)}
                    if not ev_paid:
                        member_pending += event.get("amount", 0)
                        summary_data["event_pending_total"] += event.get("amount", 0)
        
        # Calculate totals
        kitty_total = sum(d["paid"] for d in member_data["kitty"].values())
        mf_total = sum(d["paid"] for d in member_data["meetingfee"].values())
        event_total = sum(d["paid"] for d in member_data["events"].values())
        member_total = kitty_total + mf_total + event_total
        
        # Track member paid/pending status for summary
        if member_pending > 0:
            summary_data["pending_members"] += 1
            summary_data["total_pending"] += member_pending
        elif member_total > 0:
            summary_data["paid_members"] += 1
        
        # Payment status filter (Points 3, 4, 5, 6)
        is_fully_paid = member_pending == 0 and member_total > 0
        is_pending = member_pending > 0
        
        if payment_status == 'paid' and not is_fully_paid:
            continue
        if payment_status == 'pending' and not is_pending:
            continue
        
        # Write row
        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = member.get("unique_member_id", "")
        ws.cell(row=row, column=3).value = member.get("full_name", "")
        
        col = 4
        
        # Kitty month-wise (Point 7)
        if include_kitty:
            for m, y in month_year_list:
                data = member_data["kitty"].get((m, y), {"paid": 0})
                cell = ws.cell(row=row, column=col)
                cell.value = data["paid"]
                if data["paid"] == 0 and data.get("expected", 0) > 0:
                    cell.fill = red_fill  # Highlight pending
                elif data["paid"] > 0:
                    cell.fill = green_fill  # Highlight paid
                grand_totals["kitty_months"][(m, y)] += data["paid"]
                col += 1
            ws.cell(row=row, column=col).value = kitty_total
            grand_totals["kitty_total"] += kitty_total
            col += 1
        
        # Meeting Fee month-wise
        if include_meetingfee:
            for m, y in month_year_list:
                data = member_data["meetingfee"].get((m, y), {"paid": 0})
                cell = ws.cell(row=row, column=col)
                cell.value = data["paid"]
                if data["paid"] == 0 and data.get("expected", 0) > 0:
                    cell.fill = red_fill
                elif data["paid"] > 0:
                    cell.fill = green_fill
                grand_totals["mf_months"][(m, y)] += data["paid"]
                col += 1
            ws.cell(row=row, column=col).value = mf_total
            grand_totals["mf_total"] += mf_total
            col += 1
            ws.cell(row=row, column=col).value = mf_payment_mode
            col += 1
        
        # Events (Point 8)
        if include_events:
            for event in all_events:
                if not event_id or event["event_id"] == event_id:
                    data = member_data["events"].get(event["event_id"], {"paid": 0})
                    cell = ws.cell(row=row, column=col)
                    cell.value = data["paid"]
                    if data["paid"] == 0 and data.get("expected", 0) > 0:
                        cell.fill = red_fill
                    elif data["paid"] > 0:
                        cell.fill = green_fill
                    col += 1
            ws.cell(row=row, column=col).value = event_total
            grand_totals["event_total"] += event_total
            col += 1
        
        # Grand Total
        ws.cell(row=row, column=col).value = member_total
        grand_totals["grand"] += member_total
        col += 1
        
        # Status
        status_cell = ws.cell(row=row, column=col)
        if member_pending > 0:
            status_cell.value = f"Pending: Rs.{member_pending}"
            status_cell.fill = red_fill
        elif member_total > 0:
            status_cell.value = "Paid"
            status_cell.fill = green_fill
        else:
            status_cell.value = "No Activity"
        
        row += 1
    
    # Total row
    row += 1
    ws.cell(row=row, column=2).value = "TOTAL"
    ws.cell(row=row, column=2).font = header_font
    
    col = 4
    if include_kitty:
        for m, y in month_year_list:
            ws.cell(row=row, column=col).value = grand_totals["kitty_months"][(m, y)]
            ws.cell(row=row, column=col).font = header_font
            col += 1
        ws.cell(row=row, column=col).value = grand_totals["kitty_total"]
        ws.cell(row=row, column=col).font = header_font
        ws.cell(row=row, column=col).fill = yellow_fill
        col += 1
    
    if include_meetingfee:
        for m, y in month_year_list:
            ws.cell(row=row, column=col).value = grand_totals["mf_months"][(m, y)]
            ws.cell(row=row, column=col).font = header_font
            col += 1
        ws.cell(row=row, column=col).value = grand_totals["mf_total"]
        ws.cell(row=row, column=col).font = header_font
        ws.cell(row=row, column=col).fill = yellow_fill
        col += 1
        col += 1  # Skip mode column
    
    if include_events:
        col += len([e for e in all_events if not event_id or e["event_id"] == event_id])
        ws.cell(row=row, column=col).value = grand_totals["event_total"]
        ws.cell(row=row, column=col).font = header_font
        ws.cell(row=row, column=col).fill = yellow_fill
        col += 1
    
    ws.cell(row=row, column=col).value = grand_totals["grand"]
    ws.cell(row=row, column=col).font = header_font
    ws.cell(row=row, column=col).fill = yellow_fill
    
    # Adjust column widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    
    # ========== SUMMARY SECTION ==========
    row += 3
    summary_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    # Summary Title
    ws.cell(row=row, column=1).value = "═══════ SUMMARY ═══════"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 2
    
    # Member Count Summary
    ws.cell(row=row, column=1).value = "MEMBER STATUS"
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).fill = summary_fill
    row += 1
    ws.cell(row=row, column=1).value = f"Total Members: {summary_data['total_members']}"
    row += 1
    ws.cell(row=row, column=1).value = f"Fully Paid Members: {summary_data['paid_members']}"
    ws.cell(row=row, column=1).fill = green_fill
    row += 1
    ws.cell(row=row, column=1).value = f"Pending Members: {summary_data['pending_members']}"
    ws.cell(row=row, column=1).fill = red_fill
    row += 2
    
    # Category-wise Summary
    ws.cell(row=row, column=1).value = "CATEGORY-WISE COLLECTION"
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).fill = summary_fill
    row += 1
    
    if include_kitty:
        ws.cell(row=row, column=1).value = f"KITTY - Collected: Rs.{grand_totals['kitty_total']} | Pending: Rs.{summary_data['kitty_pending_total']}"
        row += 1
    if include_meetingfee:
        ws.cell(row=row, column=1).value = f"MEETING FEE - Collected: Rs.{grand_totals['mf_total']} | Pending: Rs.{summary_data['mf_pending_total']}"
        row += 1
    if include_events:
        ws.cell(row=row, column=1).value = f"EVENTS - Collected: Rs.{grand_totals['event_total']} | Pending: Rs.{summary_data['event_pending_total']}"
        row += 1
    
    row += 1
    ws.cell(row=row, column=1).value = f"GRAND TOTAL - Collected: Rs.{grand_totals['grand']} | Pending: Rs.{summary_data['total_pending']}"
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).fill = yellow_fill
    row += 2
    
    # Month-wise Summary
    ws.cell(row=row, column=1).value = "MONTH-WISE BREAKDOWN"
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).fill = summary_fill
    row += 1
    
    for m, y in month_year_list:
        month_collected = grand_totals["kitty_months"].get((m, y), 0) + grand_totals["mf_months"].get((m, y), 0)
        month_pending = summary_data["kitty_pending_months"].get((m, y), 0) + summary_data["mf_pending_months"].get((m, y), 0)
        ws.cell(row=row, column=1).value = f"{month_names[m-1]} {y}: Collected Rs.{month_collected} | Pending Rs.{month_pending}"
        row += 1
    
    # Save
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=fund_report_{chapter_id}.xlsx"}
    )

# Fund Report Export - PDF with Date Filter
@api_router.get("/admin/fund/reports/export/pdf")
async def export_fund_report_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    months: Optional[str] = None,
    year: Optional[int] = None,
    category: Optional[str] = None,
    categories: Optional[str] = None,
    payment_status: Optional[str] = None,
    event_id: Optional[str] = None,
    user = Depends(get_current_user)
):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    
    chapter_id = user.get("chapter_id")
    chapter_name = user.get("chapter_name", "Chapter")
    
    # Parse filters with proper year handling
    current_year = datetime.now().year
    current_month = datetime.now().month
    filter_year = year or current_year
    
    # Month-year list: Store tuples of (month, year) for proper cross-year handling
    month_year_list = []
    
    if months:
        # Validate month range 1-12
        month_list = [int(m.strip()) for m in months.split(",") if m.strip().isdigit()]
        month_year_list = [(m, filter_year) for m in month_list if 1 <= m <= 12]
    else:
        # Default to last 3 months with proper year handling
        for i in range(3):
            m = current_month - i
            y = current_year
            if m <= 0:
                m += 12
                y -= 1
            month_year_list.append((m, y))
        month_year_list.reverse()
    
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    # Build filter description with year for each month
    filter_parts = []
    for m, y in month_year_list:
        filter_parts.append(f"{month_names[m-1]}'{str(y)[-2:]}")
    filter_desc = ', '.join(filter_parts)
    
    # Categories
    cat_list = []
    if categories:
        cat_list = [c.strip() for c in categories.split(',')]
    elif category and category != 'all':
        cat_list = [category]
    
    include_kitty = len(cat_list) == 0 or 'kitty' in cat_list
    include_meetingfee = len(cat_list) == 0 or 'meetingfee' in cat_list
    include_events = len(cat_list) == 0 or 'events' in cat_list
    
    category_desc = ""
    if cat_list:
        cat_names = {'kitty': 'Kitty', 'meetingfee': 'M.Fee', 'events': 'Events'}
        category_desc = f" - {', '.join([cat_names.get(c, c) for c in cat_list])}"
    
    status_desc = ""
    if payment_status and payment_status != 'all':
        status_desc = f" ({payment_status.capitalize()})"
    
    # Get data
    members = await db.members.find({"chapter_id": chapter_id, "status": "Active"}, {"_id": 0}).to_list(500)
    
    all_kitty = await db.kitty_payments.find({"chapter_id": chapter_id, "status": "paid"}, {"_id": 0}).to_list(5000)
    all_kitty_settings = await db.kitty_settings.find({"chapter_id": chapter_id}, {"_id": 0}).to_list(100)
    kitty_setting_map = {(s["month"], s["year"]): s.get("amount", 0) for s in all_kitty_settings}
    
    all_meetingfee = await db.meetingfee_payments.find({"chapter_id": chapter_id, "status": "paid"}, {"_id": 0}).to_list(5000)
    all_mf_settings = await db.meetingfee_settings.find({"chapter_id": chapter_id}, {"_id": 0}).to_list(100)
    mf_setting_map = {(s["month"], s["year"]): s.get("amount", 0) for s in all_mf_settings}
    
    all_events = await db.fund_events.find({"chapter_id": chapter_id}, {"_id": 0}).to_list(100)
    all_event_payments = await db.event_payments.find({"status": "paid"}, {"_id": 0}).to_list(5000)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=15, bottomMargin=15, leftMargin=10, rightMargin=10)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=14, alignment=TA_CENTER)
    elements.append(Paragraph(f"FUND REPORT - {chapter_name}{category_desc}{status_desc}", title_style))
    elements.append(Paragraph(f"Period: {filter_desc} | Generated: {datetime.now(IST).strftime('%d-%b-%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 5*mm))
    
    # Build headers with proper month-year
    headers = ["Sr", "Name"]
    if include_kitty:
        for m, y in month_year_list:
            headers.append(f"K-{month_names[m-1][:3]}'{str(y)[-2:]}")
        headers.append("K.Tot")
    if include_meetingfee:
        for m, y in month_year_list:
            headers.append(f"MF-{month_names[m-1][:3]}'{str(y)[-2:]}")
        headers.append("MF.Tot")
    if include_events:
        headers.append("Events")
    headers.append("Total")
    headers.append("Status")
    
    table_data = [headers]
    grand_totals = {"kitty": 0, "mf": 0, "event": 0, "grand": 0}
    
    # Summary tracking for PDF
    summary_data = {
        "total_members": len(members),
        "paid_members": 0,
        "pending_members": 0,
        "kitty_months_collected": {(m, y): 0 for m, y in month_year_list},
        "kitty_months_pending": {(m, y): 0 for m, y in month_year_list},
        "kitty_pending_total": 0,
        "mf_months_collected": {(m, y): 0 for m, y in month_year_list},
        "mf_months_pending": {(m, y): 0 for m, y in month_year_list},
        "mf_pending_total": 0,
        "event_pending_total": 0,
        "total_pending": 0
    }
    
    for idx, member in enumerate(members, 1):
        member_id = member["member_id"]
        row_data = [idx, member.get("full_name", "")[:15]]
        
        kitty_total = 0
        mf_total = 0
        event_total = 0
        pending = 0
        
        if include_kitty:
            for m, y in month_year_list:
                paid = sum(k.get("amount", 0) for k in all_kitty if k.get("member_id") == member_id and k.get("month") == m and k.get("year") == y)
                expected = kitty_setting_map.get((m, y), 0)
                row_data.append(paid if paid > 0 else ("-" if expected == 0 else "P"))
                kitty_total += paid
                summary_data["kitty_months_collected"][(m, y)] += paid
                if paid == 0 and expected > 0:
                    pending += expected
                    summary_data["kitty_months_pending"][(m, y)] += expected
                    summary_data["kitty_pending_total"] += expected
            row_data.append(kitty_total)
            grand_totals["kitty"] += kitty_total
        
        if include_meetingfee:
            for m, y in month_year_list:
                paid = sum(mf.get("amount", 0) for mf in all_meetingfee if mf.get("member_id") == member_id and mf.get("month") == m and mf.get("year") == y)
                expected = mf_setting_map.get((m, y), 0)
                row_data.append(paid if paid > 0 else ("-" if expected == 0 else "P"))
                mf_total += paid
                summary_data["mf_months_collected"][(m, y)] += paid
                if paid == 0 and expected > 0:
                    pending += expected
                    summary_data["mf_months_pending"][(m, y)] += expected
                    summary_data["mf_pending_total"] += expected
            row_data.append(mf_total)
            grand_totals["mf"] += mf_total
        
        if include_events:
            for event in all_events:
                if not event_id or event["event_id"] == event_id:
                    ev_paid = [ep for ep in all_event_payments if ep.get("member_id") == member_id and ep.get("event_id") == event["event_id"]]
                    if ev_paid:
                        event_total += event.get("amount", 0)
                    else:
                        pending += event.get("amount", 0)
                        summary_data["event_pending_total"] += event.get("amount", 0)
            row_data.append(event_total)
            grand_totals["event"] += event_total
        
        total = kitty_total + mf_total + event_total
        row_data.append(total)
        grand_totals["grand"] += total
        
        # Track member paid/pending status for summary
        if pending > 0:
            summary_data["pending_members"] += 1
            summary_data["total_pending"] += pending
        elif total > 0:
            summary_data["paid_members"] += 1
        
        # Status
        if pending > 0:
            row_data.append(f"P:{pending}")
        elif total > 0:
            row_data.append("OK")
        else:
            row_data.append("-")
        
        # Filter by payment status
        if payment_status == 'paid' and pending > 0:
            continue
        if payment_status == 'pending' and pending == 0:
            continue
        
        table_data.append(row_data)
    
    # Total row
    total_row = ["", "TOTAL"]
    if include_kitty:
        total_row.extend([""] * len(month_year_list))
        total_row.append(grand_totals["kitty"])
    if include_meetingfee:
        total_row.extend([""] * len(month_year_list))
        total_row.append(grand_totals["mf"])
    if include_events:
        total_row.append(grand_totals["event"])
    total_row.append(grand_totals["grand"])
    total_row.append("")
    table_data.append(total_row)
    
    # Calculate column widths
    num_cols = len(headers)
    base_width = 750 / num_cols
    col_widths = [20, 80] + [min(base_width, 35)] * (num_cols - 2)
    
    # Create table
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, -1), (-1, -1), colors.yellow),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    
    elements.append(table)
    
    # Legend
    elements.append(Spacer(1, 5*mm))
    elements.append(Paragraph("Legend: P = Pending, OK = Paid, - = No Activity", styles['Normal']))
    
    # ========== SUMMARY SECTION FOR PDF ==========
    elements.append(Spacer(1, 8*mm))
    
    summary_title_style = ParagraphStyle('SummaryTitle', parent=styles['Heading2'], fontSize=12, alignment=TA_CENTER)
    elements.append(Paragraph("═══════ SUMMARY ═══════", summary_title_style))
    elements.append(Spacer(1, 4*mm))
    
    # Summary Table Data
    summary_table_data = []
    
    # Member Status Section
    summary_table_data.append(["MEMBER STATUS", "", "", ""])
    summary_table_data.append(["Total Members", str(summary_data['total_members']), "Fully Paid", str(summary_data['paid_members'])])
    summary_table_data.append(["Pending Members", str(summary_data['pending_members']), "", ""])
    summary_table_data.append(["", "", "", ""])
    
    # Category-wise Collection Section
    summary_table_data.append(["CATEGORY-WISE COLLECTION", "Collected (Rs.)", "Pending (Rs.)", ""])
    if include_kitty:
        summary_table_data.append(["Kitty", str(grand_totals['kitty']), str(summary_data['kitty_pending_total']), ""])
    if include_meetingfee:
        summary_table_data.append(["Meeting Fee", str(grand_totals['mf']), str(summary_data['mf_pending_total']), ""])
    if include_events:
        summary_table_data.append(["Events", str(grand_totals['event']), str(summary_data['event_pending_total']), ""])
    summary_table_data.append(["GRAND TOTAL", str(grand_totals['grand']), str(summary_data['total_pending']), ""])
    summary_table_data.append(["", "", "", ""])
    
    # Month-wise Breakdown Section
    summary_table_data.append(["MONTH-WISE BREAKDOWN", "Collected (Rs.)", "Pending (Rs.)", ""])
    for m, y in month_year_list:
        month_collected = summary_data["kitty_months_collected"].get((m, y), 0) + summary_data["mf_months_collected"].get((m, y), 0)
        month_pending = summary_data["kitty_months_pending"].get((m, y), 0) + summary_data["mf_months_pending"].get((m, y), 0)
        summary_table_data.append([f"{month_names[m-1]} {y}", str(month_collected), str(month_pending), ""])
    
    # Create Summary Table
    summary_table = Table(summary_table_data, colWidths=[120, 80, 80, 50])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.9, 0.95, 1)),  # Light blue for section headers
        ('BACKGROUND', (0, 4), (-1, 4), colors.Color(0.9, 0.95, 1)),
        ('BACKGROUND', (0, len(summary_table_data)-len(month_year_list)-1), (-1, len(summary_table_data)-len(month_year_list)-1), colors.Color(0.9, 0.95, 1)),
        ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 4), (0, 4), 'Helvetica-Bold'),
        ('FONTNAME', (0, len(summary_table_data)-len(month_year_list)-1), (0, len(summary_table_data)-len(month_year_list)-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 3+len([c for c in [include_kitty, include_meetingfee, include_events] if c])+1), (-1, 3+len([c for c in [include_kitty, include_meetingfee, include_events] if c])+1), colors.yellow),  # Grand total row
        ('FONTNAME', (0, 3+len([c for c in [include_kitty, include_meetingfee, include_events] if c])+1), (-1, 3+len([c for c in [include_kitty, include_meetingfee, include_events] if c])+1), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    elements.append(summary_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=fund_report_{chapter_id}.pdf"}
    )

# Quick View - Get payments by date with month-wise breakdown
@api_router.get("/admin/fund/quick-view")
async def get_payments_by_date(date: str = None, category: str = "kitty", user = Depends(get_current_user)):
    """Get all payments made on a specific date for Quick View popup with month-wise breakdown"""
    chapter_id = user.get("chapter_id")
    
    # If no date provided, use today's date in IST
    if not date:
        now_ist = datetime.now(IST)
        date = now_ist.strftime("%Y-%m-%d")
    
    # Parse the date
    try:
        target_date = datetime.strptime(date, "%Y-%m-%d").date()
    except:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    # Get payments based on category
    payments = []
    total_amount = 0
    month_breakdown = {}  # {(month, year): {"count": 0, "amount": 0}}
    
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    
    if category == "kitty":
        all_payments = await db.kitty_payments.find(
            {"chapter_id": chapter_id, "status": "paid"},
            {"_id": 0}
        ).to_list(5000)
        
        for p in all_payments:
            paid_date = p.get("paid_date")
            if paid_date:
                try:
                    if isinstance(paid_date, str):
                        paid_dt = datetime.fromisoformat(paid_date.replace('Z', '+00:00'))
                        paid_date_only = paid_dt.astimezone(IST).date()
                    else:
                        paid_date_only = paid_date.date() if hasattr(paid_date, 'date') else paid_date
                    
                    if paid_date_only == target_date:
                        member = await db.members.find_one({"member_id": p.get("member_id")}, {"_id": 0, "full_name": 1})
                        payment_month = p.get("month", 0)
                        payment_year = p.get("year", 0)
                        amount = p.get("amount", 0)
                        
                        payments.append({
                            "member_name": member.get("full_name", "Unknown") if member else "Unknown",
                            "amount": amount,
                            "paid_date": paid_date,
                            "for_month": payment_month,
                            "for_year": payment_year
                        })
                        total_amount += amount
                        
                        # Update month breakdown
                        key = (payment_month, payment_year)
                        if key not in month_breakdown:
                            month_breakdown[key] = {"count": 0, "amount": 0, "month": payment_month, "year": payment_year}
                        month_breakdown[key]["count"] += 1
                        month_breakdown[key]["amount"] += amount
                except Exception as e:
                    continue
    
    elif category == "meetingfee":
        all_payments = await db.meetingfee_payments.find(
            {"chapter_id": chapter_id, "status": "paid"},
            {"_id": 0}
        ).to_list(5000)
        
        for p in all_payments:
            paid_date = p.get("paid_date")
            if paid_date:
                try:
                    if isinstance(paid_date, str):
                        paid_dt = datetime.fromisoformat(paid_date.replace('Z', '+00:00'))
                        paid_date_only = paid_dt.astimezone(IST).date()
                    else:
                        paid_date_only = paid_date.date() if hasattr(paid_date, 'date') else paid_date
                    
                    if paid_date_only == target_date:
                        member = await db.members.find_one({"member_id": p.get("member_id")}, {"_id": 0, "full_name": 1})
                        payment_month = p.get("month", 0)
                        payment_year = p.get("year", 0)
                        amount = p.get("amount", 0)
                        
                        payments.append({
                            "member_name": member.get("full_name", "Unknown") if member else "Unknown",
                            "amount": amount,
                            "paid_date": paid_date,
                            "for_month": payment_month,
                            "for_year": payment_year
                        })
                        total_amount += amount
                        
                        key = (payment_month, payment_year)
                        if key not in month_breakdown:
                            month_breakdown[key] = {"count": 0, "amount": 0, "month": payment_month, "year": payment_year}
                        month_breakdown[key]["count"] += 1
                        month_breakdown[key]["amount"] += amount
                except:
                    continue
    
    elif category == "events":
        all_payments = await db.event_payments.find(
            {"chapter_id": chapter_id, "status": "paid"},
            {"_id": 0}
        ).to_list(5000)
        
        for p in all_payments:
            paid_date = p.get("paid_date")
            if paid_date:
                try:
                    if isinstance(paid_date, str):
                        paid_dt = datetime.fromisoformat(paid_date.replace('Z', '+00:00'))
                        paid_date_only = paid_dt.astimezone(IST).date()
                    else:
                        paid_date_only = paid_date.date() if hasattr(paid_date, 'date') else paid_date
                    
                    if paid_date_only == target_date:
                        member = await db.members.find_one({"member_id": p.get("member_id")}, {"_id": 0, "full_name": 1})
                        event = await db.events.find_one({"event_id": p.get("event_id")}, {"_id": 0, "event_name": 1})
                        amount = p.get("amount", 0)
                        event_name = event.get("event_name", "Unknown Event") if event else "Unknown Event"
                        
                        payments.append({
                            "member_name": member.get("full_name", "Unknown") if member else "Unknown",
                            "amount": amount,
                            "event_name": event_name,
                            "paid_date": paid_date
                        })
                        total_amount += amount
                        
                        # For events, group by event name
                        key = event_name
                        if key not in month_breakdown:
                            month_breakdown[key] = {"count": 0, "amount": 0, "event_name": event_name}
                        month_breakdown[key]["count"] += 1
                        month_breakdown[key]["amount"] += amount
                except:
                    continue
    
    # Convert month_breakdown to sorted list
    if category == "events":
        breakdown_list = list(month_breakdown.values())
    else:
        breakdown_list = sorted(
            month_breakdown.values(),
            key=lambda x: (x.get("year", 0), x.get("month", 0)),
            reverse=True
        )
        # Add month name to each item
        for item in breakdown_list:
            if item.get("month") and item.get("month") > 0:
                item["month_name"] = month_names[item["month"] - 1]
    
    return {
        "date": date,
        "category": category,
        "payments": payments,
        "count": len(payments),
        "total_amount": total_amount,
        "month_breakdown": breakdown_list
    }

# ===== DEVELOPER ENDPOINTS =====
from uuid import uuid4 as _uuid4

@api_router.post("/developer/seed")
async def seed_developer():
    """One-time seed developer account"""
    data = DeveloperSeedRequest()

    existing = await db.developers.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Developer account already exists")

    dev_data = {
        "dev_id": str(_uuid4()),
        "email": data.email,
        "password_hash": hash_password(data.password),
        "name": data.name,
        "company": "AasaanApp",
        "role": "developer",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.developers.insert_one(dev_data)
    return {"message": "Developer account seeded successfully", "email": data.email}

@api_router.post("/developer/login", response_model=DeveloperLoginResponse)
async def developer_login(data: DeveloperLoginRequest):
    developer = await db.developers.find_one({"email": data.email}, {"_id": 0})
    if not developer or not verify_password(data.password, developer["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")

    token = create_access_token({"email": data.email, "role": "developer", "dev_id": developer["dev_id"]})
    return DeveloperLoginResponse(token=token, role="developer", email=data.email, name=developer["name"])

@api_router.post("/developer/superadmins", response_model=SuperAdminResponse)
async def create_superadmin(data: SuperAdminCreate, user=Depends(require_role("developer"))):
    """Developer creates a new Super Admin / ED"""
    # Check if mobile already exists
    existing = await db.superadmins.find_one({"mobile": data.mobile})
    if existing:
        raise HTTPException(status_code=400, detail="Super Admin with this mobile already exists")

    superadmin_id = str(_uuid4())
    superadmin_data = {
        "superadmin_id": superadmin_id,
        "name": data.name,
        "email": data.email,
        "mobile": data.mobile,
        "password_hash": hash_password(data.password),
        "region": data.region,
        "state": data.state,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user.get("email")
    }
    await db.superadmins.insert_one(superadmin_data)

    # Auto-assign FREE TRIAL subscription
    settings = await db.subscription_settings.find_one({"setting_id": "default"}, {"_id": 0})
    if settings and settings.get("free_trial", {}).get("enabled", True):
        trial_days = settings.get("free_trial", {}).get("duration_days", 30)
        trial_chapters = settings.get("free_trial", {}).get("max_chapters", 1)
    else:
        trial_days = 30
        trial_chapters = 1

    trial_sub = {
        "subscription_id": str(_uuid4()),
        "superadmin_id": superadmin_id,
        "plan_type": "trial",
        "billing_cycle": "monthly",
        "chapters_allowed": trial_chapters,
        "start_date": datetime.now(timezone.utc).isoformat(),
        "end_date": (datetime.now(timezone.utc) + timedelta(days=trial_days)).isoformat(),
        "amount_paid": 0,
        "payment_method": "manual",
        "payment_ref": "free_trial",
        "status": "active",
        "auto_renew": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.subscriptions.insert_one(trial_sub)

    # Count chapters for this superadmin
    chapter_count = await db.chapters.count_documents({"created_by": data.mobile})

    return SuperAdminResponse(
        superadmin_id=superadmin_id,
        name=data.name,
        email=data.email,
        mobile=data.mobile,
        region=data.region,
        state=data.state,
        is_active=True,
        created_at=superadmin_data["created_at"],
        chapter_count=chapter_count
    )

@api_router.get("/developer/superadmins")
async def list_superadmins(user=Depends(require_role("developer"))):
    """List all Super Admins with their chapter counts"""
    superadmins = await db.superadmins.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)

    result = []
    for sa in superadmins:
        chapter_count = await db.chapters.count_documents({"created_by": sa.get("mobile", "")})
        result.append({
            "superadmin_id": sa.get("superadmin_id", sa.get("mobile", "")),
            "name": sa.get("name", ""),
            "email": sa.get("email", ""),
            "mobile": sa.get("mobile", ""),
            "region": sa.get("region", ""),
            "state": sa.get("state", ""),
            "is_active": sa.get("is_active", True),
            "created_at": sa.get("created_at", ""),
            "chapter_count": chapter_count
        })

    return result

@api_router.put("/developer/superadmins/{superadmin_id}")
async def update_superadmin(superadmin_id: str, data: SuperAdminUpdate, user=Depends(require_role("developer"))):
    """Update ED details, activate/deactivate"""
    update_data = {}
    for k, v in data.dict(exclude_unset=True).items():
        if isinstance(v, bool) or v is not None:
            update_data[k] = v

    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")

    # Try finding by superadmin_id first, then by mobile (backward compat)
    result = await db.superadmins.update_one(
        {"$or": [{"superadmin_id": superadmin_id}, {"mobile": superadmin_id}]},
        {"$set": update_data}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Super Admin not found")

    return {"message": "Super Admin updated successfully"}

@api_router.delete("/developer/superadmins/{superadmin_id}")
async def delete_superadmin(superadmin_id: str, user=Depends(require_role("developer"))):
    """Soft delete a Super Admin"""
    result = await db.superadmins.update_one(
        {"$or": [{"superadmin_id": superadmin_id}, {"mobile": superadmin_id}]},
        {"$set": {"is_active": False, "deleted_at": datetime.now(timezone.utc).isoformat()}}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Super Admin not found")

    return {"message": "Super Admin deactivated successfully"}

@api_router.get("/developer/dashboard/stats")
async def developer_dashboard_stats(user=Depends(require_role("developer"))):
    """Dashboard stats: total EDs, chapters, members, revenue"""
    total_eds = await db.superadmins.count_documents({"is_active": {"$ne": False}})
    total_chapters = await db.chapters.count_documents({})
    total_members = await db.members.count_documents({"status": "Active"})

    # Calculate total revenue from all payment collections
    kitty_pipeline = [{"$match": {"status": "paid"}}, {"$group": {"_id": None, "total": {"$sum": "$amount"}}}]
    meeting_fee_pipeline = [{"$match": {"status": "paid"}}, {"$group": {"_id": None, "total": {"$sum": "$amount"}}}]
    misc_pipeline = [{"$match": {"status": "paid"}}, {"$group": {"_id": None, "total": {"$sum": "$amount"}}}]
    event_pipeline = [{"$match": {"status": "paid"}}, {"$group": {"_id": None, "total": {"$sum": "$amount"}}}]

    kitty_rev = await db.kitty_payments.aggregate(kitty_pipeline).to_list(1)
    meeting_fee_rev = await db.meeting_fee_payments.aggregate(meeting_fee_pipeline).to_list(1)
    misc_rev = await db.misc_payment_records.aggregate(misc_pipeline).to_list(1)
    event_rev = await db.event_payments.aggregate(event_pipeline).to_list(1)

    total_revenue = (
        (kitty_rev[0]["total"] if kitty_rev else 0) +
        (meeting_fee_rev[0]["total"] if meeting_fee_rev else 0) +
        (misc_rev[0]["total"] if misc_rev else 0) +
        (event_rev[0]["total"] if event_rev else 0)
    )

    return {
        "total_eds": total_eds,
        "total_chapters": total_chapters,
        "total_members": total_members,
        "total_revenue": total_revenue
    }

# ===== SUBSCRIPTION SETTINGS ENDPOINTS =====

DEFAULT_SUBSCRIPTION_SETTINGS = {
    "setting_id": "default",
    "pricing_model": "per_chapter",
    "billing_cycles": [
        {"cycle": "monthly", "months": 1, "discount_percent": 0, "enabled": True},
        {"cycle": "quarterly", "months": 3, "discount_percent": 0, "enabled": True},
        {"cycle": "half_yearly", "months": 6, "discount_percent": 0, "enabled": True},
        {"cycle": "yearly", "months": 12, "discount_percent": 0, "enabled": True}
    ],
    "per_chapter_rate": 0,
    "slab_rates": [
        {"min_chapters": 1, "max_chapters": 3, "rate": 0},
        {"min_chapters": 4, "max_chapters": 10, "rate": 0},
        {"min_chapters": 11, "max_chapters": 25, "rate": 0},
        {"min_chapters": 26, "max_chapters": 9999, "rate": 0}
    ],
    "per_member_rate": 0,
    "free_trial": {
        "enabled": True,
        "duration_days": 30,
        "max_chapters": 1
    },
    "gst_percent": 18,
    "updated_at": datetime.now(timezone.utc).isoformat()
}

BILLING_CYCLE_MONTHS = {
    "monthly": 1,
    "quarterly": 3,
    "half_yearly": 6,
    "yearly": 12
}

@api_router.get("/developer/subscription-settings")
async def get_subscription_settings(user=Depends(require_role("developer"))):
    """Get current subscription settings, create defaults if not exist"""
    settings = await db.subscription_settings.find_one({"setting_id": "default"}, {"_id": 0})
    if not settings:
        await db.subscription_settings.insert_one({**DEFAULT_SUBSCRIPTION_SETTINGS})
        settings = {k: v for k, v in DEFAULT_SUBSCRIPTION_SETTINGS.items()}
    return settings

@api_router.put("/developer/subscription-settings")
async def update_subscription_settings(data: SubscriptionSettingsUpdate, user=Depends(require_role("developer"))):
    """Update pricing, billing cycles, discounts, free trial config"""
    update_data = {}
    for k, v in data.dict(exclude_unset=True).items():
        if v is not None:
            update_data[k] = v if not isinstance(v, list) else [item.dict() if hasattr(item, 'dict') else item for item in v]

    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")

    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()

    # Upsert settings
    await db.subscription_settings.update_one(
        {"setting_id": "default"},
        {"$set": update_data},
        upsert=True
    )
    settings = await db.subscription_settings.find_one({"setting_id": "default"}, {"_id": 0})
    return settings

@api_router.get("/developer/subscriptions")
async def list_subscriptions(user=Depends(require_role("developer"))):
    """List all ED subscriptions with ED details"""
    subs = await db.subscriptions.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)

    result = []
    for sub in subs:
        sa = await db.superadmins.find_one(
            {"$or": [{"superadmin_id": sub["superadmin_id"]}, {"mobile": sub["superadmin_id"]}]},
            {"_id": 0, "password_hash": 0}
        )
        sa_name = sa.get("name", "Unknown") if sa else "Unknown"
        sa_mobile = sa.get("mobile", "") if sa else ""

        # Count chapters used
        chapters_used = await db.chapters.count_documents({"created_by": sa_mobile}) if sa_mobile else 0

        result.append({
            **sub,
            "ed_name": sa_name,
            "ed_mobile": sa_mobile,
            "chapters_used": chapters_used
        })

    return result

@api_router.post("/developer/subscriptions/activate")
async def activate_subscription(data: SubscriptionActivate, user=Depends(require_role("developer"))):
    """Manually activate a subscription for an ED"""
    # Verify ED exists
    sa = await db.superadmins.find_one(
        {"$or": [{"superadmin_id": data.superadmin_id}, {"mobile": data.superadmin_id}]},
        {"_id": 0}
    )
    if not sa:
        raise HTTPException(status_code=404, detail="Super Admin not found")

    # Deactivate any existing active subscription
    await db.subscriptions.update_many(
        {"superadmin_id": data.superadmin_id, "status": "active"},
        {"$set": {"status": "expired"}}
    )

    # Calculate end date
    months = BILLING_CYCLE_MONTHS.get(data.billing_cycle, 1)
    start = datetime.now(timezone.utc)
    end = start + timedelta(days=months * 30)

    sub_id = str(_uuid4())
    sub_data = {
        "subscription_id": sub_id,
        "superadmin_id": data.superadmin_id,
        "plan_type": "paid",
        "billing_cycle": data.billing_cycle,
        "chapters_allowed": data.chapters_allowed,
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "amount_paid": data.amount_paid,
        "payment_method": data.payment_method,
        "payment_ref": data.payment_ref,
        "status": "active",
        "auto_renew": False,
        "created_at": start.isoformat()
    }
    await db.subscriptions.insert_one(sub_data)

    # Create recharge history entry
    settings = await db.subscription_settings.find_one({"setting_id": "default"}, {"_id": 0})
    gst_pct = settings.get("gst_percent", 18) if settings else 18
    gst_amount = round(data.amount_paid * gst_pct / (100 + gst_pct), 2)

    recharge = {
        "recharge_id": str(_uuid4()),
        "superadmin_id": data.superadmin_id,
        "subscription_id": sub_id,
        "amount": round(data.amount_paid - gst_amount, 2),
        "gst_amount": gst_amount,
        "total_amount": data.amount_paid,
        "chapters_count": data.chapters_allowed,
        "billing_cycle": data.billing_cycle,
        "payment_method": data.payment_method,
        "payment_ref": data.payment_ref,
        "status": "success",
        "created_at": start.isoformat()
    }
    await db.recharge_history.insert_one(recharge)

    return {"message": "Subscription activated", "subscription_id": sub_id}

@api_router.post("/developer/subscriptions/extend")
async def extend_subscription(data: SubscriptionExtend, user=Depends(require_role("developer"))):
    """Extend an existing subscription"""
    sub = await db.subscriptions.find_one({"subscription_id": data.subscription_id}, {"_id": 0})
    if not sub:
        raise HTTPException(status_code=404, detail="Subscription not found")

    # Calculate new end date from current end_date
    current_end = datetime.fromisoformat(sub["end_date"]) if isinstance(sub["end_date"], str) else sub["end_date"]
    if current_end < datetime.now(timezone.utc):
        current_end = datetime.now(timezone.utc)

    new_end = current_end + timedelta(days=data.additional_months * 30)

    await db.subscriptions.update_one(
        {"subscription_id": data.subscription_id},
        {"$set": {"end_date": new_end.isoformat(), "status": "active"}}
    )

    # Recharge history
    recharge = {
        "recharge_id": str(_uuid4()),
        "superadmin_id": sub["superadmin_id"],
        "subscription_id": data.subscription_id,
        "amount": data.amount_paid,
        "gst_amount": 0,
        "total_amount": data.amount_paid,
        "chapters_count": sub.get("chapters_allowed", 0),
        "billing_cycle": sub.get("billing_cycle", ""),
        "payment_method": "manual",
        "payment_ref": data.payment_ref,
        "status": "success",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.recharge_history.insert_one(recharge)

    return {"message": "Subscription extended", "new_end_date": new_end.isoformat()}

@api_router.post("/developer/subscriptions/cancel")
async def cancel_subscription(data: SubscriptionCancel, user=Depends(require_role("developer"))):
    """Cancel/deactivate a subscription"""
    result = await db.subscriptions.update_one(
        {"subscription_id": data.subscription_id},
        {"$set": {"status": "cancelled"}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Subscription not found")

    return {"message": "Subscription cancelled"}

# ===== SUPERADMIN SUBSCRIPTION + DASHBOARD ENDPOINTS =====

@api_router.get("/superadmin/my-subscription")
async def get_my_subscription(user=Depends(get_current_user)):
    """Returns current subscription for the logged-in ED"""
    if user["role"] not in ("superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    mobile = user.get("mobile", "")
    sa = await db.superadmins.find_one({"mobile": mobile}, {"_id": 0})
    if not sa:
        return {"subscription": None, "days_remaining": 0, "chapters_used": 0, "chapters_allowed": 0}

    sa_id = sa.get("superadmin_id", mobile)

    sub = await db.subscriptions.find_one(
        {"superadmin_id": sa_id, "status": "active"},
        {"_id": 0},
        sort=[("end_date", -1)]
    )

    if not sub:
        # Check for any subscription (expired/cancelled)
        sub = await db.subscriptions.find_one(
            {"superadmin_id": sa_id},
            {"_id": 0},
            sort=[("end_date", -1)]
        )

    chapters_used = await db.chapters.count_documents({"created_by": mobile})
    chapters_allowed = sub.get("chapters_allowed", 0) if sub else 0

    if sub:
        end_dt = datetime.fromisoformat(sub["end_date"]) if isinstance(sub["end_date"], str) else sub["end_date"]
        days_remaining = max(0, (end_dt - datetime.now(timezone.utc)).days)

        # Auto-expire if past end date
        if days_remaining == 0 and sub.get("status") == "active":
            await db.subscriptions.update_one(
                {"subscription_id": sub["subscription_id"]},
                {"$set": {"status": "expired"}}
            )
            sub["status"] = "expired"
    else:
        days_remaining = 0

    return {
        "subscription": sub,
        "days_remaining": days_remaining,
        "chapters_used": chapters_used,
        "chapters_allowed": chapters_allowed
    }

@api_router.get("/superadmin/dashboard/stats")
async def superadmin_dashboard_stats(user=Depends(get_current_user)):
    """Dashboard stats for SuperAdmin: my chapters, members, active/inactive, collection"""
    if user["role"] not in ("superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    mobile = user.get("mobile", "")
    my_chapters = await db.chapters.find({"created_by": mobile}, {"_id": 0}).to_list(1000)
    chapter_ids = [ch["chapter_id"] for ch in my_chapters]

    total_chapters = len(my_chapters)
    active_chapters = sum(1 for ch in my_chapters if ch.get("status", "active") == "active")
    inactive_chapters = total_chapters - active_chapters

    total_members = await db.members.count_documents({"chapter_id": {"$in": chapter_ids}, "status": "Active"}) if chapter_ids else 0
    pending_members = await db.members.count_documents({"chapter_id": {"$in": chapter_ids}, "membership_status": "pending"}) if chapter_ids else 0

    # Fund collection summary (this month)
    now = datetime.now(IST)
    current_month = now.month
    current_year = now.year

    kitty_pipeline = [
        {"$match": {"chapter_id": {"$in": chapter_ids}, "status": "paid", "month": current_month, "year": current_year}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    kitty_sum = await db.kitty_payments.aggregate(kitty_pipeline).to_list(1) if chapter_ids else []

    mf_pipeline = [
        {"$match": {"chapter_id": {"$in": chapter_ids}, "status": "paid", "month": current_month, "year": current_year}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    mf_sum = await db.meetingfee_payments.aggregate(mf_pipeline).to_list(1) if chapter_ids else []

    this_month_collection = (
        (kitty_sum[0]["total"] if kitty_sum else 0) +
        (mf_sum[0]["total"] if mf_sum else 0)
    )

    return {
        "total_chapters": total_chapters,
        "active_chapters": active_chapters,
        "inactive_chapters": inactive_chapters,
        "pending_members": pending_members,
        "total_members": total_members,
        "this_month_collection": this_month_collection
    }

@api_router.get("/superadmin/chapters/overview")
async def superadmin_chapters_overview(user=Depends(get_current_user)):
    """All chapters with member_count, last_meeting_date, fund_status"""
    if user["role"] not in ("superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    mobile = user.get("mobile", "")
    chapters = await db.chapters.find({"created_by": mobile}, {"_id": 0, "admin_password_hash": 0}).to_list(1000)

    result = []
    for ch in chapters:
        cid = ch["chapter_id"]
        member_count = await db.members.count_documents({"chapter_id": cid, "status": "Active"})

        # Last meeting
        last_meeting = await db.meetings.find_one(
            {"chapter_id": cid},
            {"_id": 0, "date": 1, "meeting_id": 1},
            sort=[("date", -1)]
        )

        # Admin info
        admin_name = ch.get("admin_mobile", "")

        result.append({
            **ch,
            "member_count": member_count,
            "last_meeting_date": last_meeting.get("date", "") if last_meeting else "",
            "admin_name": admin_name
        })

    return result

app.include_router(api_router)

# ===== MOUNT PAYMENT SYSTEM ROUTE MODULES =====
from routes.member_auth import router as member_auth_router
from routes.payment_config import router as payment_config_router
from routes.fee_ledger import router as fee_ledger_router
from routes.member_portal import router as member_portal_router
from routes.admin_verification import router as admin_verification_router
app.include_router(member_auth_router)
app.include_router(payment_config_router)
app.include_router(fee_ledger_router)
app.include_router(member_portal_router)
app.include_router(admin_verification_router)

# Serve uploaded files (payment proofs, etc.)
import os as _os
_uploads_dir = _os.path.join(_os.path.dirname(__file__), "uploads")
_os.makedirs(_uploads_dir, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=_uploads_dir), name="uploads")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup():
    # Initialize super admin if not exists
    existing = await db.superadmins.find_one({"mobile": "919893452545"})
    if not existing:
        await db.superadmins.insert_one({
            "mobile": "919893452545",
            "password_hash": hash_password("superadmin123@"),
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info("Super admin created")

    # Seed default subscription settings if not exists
    existing_settings = await db.subscription_settings.find_one({"setting_id": "default"})
    if not existing_settings:
        await db.subscription_settings.insert_one({**DEFAULT_SUBSCRIPTION_SETTINGS})
        logger.info("Default subscription settings seeded")

    # ===== Phase 1.3 Migration: Enhanced Member Fields =====
    migrated = await db.members.update_many(
        {"membership_status": {"$exists": False}},
        {"$set": {
            "membership_status": "active",
            "email": None,
            "business_name": None,
            "business_category": None,
            "joining_date": None,
            "renewal_date": None,
            "induction_fee": None,
            "status_history": [],
            "archived": False,
            "transfer_from_chapter": None,
            "transfer_date": None,
        }}
    )
    if migrated.modified_count > 0:
        logger.info(f"Migrated {migrated.modified_count} members with enhanced fields")

    # Create indexes for member search/filter
    await db.members.create_index([("chapter_id", 1), ("membership_status", 1)])
    await db.members.create_index([("chapter_id", 1), ("status", 1)])
    await db.members.create_index([("membership_status", 1)])
    logger.info("Member indexes ensured")

    # ===== Payment System Indexes =====
    await db.member_credentials.create_index("mobile", unique=True)
    await db.member_credentials.create_index("member_id", unique=True)
    await db.payment_config.create_index("superadmin_id", unique=True)
    await db.chapter_fee_config.create_index("chapter_id", unique=True)
    await db.fee_ledger.create_index([("chapter_id", 1), ("member_id", 1), ("status", 1)])
    await db.fee_ledger.create_index([("chapter_id", 1), ("status", 1), ("fee_type", 1)])
    await db.fee_ledger.create_index([("member_id", 1), ("status", 1)])
    await db.accountant_credentials.create_index("mobile", unique=True)
    logger.info("Payment system indexes ensured")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()