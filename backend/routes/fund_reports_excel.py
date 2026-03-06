# MAX 400 LINES - Fund report Excel export
"""Fund report Excel export with date filter, category filter, payment status filter."""

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from datetime import datetime
from typing import Optional
from database import db
from deps import get_current_user
import pytz
import io

IST = pytz.timezone('Asia/Kolkata')

router = APIRouter(prefix="/api", tags=["fund-reports-excel"])


# Fund Report Export - Excel with Date Filter
@router.get("/admin/fund/reports/export/excel")
async def export_fund_report_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    months: Optional[str] = None,
    year: Optional[int] = None,
    category: Optional[str] = None,
    categories: Optional[str] = None,
    payment_status: Optional[str] = None,  # all, paid, pending
    event_id: Optional[str] = None,  # Filter by specific event
    user = Depends(get_current_user)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    chapter_id = user.get("chapter_id")
    chapter_name = user.get("chapter_name", "Chapter")

    # Parse date filters
    current_year = datetime.now().year
    current_month = datetime.now().month
    filter_year = year or current_year

    # Month-year list: Store tuples of (month, year) for proper cross-year handling
    month_year_list = []

    if months:
        # User specified months - use with filter_year, validate range 1-12
        month_list = [int(m.strip()) for m in months.split(",") if m.strip().isdigit()]
        month_year_list = [(m, filter_year) for m in month_list if 1 <= m <= 12]
    else:
        # Default to last 3 months with proper year handling
        for i in range(3):
            m = current_month - i
            y = current_year
            if m <= 0:
                m += 12
                y -= 1
            month_year_list.append((m, y))
        month_year_list.reverse()  # Oldest first

    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    # Build filter description with year for each month
    filter_parts = []
    for m, y in month_year_list:
        filter_parts.append(f"{month_names[m-1]}'{str(y)[-2:]}")
    filter_desc = ', '.join(filter_parts)

    # Categories
    cat_list = []
    if categories:
        cat_list = [c.strip() for c in categories.split(',')]
    elif category and category != 'all':
        cat_list = [category]

    include_kitty = len(cat_list) == 0 or 'kitty' in cat_list
    include_meetingfee = len(cat_list) == 0 or 'meetingfee' in cat_list
    include_events = len(cat_list) == 0 or 'events' in cat_list

    category_desc = ""
    if cat_list:
        cat_names = {'kitty': 'Kitty', 'meetingfee': 'Meeting Fee', 'events': 'Events'}
        category_desc = f" - {', '.join([cat_names.get(c, c) for c in cat_list])}"

    status_desc = ""
    if payment_status and payment_status != 'all':
        status_desc = f" ({payment_status.capitalize()} Only)"

    # Get all members
    members = await db.members.find(
        {"chapter_id": chapter_id, "status": "Active"},
        {"_id": 0}
    ).to_list(500)

    # Pre-fetch all payments
    all_kitty = await db.kitty_payments.find(
        {"chapter_id": chapter_id, "status": "paid"},
        {"_id": 0}
    ).to_list(5000)

    all_kitty_settings = await db.kitty_settings.find(
        {"chapter_id": chapter_id},
        {"_id": 0}
    ).to_list(100)
    kitty_setting_map = {(s["month"], s["year"]): s.get("amount", 0) for s in all_kitty_settings}

    all_meetingfee = await db.meetingfee_payments.find(
        {"chapter_id": chapter_id, "status": "paid"},
        {"_id": 0}
    ).to_list(5000)

    all_mf_settings = await db.meetingfee_settings.find(
        {"chapter_id": chapter_id},
        {"_id": 0}
    ).to_list(100)
    mf_setting_map = {(s["month"], s["year"]): s.get("amount", 0) for s in all_mf_settings}

    # Get events
    all_events = await db.fund_events.find(
        {"chapter_id": chapter_id},
        {"_id": 0}
    ).to_list(100)
    events_map = {e["event_id"]: e for e in all_events}

    all_event_payments = await db.event_payments.find(
        {"chapter_id": chapter_id, "status": "paid"} if chapter_id else {"status": "paid"},
        {"_id": 0}
    ).to_list(5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Fund Report"

    # Styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)

    # Title
    ws['A1'] = f"FUND REPORT - {chapter_name}{category_desc}{status_desc}"
    ws['A1'].font = title_font
    ws['A1'].fill = yellow_fill
    ws['A2'] = f"Period: {filter_desc}"
    ws['A3'] = f"Generated: {datetime.now(IST).strftime('%d-%b-%Y %H:%M')}"

    # Build dynamic headers (Point 2: Month-wise columns)
    headers = ["Sr", "Member ID", "Member Name"]

    # Add month-wise columns for each category with year
    if include_kitty:
        for m, y in month_year_list:
            headers.append(f"K-{month_names[m-1]}'{str(y)[-2:]}")
        headers.append("Kitty Total")

    if include_meetingfee:
        for m, y in month_year_list:
            headers.append(f"MF-{month_names[m-1]}'{str(y)[-2:]}")
        headers.append("M.Fee Total")
        headers.append("Pay Mode")

    if include_events:
        # Add event columns
        for event in all_events:
            if not event_id or event["event_id"] == event_id:
                headers.append(f"Event:{event.get('name', 'Unknown')[:10]}")
        headers.append("Event Total")

    headers.append("Grand Total")
    headers.append("Status")  # Paid/Pending status

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    row = 6
    grand_totals = {
        "kitty_months": {(m, y): 0 for m, y in month_year_list},
        "kitty_total": 0,
        "mf_months": {(m, y): 0 for m, y in month_year_list},
        "mf_total": 0,
        "event_total": 0,
        "grand": 0
    }

    # Summary tracking
    summary_data = {
        "total_members": len(members),
        "paid_members": 0,
        "pending_members": 0,
        "kitty_pending_months": {(m, y): 0 for m, y in month_year_list},
        "kitty_pending_total": 0,
        "mf_pending_months": {(m, y): 0 for m, y in month_year_list},
        "mf_pending_total": 0,
        "event_pending_total": 0,
        "total_pending": 0
    }

    for idx, member in enumerate(members, 1):
        member_id = member["member_id"]
        member_data = {"kitty": {}, "meetingfee": {}, "events": {}}
        member_total = 0
        member_pending = 0

        # Calculate Kitty month-wise with proper year
        if include_kitty:
            for m, y in month_year_list:
                kitty_paid = [k for k in all_kitty if k.get("member_id") == member_id and k.get("month") == m and k.get("year") == y]
                paid_amt = sum(k.get("amount", 0) for k in kitty_paid)
                expected_amt = kitty_setting_map.get((m, y), 0)
                member_data["kitty"][(m, y)] = {"paid": paid_amt, "expected": expected_amt}
                if paid_amt == 0 and expected_amt > 0:
                    member_pending += expected_amt
                    summary_data["kitty_pending_months"][(m, y)] += expected_amt
                    summary_data["kitty_pending_total"] += expected_amt

        # Calculate Meeting Fee month-wise with proper year
        mf_payment_mode = "-"
        if include_meetingfee:
            for m, y in month_year_list:
                mf_paid = [mf for mf in all_meetingfee if mf.get("member_id") == member_id and mf.get("month") == m and mf.get("year") == y]
                paid_amt = sum(mf.get("amount", 0) for mf in mf_paid)
                expected_amt = mf_setting_map.get((m, y), 0)
                member_data["meetingfee"][(m, y)] = {"paid": paid_amt, "expected": expected_amt}
                if mf_paid and mf_paid[0].get("payment_mode"):
                    mf_payment_mode = mf_paid[0]["payment_mode"]
                if paid_amt == 0 and expected_amt > 0:
                    member_pending += expected_amt
                    summary_data["mf_pending_months"][(m, y)] += expected_amt
                    summary_data["mf_pending_total"] += expected_amt

        # Calculate Events
        if include_events:
            for event in all_events:
                if not event_id or event["event_id"] == event_id:
                    ev_paid = [ep for ep in all_event_payments if ep.get("member_id") == member_id and ep.get("event_id") == event["event_id"]]
                    paid_amt = event.get("amount", 0) if ev_paid else 0
                    member_data["events"][event["event_id"]] = {"paid": paid_amt, "expected": event.get("amount", 0)}
                    if not ev_paid:
                        member_pending += event.get("amount", 0)
                        summary_data["event_pending_total"] += event.get("amount", 0)

        # Calculate totals
        kitty_total = sum(d["paid"] for d in member_data["kitty"].values())
        mf_total = sum(d["paid"] for d in member_data["meetingfee"].values())
        event_total = sum(d["paid"] for d in member_data["events"].values())
        member_total = kitty_total + mf_total + event_total

        # Track member paid/pending status for summary
        if member_pending > 0:
            summary_data["pending_members"] += 1
            summary_data["total_pending"] += member_pending
        elif member_total > 0:
            summary_data["paid_members"] += 1

        # Payment status filter (Points 3, 4, 5, 6)
        is_fully_paid = member_pending == 0 and member_total > 0
        is_pending = member_pending > 0

        if payment_status == 'paid' and not is_fully_paid:
            continue
        if payment_status == 'pending' and not is_pending:
            continue

        # Write row
        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = member.get("unique_member_id", "")
        ws.cell(row=row, column=3).value = member.get("full_name", "")

        col = 4

        # Kitty month-wise (Point 7)
        if include_kitty:
            for m, y in month_year_list:
                data = member_data["kitty"].get((m, y), {"paid": 0})
                cell = ws.cell(row=row, column=col)
                cell.value = data["paid"]
                if data["paid"] == 0 and data.get("expected", 0) > 0:
                    cell.fill = red_fill  # Highlight pending
                elif data["paid"] > 0:
                    cell.fill = green_fill  # Highlight paid
                grand_totals["kitty_months"][(m, y)] += data["paid"]
                col += 1
            ws.cell(row=row, column=col).value = kitty_total
            grand_totals["kitty_total"] += kitty_total
            col += 1

        # Meeting Fee month-wise
        if include_meetingfee:
            for m, y in month_year_list:
                data = member_data["meetingfee"].get((m, y), {"paid": 0})
                cell = ws.cell(row=row, column=col)
                cell.value = data["paid"]
                if data["paid"] == 0 and data.get("expected", 0) > 0:
                    cell.fill = red_fill
                elif data["paid"] > 0:
                    cell.fill = green_fill
                grand_totals["mf_months"][(m, y)] += data["paid"]
                col += 1
            ws.cell(row=row, column=col).value = mf_total
            grand_totals["mf_total"] += mf_total
            col += 1
            ws.cell(row=row, column=col).value = mf_payment_mode
            col += 1

        # Events (Point 8)
        if include_events:
            for event in all_events:
                if not event_id or event["event_id"] == event_id:
                    data = member_data["events"].get(event["event_id"], {"paid": 0})
                    cell = ws.cell(row=row, column=col)
                    cell.value = data["paid"]
                    if data["paid"] == 0 and data.get("expected", 0) > 0:
                        cell.fill = red_fill
                    elif data["paid"] > 0:
                        cell.fill = green_fill
                    col += 1
            ws.cell(row=row, column=col).value = event_total
            grand_totals["event_total"] += event_total
            col += 1

        # Grand Total
        ws.cell(row=row, column=col).value = member_total
        grand_totals["grand"] += member_total
        col += 1

        # Status
        status_cell = ws.cell(row=row, column=col)
        if member_pending > 0:
            status_cell.value = f"Pending: Rs.{member_pending}"
            status_cell.fill = red_fill
        elif member_total > 0:
            status_cell.value = "Paid"
            status_cell.fill = green_fill
        else:
            status_cell.value = "No Activity"

        row += 1

    # Total row
    row += 1
    ws.cell(row=row, column=2).value = "TOTAL"
    ws.cell(row=row, column=2).font = header_font

    col = 4
    if include_kitty:
        for m, y in month_year_list:
            ws.cell(row=row, column=col).value = grand_totals["kitty_months"][(m, y)]
            ws.cell(row=row, column=col).font = header_font
            col += 1
        ws.cell(row=row, column=col).value = grand_totals["kitty_total"]
        ws.cell(row=row, column=col).font = header_font
        ws.cell(row=row, column=col).fill = yellow_fill
        col += 1

    if include_meetingfee:
        for m, y in month_year_list:
            ws.cell(row=row, column=col).value = grand_totals["mf_months"][(m, y)]
            ws.cell(row=row, column=col).font = header_font
            col += 1
        ws.cell(row=row, column=col).value = grand_totals["mf_total"]
        ws.cell(row=row, column=col).font = header_font
        ws.cell(row=row, column=col).fill = yellow_fill
        col += 1
        col += 1  # Skip mode column

    if include_events:
        col += len([e for e in all_events if not event_id or e["event_id"] == event_id])
        ws.cell(row=row, column=col).value = grand_totals["event_total"]
        ws.cell(row=row, column=col).font = header_font
        ws.cell(row=row, column=col).fill = yellow_fill
        col += 1

    ws.cell(row=row, column=col).value = grand_totals["grand"]
    ws.cell(row=row, column=col).font = header_font
    ws.cell(row=row, column=col).fill = yellow_fill

    # Adjust column widths
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20

    # ========== SUMMARY SECTION ==========
    row += 3
    summary_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

    # Summary Title
    ws.cell(row=row, column=1).value = "═══════ SUMMARY ═══════"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 2

    # Member Count Summary
    ws.cell(row=row, column=1).value = "MEMBER STATUS"
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).fill = summary_fill
    row += 1
    ws.cell(row=row, column=1).value = f"Total Members: {summary_data['total_members']}"
    row += 1
    ws.cell(row=row, column=1).value = f"Fully Paid Members: {summary_data['paid_members']}"
    ws.cell(row=row, column=1).fill = green_fill
    row += 1
    ws.cell(row=row, column=1).value = f"Pending Members: {summary_data['pending_members']}"
    ws.cell(row=row, column=1).fill = red_fill
    row += 2

    # Category-wise Summary
    ws.cell(row=row, column=1).value = "CATEGORY-WISE COLLECTION"
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).fill = summary_fill
    row += 1

    if include_kitty:
        ws.cell(row=row, column=1).value = f"KITTY - Collected: Rs.{grand_totals['kitty_total']} | Pending: Rs.{summary_data['kitty_pending_total']}"
        row += 1
    if include_meetingfee:
        ws.cell(row=row, column=1).value = f"MEETING FEE - Collected: Rs.{grand_totals['mf_total']} | Pending: Rs.{summary_data['mf_pending_total']}"
        row += 1
    if include_events:
        ws.cell(row=row, column=1).value = f"EVENTS - Collected: Rs.{grand_totals['event_total']} | Pending: Rs.{summary_data['event_pending_total']}"
        row += 1

    row += 1
    ws.cell(row=row, column=1).value = f"GRAND TOTAL - Collected: Rs.{grand_totals['grand']} | Pending: Rs.{summary_data['total_pending']}"
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).fill = yellow_fill
    row += 2

    # Month-wise Summary
    ws.cell(row=row, column=1).value = "MONTH-WISE BREAKDOWN"
    ws.cell(row=row, column=1).font = header_font
    ws.cell(row=row, column=1).fill = summary_fill
    row += 1

    for m, y in month_year_list:
        month_collected = grand_totals["kitty_months"].get((m, y), 0) + grand_totals["mf_months"].get((m, y), 0)
        month_pending = summary_data["kitty_pending_months"].get((m, y), 0) + summary_data["mf_pending_months"].get((m, y), 0)
        ws.cell(row=row, column=1).value = f"{month_names[m-1]} {y}: Collected Rs.{month_collected} | Pending Rs.{month_pending}"
        row += 1

    # Save
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=fund_report_{chapter_id}.xlsx"}
    )
