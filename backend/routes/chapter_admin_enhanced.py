# MAX 400 LINES - Enhanced member management endpoints
"""Chapter Admin enhanced member management: stats, expiring, export, profile, status, archive."""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone, timedelta
from database import db
from deps import get_current_user
from models import MemberStatusChange, BulkMemberStatus
import io

router = APIRouter(prefix="/api", tags=["chapter-admin-enhanced"])


def _sync_status(membership_status: str) -> str:
    """Return legacy status field value based on membership_status."""
    return "Active" if membership_status == "active" else "Inactive"


@router.get("/admin/members/stats")
async def get_member_stats(user=Depends(get_current_user)):
    """Return member counts by status for the chapter."""
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    chapter_id = user.get("chapter_id")
    if not chapter_id and user["role"] == "superadmin":
        sa = await db.superadmins.find_one({"mobile": user.get("mobile")}, {"_id": 0})
        sa_id = sa.get("superadmin_id", user.get("mobile")) if sa else user.get("mobile")
        chapters = await db.chapters.find({"created_by": sa_id}, {"chapter_id": 1}).to_list(100)
        chapter_ids = [c["chapter_id"] for c in chapters]
        base_filter = {"chapter_id": {"$in": chapter_ids}}
    elif chapter_id:
        base_filter = {"chapter_id": chapter_id}
    else:
        base_filter = {}

    total = await db.members.count_documents({**base_filter, "archived": {"$ne": True}})
    active = await db.members.count_documents({**base_filter, "membership_status": "active", "archived": {"$ne": True}})
    pending = await db.members.count_documents({**base_filter, "membership_status": "pending"})
    inactive = await db.members.count_documents({**base_filter, "membership_status": "inactive", "archived": {"$ne": True}})
    suspended = await db.members.count_documents({**base_filter, "membership_status": "suspended"})

    from datetime import date
    today_str = date.today().isoformat()
    thirty_days = (date.today() + timedelta(days=30)).isoformat()
    expiring_soon = await db.members.count_documents({
        **base_filter,
        "membership_status": "active",
        "renewal_date": {"$ne": None, "$lte": thirty_days, "$gte": today_str}
    })

    return {
        "total": total, "active": active, "pending": pending,
        "inactive": inactive, "suspended": suspended, "expiring_soon": expiring_soon
    }


@router.get("/admin/members/expiring")
async def get_expiring_members(user=Depends(get_current_user)):
    """Members with renewal_date within next 30 days."""
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    chapter_id = user.get("chapter_id")
    if not chapter_id:
        raise HTTPException(status_code=400, detail="Chapter ID required")

    from datetime import date
    today_str = date.today().isoformat()
    thirty_days = (date.today() + timedelta(days=30)).isoformat()

    members = await db.members.find({
        "chapter_id": chapter_id,
        "membership_status": "active",
        "renewal_date": {"$ne": None, "$lte": thirty_days, "$gte": today_str}
    }, {"_id": 0}).to_list(500)
    return members


@router.get("/admin/members/export")
async def export_members(user=Depends(get_current_user)):
    """Export member list as Excel."""
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    chapter_id = user.get("chapter_id")
    if not chapter_id:
        raise HTTPException(status_code=400, detail="Chapter ID required")

    members = await db.members.find(
        {"chapter_id": chapter_id, "archived": {"$ne": True}},
        {"_id": 0}
    ).sort("unique_member_id", 1).to_list(2000)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Members"

    headers = [
        "Member ID", "Full Name", "Primary Mobile", "Secondary Mobile",
        "Email", "Business Name", "Business Category",
        "Joining Date", "Renewal Date", "Induction Fee",
        "Membership Status", "Status"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="CF2030", end_color="CF2030", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for m in members:
        ws.append([
            m.get("unique_member_id", ""), m.get("full_name", ""),
            m.get("primary_mobile", ""), m.get("secondary_mobile", ""),
            m.get("email", ""), m.get("business_name", ""),
            m.get("business_category", ""), m.get("joining_date", ""),
            m.get("renewal_date", ""), m.get("induction_fee", ""),
            m.get("membership_status", "active"), m.get("status", "Active"),
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=members_export.xlsx"}
    )


@router.get("/admin/members/{member_id}/profile")
async def get_member_profile(member_id: str, user=Depends(get_current_user)):
    """Full member profile with attendance & payment history."""
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    member = await db.members.find_one({"member_id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    attendance = await db.attendance.find(
        {"unique_member_id": member.get("unique_member_id"), "meeting_id": {"$exists": True}},
        {"_id": 0}
    ).sort("timestamp", -1).to_list(20)

    for att in attendance:
        meeting = await db.meetings.find_one({"meeting_id": att["meeting_id"]}, {"_id": 0, "date": 1})
        att["meeting_date"] = meeting.get("date") if meeting else None

    all_att = await db.attendance.find(
        {"unique_member_id": member.get("unique_member_id"), "type": "member"},
        {"_id": 0, "status": 1, "late_type": 1}
    ).to_list(5000)
    total_meetings_attended = len(all_att)
    present_count = sum(1 for a in all_att if a.get("status") == "present")
    late_count = sum(1 for a in all_att if a.get("late_type") in ("late", "very_late"))

    kitty_payments = await db.kitty_payments.find(
        {"member_id": member["member_id"]}, {"_id": 0}
    ).sort("year", -1).to_list(20)

    meetingfee_payments = await db.meetingfee_payments.find(
        {"member_id": member["member_id"]}, {"_id": 0}
    ).sort("year", -1).to_list(20)

    misc_records = await db.misc_payment_records.find(
        {"member_id": member["member_id"]}, {"_id": 0}
    ).to_list(20)

    event_payments = await db.event_payments.find(
        {"member_id": member["member_id"]}, {"_id": 0}
    ).to_list(20)

    chapter = await db.chapters.find_one({"chapter_id": member.get("chapter_id")}, {"_id": 0, "name": 1})

    return {
        "member": member,
        "chapter_name": chapter.get("name") if chapter else "",
        "attendance": {
            "recent": attendance, "total_attended": total_meetings_attended,
            "present_count": present_count, "late_count": late_count,
        },
        "payments": {
            "kitty": kitty_payments, "meeting_fee": meetingfee_payments,
            "misc": misc_records, "events": event_payments,
        }
    }


@router.post("/admin/members/{member_id}/status")
async def change_member_status(member_id: str, data: MemberStatusChange, user=Depends(get_current_user)):
    """Change member status: deactivate, suspend, reactivate."""
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    member = await db.members.find_one({"member_id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    action_map = {"deactivate": "inactive", "suspend": "suspended", "reactivate": "active"}
    new_ms = action_map.get(data.action)
    if not new_ms:
        raise HTTPException(status_code=400, detail=f"Invalid action: {data.action}")

    old_ms = member.get("membership_status", "active")
    history_entry = {
        "action": data.action, "from_status": old_ms, "to_status": new_ms,
        "reason": data.reason,
        "changed_by": user.get("mobile", user.get("email", "system")),
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

    await db.members.update_one(
        {"member_id": member_id},
        {"$set": {"membership_status": new_ms, "status": _sync_status(new_ms)},
         "$push": {"status_history": history_entry}}
    )

    return {"message": f"Member {data.action}d successfully", "new_status": new_ms}


@router.post("/admin/members/bulk-status")
async def bulk_change_status(data: BulkMemberStatus, user=Depends(get_current_user)):
    """Bulk activate/deactivate members."""
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    action_map = {"activate": "active", "deactivate": "inactive"}
    new_ms = action_map.get(data.action)
    if not new_ms:
        raise HTTPException(status_code=400, detail="Action must be activate or deactivate")

    history_entry = {
        "action": data.action, "from_status": "bulk", "to_status": new_ms,
        "reason": data.reason,
        "changed_by": user.get("mobile", user.get("email", "system")),
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

    result = await db.members.update_many(
        {"member_id": {"$in": data.member_ids}},
        {"$set": {"membership_status": new_ms, "status": _sync_status(new_ms)},
         "$push": {"status_history": history_entry}}
    )

    return {"message": f"{result.modified_count} members updated", "new_status": new_ms}


@router.post("/admin/members/auto-archive")
async def auto_archive_members(user=Depends(get_current_user)):
    """Archive members inactive for 6+ months."""
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    chapter_id = user.get("chapter_id")
    if not chapter_id:
        raise HTTPException(status_code=400, detail="Chapter ID required")

    six_months_ago = (datetime.now(timezone.utc) - timedelta(days=180)).isoformat()

    inactive_members = await db.members.find({
        "chapter_id": chapter_id,
        "membership_status": "inactive",
        "archived": {"$ne": True},
    }, {"_id": 0, "member_id": 1, "status_history": 1}).to_list(1000)

    archived_ids = []
    for m in inactive_members:
        history = m.get("status_history", [])
        if history:
            last_entry = history[-1]
            if last_entry.get("timestamp", "") < six_months_ago:
                archived_ids.append(m["member_id"])
        else:
            archived_ids.append(m["member_id"])

    if archived_ids:
        await db.members.update_many(
            {"member_id": {"$in": archived_ids}},
            {"$set": {"archived": True}}
        )

    return {"message": f"{len(archived_ids)} members archived"}
