"""Accountant reports: summary, ledger, export."""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
from typing import Optional
import io

from database import db
from deps import require_role

router = APIRouter(prefix="/api", tags=["accountant-reports"])


async def _get_sa_chapter_ids(user):
    """Get chapter IDs for the accountant's superadmin."""
    sa_id = user.get("superadmin_id")
    if not sa_id:
        return []
    chapters = await db.chapters.find(
        {"superadmin_id": sa_id}, {"chapter_id": 1, "_id": 0}
    ).to_list(500)
    return [c["chapter_id"] for c in chapters]


@router.get("/accountant/reports/summary")
async def reports_summary(
    period: Optional[str] = Query("this_month"),
    user=Depends(require_role("accountant", "superadmin")),
):
    """Monthly/quarterly payment summary."""
    if user["role"] == "superadmin":
        mobile = user.get("mobile", "")
        sa = await db.superadmins.find_one({"mobile": mobile}, {"_id": 0})
        sa_id = sa.get("superadmin_id", mobile) if sa else mobile
        chapters = await db.chapters.find(
            {"$or": [{"created_by": sa_id}, {"created_by": mobile}, {"superadmin_id": sa_id}]},
            {"chapter_id": 1, "_id": 0}
        ).to_list(500)
        chapter_ids = list({c["chapter_id"] for c in chapters})
    else:
        chapter_ids = await _get_sa_chapter_ids(user)

    now = datetime.now(timezone.utc)
    if period == "this_month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
    elif period == "last_month":
        if now.month == 1:
            start = now.replace(year=now.year - 1, month=12, day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
        else:
            start = now.replace(month=now.month - 1, day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
    elif period == "this_quarter":
        q_month = ((now.month - 1) // 3) * 3 + 1
        start = now.replace(month=q_month, day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
    else:
        start = None

    query = {"chapter_id": {"$in": chapter_ids}}
    if start:
        query["created_at"] = {"$gte": start}

    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": "$status",
            "total": {"$sum": "$amount"},
            "count": {"$sum": 1},
        }}
    ]
    results = await db.fee_ledger.aggregate(pipeline).to_list(None)

    summary = {"total_collected": 0, "pending": 0, "verified": 0, "rejected": 0,
               "total_collected_count": 0, "pending_count": 0, "verified_count": 0, "rejected_count": 0}
    for r in results:
        s = r["_id"]
        if s == "verified":
            summary["total_collected"] = r["total"]
            summary["total_collected_count"] = r["count"]
        elif s in ("pending", "submitted"):
            summary["pending"] += r["total"]
            summary["pending_count"] += r["count"]
        elif s == "admin_confirmed":
            summary["verified"] = r["total"]
            summary["verified_count"] = r["count"]
        elif s == "rejected":
            summary["rejected"] = r["total"]
            summary["rejected_count"] = r["count"]

    return summary


@router.get("/accountant/reports/ledger")
async def reports_ledger(
    status: Optional[str] = Query(None),
    fee_type: Optional[str] = Query(None),
    chapter_id: Optional[str] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    user=Depends(require_role("accountant", "superadmin")),
):
    """Full payment ledger with filters."""
    if user["role"] == "superadmin":
        mobile = user.get("mobile", "")
        sa = await db.superadmins.find_one({"mobile": mobile}, {"_id": 0})
        sa_id = sa.get("superadmin_id", mobile) if sa else mobile
        chapters = await db.chapters.find(
            {"$or": [{"created_by": sa_id}, {"created_by": mobile}, {"superadmin_id": sa_id}]},
            {"chapter_id": 1, "_id": 0}
        ).to_list(500)
        all_chapter_ids = list({c["chapter_id"] for c in chapters})
    else:
        all_chapter_ids = await _get_sa_chapter_ids(user)

    query = {"chapter_id": {"$in": all_chapter_ids}}
    if status:
        query["status"] = status
    if fee_type:
        query["fee_type"] = fee_type
    if chapter_id:
        query["chapter_id"] = chapter_id
    if from_date:
        query.setdefault("created_at", {})["$gte"] = from_date
    if to_date:
        query.setdefault("created_at", {})["$lte"] = to_date

    total = await db.fee_ledger.count_documents(query)
    skip = (page - 1) * limit
    entries = await db.fee_ledger.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)

    # Enrich with chapter names
    ch_cache = {}
    for e in entries:
        cid = e.get("chapter_id")
        if cid not in ch_cache:
            ch = await db.chapters.find_one({"chapter_id": cid}, {"_id": 0, "name": 1})
            ch_cache[cid] = ch.get("name", "") if ch else ""
        e["chapter_name"] = ch_cache[cid]

    return {"entries": entries, "total": total, "page": page, "limit": limit}


@router.get("/accountant/reports/export")
async def export_report(
    format: str = Query("excel"),
    status: Optional[str] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    user=Depends(require_role("accountant", "superadmin")),
):
    """Export payment ledger as Excel or PDF."""
    if user["role"] == "superadmin":
        mobile = user.get("mobile", "")
        sa = await db.superadmins.find_one({"mobile": mobile}, {"_id": 0})
        sa_id = sa.get("superadmin_id", mobile) if sa else mobile
        chapters = await db.chapters.find(
            {"$or": [{"created_by": sa_id}, {"created_by": mobile}, {"superadmin_id": sa_id}]},
            {"chapter_id": 1, "_id": 0}
        ).to_list(500)
        all_chapter_ids = list({c["chapter_id"] for c in chapters})
    else:
        all_chapter_ids = await _get_sa_chapter_ids(user)

    query = {"chapter_id": {"$in": all_chapter_ids}}
    if status:
        query["status"] = status
    if from_date:
        query.setdefault("created_at", {})["$gte"] = from_date
    if to_date:
        query.setdefault("created_at", {})["$lte"] = to_date

    entries = await db.fee_ledger.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)

    if format == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Payment Ledger"
        headers = ["Date", "Member", "Chapter", "Fee Type", "Amount", "Status", "Payment Method", "UTR"]
        ws.append(headers)

        header_fill = PatternFill(start_color="CF2030", end_color="CF2030", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for e in entries:
            ws.append([
                e.get("payment_date") or e.get("created_at", "")[:10],
                e.get("member_name", ""),
                e.get("chapter_id", ""),
                e.get("fee_type", ""),
                e.get("amount", 0),
                e.get("status", ""),
                e.get("payment_method", ""),
                e.get("utr_number", ""),
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=payment_ledger.xlsx"}
        )

    return {"message": "PDF export coming soon"}
