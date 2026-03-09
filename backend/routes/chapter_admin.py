# MAX 400 LINES - Chapter Admin core endpoints
"""Chapter Admin endpoints: profile, member CRUD, bulk upload, template, excel upload.
Login is now handled via member_auth (President/VP get admin role automatically).
"""
from fastapi import APIRouter, HTTPException, Depends, File, UploadFile, Request
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone, timedelta
from database import db
from deps import get_current_user
from auth import verify_password, create_access_token, hash_password
from models import LoginRequest, LoginResponse, MemberCreate, MemberResponse, MemberUpdate, List
import io

router = APIRouter(prefix="/api", tags=["chapter-admin"])


async def _generate_member_id(chapter_id: str, chapter_code: str) -> str:
    """Generate next BNI-{CODE}-{NNN} member ID for a chapter."""
    prefix = f"BNI-{chapter_code}-"
    existing = await db.members.find(
        {"chapter_id": chapter_id, "unique_member_id": {"$regex": f"^{prefix}"}},
        {"_id": 0, "unique_member_id": 1}
    ).to_list(5000)
    max_num = 0
    for m in existing:
        try:
            num = int(m["unique_member_id"].split("-")[-1])
            if num > max_num:
                max_num = num
        except (ValueError, IndexError):
            pass
    return f"{prefix}{max_num + 1:03d}"


# NOTE: POST /admin/login has been removed.
# Chapter admins now login through member login — Presidents/VPs get admin role.


@router.get("/admin/profile")
async def get_admin_profile(user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    chapter = await db.chapters.find_one({"chapter_id": user.get("chapter_id")}, {"_id": 0, "admin_password_hash": 0})
    return {
        "mobile": user.get("mobile"), "chapter_id": user.get("chapter_id"),
        "chapter_name": chapter.get("name", "") if chapter else user.get("chapter_name", ""),
        "role": user.get("role")
    }


@router.get("/admin/members")
async def get_members(search: str = "", status_filter: str = "", category: str = "", sort_by: str = "unique_member_id", user=Depends(get_current_user)):
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    chapter_id = user.get("chapter_id")
    if not chapter_id:
        raise HTTPException(status_code=400, detail="Chapter ID required")

    query = {"chapter_id": chapter_id, "archived": {"$ne": True}}
    if status_filter:
        query["membership_status"] = status_filter
    if category:
        query["business_category"] = {"$regex": category, "$options": "i"}
    if search:
        query["$or"] = [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"primary_mobile": {"$regex": search, "$options": "i"}},
            {"business_name": {"$regex": search, "$options": "i"}},
            {"unique_member_id": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
        ]

    sort_field = sort_by if sort_by in ("full_name", "unique_member_id", "created_at", "renewal_date") else "unique_member_id"
    members = await db.members.find(query, {"_id": 0}).sort(sort_field, 1).to_list(2000)
    return members


@router.post("/admin/members", response_model=MemberResponse)
async def add_member(member: MemberCreate, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    chapter_id = user.get("chapter_id")
    if user["role"] == "superadmin" and not chapter_id:
        sa = await db.superadmins.find_one({"mobile": user.get("mobile")}, {"_id": 0})
        sa_id = sa.get("superadmin_id", user.get("mobile")) if sa else user.get("mobile")
        first_ch = await db.chapters.find_one({"created_by": sa_id}, {"_id": 0, "chapter_id": 1})
        if first_ch:
            chapter_id = first_ch["chapter_id"]
        else:
            raise HTTPException(status_code=400, detail="No chapter found for this SuperAdmin")

    if not chapter_id:
        raise HTTPException(status_code=400, detail="Chapter ID required")

    # Cross-chapter mobile uniqueness check
    existing_mobile = await db.members.find_one({"primary_mobile": member.primary_mobile})
    if existing_mobile:
        if existing_mobile["chapter_id"] == chapter_id:
            raise HTTPException(status_code=400, detail="Member with this mobile number already exists in this chapter")
        other_ch = await db.chapters.find_one({"chapter_id": existing_mobile["chapter_id"]}, {"_id": 0, "name": 1})
        other_name = other_ch["name"] if other_ch else "another chapter"
        raise HTTPException(status_code=400, detail=f"Mobile already registered with {other_name}")

    # Auto-generate unique_member_id: BNI-{CODE}-{NNN}
    chapter = await db.chapters.find_one({"chapter_id": chapter_id}, {"_id": 0, "chapter_code": 1})
    chapter_code = chapter.get("chapter_code", "XXX") if chapter else "XXX"
    unique_member_id = await _generate_member_id(chapter_id, chapter_code)

    from uuid import uuid4

    if user["role"] == "admin":
        membership_status = "pending"
        sync_status = "Inactive"
    else:
        membership_status = "active"
        sync_status = "Active"

    member_data = {
        "member_id": str(uuid4()), "chapter_id": chapter_id,
        "unique_member_id": unique_member_id, "full_name": member.full_name,
        "primary_mobile": member.primary_mobile, "secondary_mobile": member.secondary_mobile,
        "email": member.email, "business_name": member.business_name,
        "business_category": member.business_category, "joining_date": member.joining_date,
        "renewal_date": member.renewal_date, "induction_fee": member.induction_fee,
        "membership_status": membership_status, "status": sync_status,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "bni_member_id": unique_member_id, "organization_id": chapter_id,
        "status_history": [{"action": "created", "from_status": None, "to_status": membership_status,
            "reason": "Member created", "changed_by": user.get("mobile", user.get("email", "system")),
            "timestamp": datetime.now(timezone.utc).isoformat()}],
        "archived": False, "transfer_from_chapter": None, "transfer_date": None,
    }

    try:
        await db.members.insert_one(member_data)

        # Auto-create login credentials: BNI@{last4digits}
        mobile = member.primary_mobile
        last4 = mobile[-4:] if len(mobile) >= 4 else mobile
        default_password = f"BNI@{last4}"
        existing_cred = await db.member_credentials.find_one({"mobile": mobile})
        if not existing_cred:
            await db.member_credentials.insert_one({
                "credential_id": str(uuid4()),
                "member_id": member_data["member_id"],
                "mobile": mobile,
                "password_hash": hash_password(default_password),
                "must_reset": True,
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": user.get("mobile", user.get("email", "system")),
            })

        return MemberResponse(**member_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to add member: {str(e)}")


@router.post("/admin/members/bulk")
async def bulk_add_members(members: List[MemberCreate], user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    from uuid import uuid4
    chapter_id = user["chapter_id"]
    chapter = await db.chapters.find_one({"chapter_id": chapter_id}, {"_id": 0, "chapter_code": 1})
    chapter_code = chapter.get("chapter_code", "XXX") if chapter else "XXX"

    # Track mobiles already used in this batch
    batch_mobiles = set()
    inserted = 0
    skipped = []
    errors = []
    duplicate_mobiles = []

    for idx, member in enumerate(members):
        # Cross-chapter mobile check
        if member.primary_mobile in batch_mobiles:
            skipped.append(f"Row {idx+1}: Duplicate mobile {member.primary_mobile} in batch")
            duplicate_mobiles.append(member.primary_mobile)
            continue
        existing_mobile = await db.members.find_one({"primary_mobile": member.primary_mobile})
        if existing_mobile:
            other_ch = await db.chapters.find_one({"chapter_id": existing_mobile["chapter_id"]}, {"_id": 0, "name": 1})
            other_name = other_ch["name"] if other_ch else "another chapter"
            skipped.append(f"Row {idx+1}: Mobile {member.primary_mobile} already registered with {other_name}")
            duplicate_mobiles.append(member.primary_mobile)
            continue

        # Auto-generate member ID
        unique_member_id = await _generate_member_id(chapter_id, chapter_code)

        member_data = {
            "member_id": str(uuid4()), "chapter_id": chapter_id,
            "unique_member_id": unique_member_id, "full_name": member.full_name,
            "primary_mobile": member.primary_mobile, "secondary_mobile": member.secondary_mobile,
            "status": member.status, "created_at": datetime.now(timezone.utc).isoformat(),
            "bni_member_id": unique_member_id, "organization_id": chapter_id,
        }
        try:
            await db.members.insert_one(member_data)
            inserted += 1
            batch_mobiles.add(member.primary_mobile)

            # Auto-create login credentials: BNI@{last4digits}
            mobile = member.primary_mobile
            last4 = mobile[-4:] if len(mobile) >= 4 else mobile
            default_password = f"BNI@{last4}"
            existing_cred = await db.member_credentials.find_one({"mobile": mobile})
            if not existing_cred:
                await db.member_credentials.insert_one({
                    "credential_id": str(uuid4()),
                    "member_id": member_data["member_id"],
                    "mobile": mobile,
                    "password_hash": hash_password(default_password),
                    "must_reset": True,
                    "is_active": True,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "updated_by": "system",
                })
        except Exception as e:
            errors.append(f"Row {idx+1}: Database error: {str(e)}")

    return {
        "message": "Bulk upload completed", "total_processed": len(members),
        "successfully_added": inserted, "duplicate_mobiles": duplicate_mobiles,
        "skipped": len(skipped), "errors": len(errors),
        "skipped_details": skipped if skipped else None, "error_details": errors if errors else None
    }


@router.get("/admin/members/template")
async def download_member_template(user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Members Template"
    headers = ["Full Name", "Primary Mobile", "Secondary Mobile", "Status"]
    ws.append(headers)

    header_fill = PatternFill(start_color="CF2030", end_color="CF2030", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws.append(["Sample Member 1", "9876543210", "9123456789", "Active"])
    ws.append(["Sample Member 2", "9999888877", "", "Active"])
    # Note row
    ws.append([])
    ws.append(["Note: Member IDs are auto-generated. Do not add an ID column."])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=members_template.xlsx"})


@router.post("/admin/members/upload")
async def upload_members_excel(file: UploadFile = File(...), user=Depends(get_current_user)):
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    from openpyxl import load_workbook
    from uuid import uuid4

    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active
    members_data = []
    errors = []
    skipped = []
    duplicate_mobiles = []

    chapter_id = user["chapter_id"]
    chapter = await db.chapters.find_one({"chapter_id": chapter_id}, {"_id": 0, "chapter_code": 1})
    chapter_code = chapter.get("chapter_code", "XXX") if chapter else "XXX"

    batch_mobiles = set()

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0] or not row[1]:
            continue
        # New column order: Full Name, Primary Mobile, Secondary Mobile, Status
        full_name = str(row[0]).strip()
        primary_mobile = str(int(row[1]) if isinstance(row[1], float) else row[1]).strip()
        secondary_mobile = str(int(row[2]) if isinstance(row[2], float) else row[2]).strip() if len(row) > 2 and row[2] else None
        status = str(row[3]).strip() if len(row) > 3 and row[3] else "Active"

        # Skip note rows
        if full_name.lower().startswith("note:"):
            continue

        # Check duplicate in batch
        if primary_mobile in batch_mobiles:
            skipped.append(f"Row {idx}: Duplicate mobile {primary_mobile} in Excel file")
            duplicate_mobiles.append(primary_mobile)
            continue

        # Cross-chapter mobile uniqueness
        existing_mobile = await db.members.find_one({"primary_mobile": primary_mobile})
        if existing_mobile:
            other_ch = await db.chapters.find_one({"chapter_id": existing_mobile["chapter_id"]}, {"_id": 0, "name": 1})
            other_name = other_ch["name"] if other_ch else "another chapter"
            skipped.append(f"Row {idx}: Mobile {primary_mobile} already registered with {other_name}")
            duplicate_mobiles.append(primary_mobile)
            continue

        try:
            unique_member_id = await _generate_member_id(chapter_id, chapter_code)
            member_data = {
                "member_id": str(uuid4()), "chapter_id": chapter_id,
                "unique_member_id": unique_member_id, "full_name": full_name,
                "primary_mobile": primary_mobile, "secondary_mobile": secondary_mobile,
                "status": status, "created_at": datetime.now(timezone.utc).isoformat(),
                "bni_member_id": unique_member_id, "organization_id": chapter_id,
            }
            members_data.append(member_data)
            batch_mobiles.add(primary_mobile)
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")

    inserted_count = 0
    for idx, member_data in enumerate(members_data):
        try:
            await db.members.insert_one(member_data)
            inserted_count += 1

            # Auto-create login credentials: BNI@{last4digits}
            mobile = member_data["primary_mobile"]
            last4 = mobile[-4:] if len(mobile) >= 4 else mobile
            default_password = f"BNI@{last4}"
            existing_cred = await db.member_credentials.find_one({"mobile": mobile})
            if not existing_cred:
                await db.member_credentials.insert_one({
                    "credential_id": str(uuid4()),
                    "member_id": member_data["member_id"],
                    "mobile": mobile,
                    "password_hash": hash_password(default_password),
                    "must_reset": True,
                    "is_active": True,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "updated_by": "system",
                })
        except Exception as e:
            errors.append(f"Insert error for row {idx+2}: {str(e)}")

    return {
        "message": f"{inserted_count} members uploaded successfully", "total": len(members_data),
        "duplicate_mobiles": duplicate_mobiles, "skipped": len(skipped), "errors": errors,
        "skipped_details": skipped if skipped else None
    }


@router.put("/admin/members/{member_id}", response_model=MemberResponse)
async def update_member(member_id: str, member: MemberUpdate, user=Depends(get_current_user)):
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    chapter_id = user.get("chapter_id")
    query = {"member_id": member_id}
    if chapter_id:
        query["chapter_id"] = chapter_id

    update_data = {k: v for k, v in member.dict(exclude_unset=True).items()}
    result = await db.members.update_one(query, {"$set": update_data})

    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Member not found")

    updated_member = await db.members.find_one({"member_id": member_id}, {"_id": 0})
    return MemberResponse(**updated_member)


@router.put("/admin/members/{member_id}/deactivate")
async def deactivate_member(member_id: str, user=Depends(get_current_user)):
    """Soft deactivate a member — sets membership_status to 'inactive'."""
    if user["role"] not in ("admin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")

    member = await db.members.find_one(
        {"member_id": member_id, "chapter_id": user["chapter_id"]}, {"_id": 0}
    )
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    new_status = "active" if member.get("membership_status") == "inactive" else "inactive"
    await db.members.update_one(
        {"member_id": member_id},
        {"$set": {"membership_status": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": f"Member {'reactivated' if new_status == 'active' else 'deactivated'} successfully", "new_status": new_status}


# ===== MOBILE AVAILABILITY CHECK =====

@router.get("/admin/check-mobile/{mobile}")
async def check_mobile_availability(mobile: str, user=Depends(get_current_user)):
    """Check if a mobile number is available (not registered in any chapter)."""
    if user["role"] not in ("admin", "superadmin", "developer"):
        raise HTTPException(status_code=403, detail="Forbidden")
    existing = await db.members.find_one({"primary_mobile": mobile}, {"_id": 0, "chapter_id": 1})
    if existing:
        chapter = await db.chapters.find_one({"chapter_id": existing["chapter_id"]}, {"_id": 0, "name": 1})
        return {"available": False, "existing_chapter": chapter.get("name", "Unknown") if chapter else "Unknown"}
    return {"available": True, "existing_chapter": None}


# ===== MEMBER ROLE MANAGEMENT =====
from models import MemberRoleUpdate
from routes.audit_log import log_audit


@router.put("/admin/members/{member_id}/role")
async def update_member_role(member_id: str, data: MemberRoleUpdate, request: Request, user=Depends(get_current_user)):
    """Update a member's chapter role. Enforces BNI hierarchy:
    - Developer: bypass (full access)
    - President role: can only be assigned by ED from SuperAdmin panel
    - Only Chapter President can assign other leadership roles
    - Cannot change your own role
    """
    now = datetime.now(timezone.utc).isoformat()

    # Developer bypass
    if user["role"] == "developer":
        pass  # allowed
    elif data.chapter_role == "president":
        raise HTTPException(status_code=403, detail="President is assigned by the Executive Director from the SuperAdmin panel")
    elif user["role"] == "superadmin":
        raise HTTPException(status_code=403, detail="Use the SuperAdmin panel to manage chapter leadership")
    elif user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Forbidden")
    elif user.get("chapter_role") != "president":
        raise HTTPException(status_code=403, detail="Only the Chapter President can assign roles")
    elif user.get("member_id") == member_id:
        raise HTTPException(status_code=400, detail="Cannot change your own role")

    chapter_id = user.get("chapter_id")
    if not chapter_id:
        raise HTTPException(status_code=400, detail="Chapter ID required")

    valid_roles = ["vice_president", "secretary", "treasurer", "secretary_treasurer", "lvh", "member"]
    # Developer can also assign president
    if user["role"] == "developer":
        valid_roles.insert(0, "president")
    if data.chapter_role not in valid_roles:
        raise HTTPException(status_code=400, detail=f"Invalid role. Must be one of: {', '.join(valid_roles)}")

    member = await db.members.find_one({"member_id": member_id, "chapter_id": chapter_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    if member.get("membership_status") != "active":
        raise HTTPException(status_code=400, detail="Member must be active to be assigned a role")

    old_role = member.get("chapter_role", "member")
    if old_role == data.chapter_role:
        return {"message": f"Member already has role {data.chapter_role}", "previous_holder": None}

    # Check if role is already assigned to another member (except 'member' role)
    current_holder = None
    if data.chapter_role != "member":
        existing = await db.members.find_one(
            {"chapter_id": chapter_id, "chapter_role": data.chapter_role, "member_id": {"$ne": member_id}},
            {"_id": 0, "full_name": 1, "member_id": 1, "chapter_role": 1}
        )
        if existing:
            current_holder = existing.get("full_name", "Unknown")
            # Demote previous holder and track history
            await db.members.update_one(
                {"member_id": existing["member_id"]},
                {"$set": {
                    "chapter_role": "member",
                    "role_assigned_by": user.get("member_id", user.get("mobile", "")),
                    "role_assigned_date": now,
                    "role_previous": existing.get("chapter_role", "member"),
                },
                "$push": {"role_change_history": {
                    "from_role": existing.get("chapter_role", "member"),
                    "to_role": "member",
                    "changed_by": user.get("member_id", user.get("mobile", "")),
                    "changed_by_role": user.get("chapter_role", user.get("role", "")),
                    "timestamp": now,
                    "reason": f"Replaced by {member.get('full_name', member_id)} as {data.chapter_role}",
                }}}
            )

    # Assign new role and track history
    await db.members.update_one(
        {"member_id": member_id},
        {"$set": {
            "chapter_role": data.chapter_role,
            "role_assigned_by": user.get("member_id", user.get("mobile", "")),
            "role_assigned_date": now,
            "role_previous": old_role,
        },
        "$push": {"role_change_history": {
            "from_role": old_role,
            "to_role": data.chapter_role,
            "changed_by": user.get("member_id", user.get("mobile", "")),
            "changed_by_role": user.get("chapter_role", user.get("role", "")),
            "timestamp": now,
            "reason": data.reason if hasattr(data, "reason") and data.reason else "",
        }}}
    )

    # Global audit log
    client_ip = request.client.host if request.client else ""
    await log_audit(
        user_id=user.get("mobile", ""),
        role=user.get("role", ""),
        action="role_assignment",
        entity_type="member",
        entity_id=member_id,
        details=f"Changed role from {old_role} to {data.chapter_role} for {member.get('full_name', '')}. Previous holder: {current_holder or 'none'}",
        ip=client_ip,
    )

    msg = f"Role updated to {data.chapter_role}"
    if current_holder:
        msg += f". Previous holder {current_holder} has been set to 'member'."

    return {"message": msg, "previous_holder": current_holder}
