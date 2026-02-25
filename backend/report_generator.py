from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
from datetime import datetime
import pytz

IST = pytz.timezone('Asia/Kolkata')

# Define colors
YELLOW_COLOR = "FFFF00"
RED_COLOR = "FF0000"
WHITE_COLOR = "FFFFFF"
BLACK_COLOR = "000000"
HEADER_BLUE = "4472C4"  # Blue for headers
LIGHT_GRAY = "F2F2F2"   # For alternating rows

# Define borders
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

def format_date_ddmmmyy(dt):
    """Format date as DD-MMM-YY (e.g., 16-Dec-25)"""
    if dt is None:
        return ""
    if isinstance(dt, str):
        dt = datetime.fromisoformat(dt)
    return dt.strftime("%d-%b-%y")

def format_time_hmmss(dt):
    """Format time as H:MM:SS (e.g., 6:30:00)"""
    if dt is None:
        return ""
    if isinstance(dt, str):
        dt = datetime.fromisoformat(dt)
    if hasattr(dt, 'strftime'):
        # Remove leading zero from hour
        time_str = dt.strftime("%H:%M:%S")
        if time_str.startswith('0'):
            time_str = time_str[1:]
        return time_str
    return ""

def format_time_12hr(dt):
    """Format time as HH:MM:SS 12-hour format for meeting info"""
    if dt is None:
        return ""
    if isinstance(dt, str):
        dt = datetime.fromisoformat(dt)
    return dt.strftime("%I:%M:%S %p") if hasattr(dt, 'strftime') else ""

def parse_datetime(value, base_date=None):
    """
    Parse various datetime formats:
    - ISO format: 2025-12-07T10:19:00.000Z
    - Simple time: 10:00
    - Date string: 2025-12-08
    """
    if value is None:
        return None
    
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return IST.localize(value)
        return value.astimezone(IST)
    
    if not isinstance(value, str):
        return None
    
    # Try ISO format first
    try:
        # Handle Z suffix
        if value.endswith('Z'):
            value = value[:-1] + '+00:00'
        dt = datetime.fromisoformat(value)
        if dt.tzinfo is None:
            dt = IST.localize(dt)
        return dt.astimezone(IST)
    except:
        pass
    
    # Try simple time format (HH:MM or HH:MM:SS)
    if ':' in value and len(value) <= 8:
        try:
            if base_date:
                if isinstance(base_date, str):
                    base_date = datetime.fromisoformat(base_date.replace('Z', '+00:00'))
                time_parts = value.split(':')
                hour = int(time_parts[0])
                minute = int(time_parts[1])
                second = int(time_parts[2]) if len(time_parts) > 2 else 0
                dt = base_date.replace(hour=hour, minute=minute, second=second, microsecond=0)
                if dt.tzinfo is None:
                    dt = IST.localize(dt)
                return dt.astimezone(IST)
        except:
            pass
    
    # Try date only format
    try:
        dt = datetime.strptime(value[:10], "%Y-%m-%d")
        return IST.localize(dt)
    except:
        pass
    
    return None

def get_attendance_status(att, cutoff_time):
    """
    Determine status and type based on entry time vs cutoff time
    Returns: (status, type)
    - Status: "Present" or "Absent"
    - Type: "On Time", "Late", or "Absent"
    """
    if att is None:
        return ("Absent", "Absent")
    
    att_time = datetime.fromisoformat(att["timestamp"]).astimezone(IST)
    cutoff = cutoff_time
    
    # Status is always "Present" if attended (even for substitutes)
    if att_time.time() <= cutoff.time():
        return ("Present", "On Time")
    else:
        return ("Present", "Late")


def generate_excel_report(meeting: dict, members: list, attendance: list, chapter_name: str = "") -> bytes:
    """Generate Excel report with 2 sheets - Attendance Report and Visitor Report (A4 Landscape)"""
    wb = Workbook()
    
    # Parse meeting times with robust parser
    meeting_date = parse_datetime(meeting.get("date"))
    if meeting_date is None:
        meeting_date = datetime.now(IST)
    
    start_time = parse_datetime(meeting.get("start_time"), meeting_date)
    if start_time is None:
        start_time = meeting_date
    
    cutoff_time = parse_datetime(meeting.get("late_cutoff_time"), meeting_date)
    if cutoff_time is None:
        cutoff_time = start_time
    
    end_time = parse_datetime(meeting.get("end_time"), meeting_date)
    if end_time is None:
        end_time = meeting_date
    
    # Check if meeting has ended
    now_ist = datetime.now(IST)
    meeting_ended = now_ist > end_time
    
    # Define fills
    yellow_fill = PatternFill(start_color=YELLOW_COLOR, end_color=YELLOW_COLOR, fill_type="solid")
    red_fill = PatternFill(start_color=RED_COLOR, end_color=RED_COLOR, fill_type="solid")
    header_fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
    alt_row_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    
    # Define fonts
    title_font = Font(bold=True, size=14, color=BLACK_COLOR)
    header_font = Font(bold=True, size=10, color=WHITE_COLOR)
    data_font = Font(size=9)
    summary_font = Font(bold=True, size=11, color=BLACK_COLOR)
    info_font = Font(size=10)
    
    # ========== SHEET 1: ATTENDANCE REPORT ==========
    ws1 = wb.active
    ws1.title = "Attendance Report"
    
    # Set page orientation to Landscape
    ws1.page_setup.orientation = 'landscape'
    ws1.page_setup.paperSize = ws1.PAPERSIZE_A4
    ws1.page_setup.fitToPage = True
    ws1.page_setup.fitToWidth = 1
    ws1.page_setup.fitToHeight = 0
    
    # Set margins (15mm left/right, 10mm top/bottom)
    ws1.page_margins = PageMargins(left=0.59, right=0.59, top=0.39, bottom=0.39)
    
    # Row 1: Title (BNI ATTENDANCE REPORT) - Yellow background
    ws1.merge_cells('A1:I1')
    title_cell = ws1['A1']
    title_cell.value = "BNI ATTENDANCE REPORT"
    title_cell.font = title_font
    title_cell.fill = yellow_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 25
    
    # Row 2: Chapter Name
    ws1.merge_cells('A2:I2')
    ws1['A2'].value = f"Chapter: {chapter_name}" if chapter_name else "Chapter: "
    ws1['A2'].font = info_font
    ws1['A2'].alignment = Alignment(horizontal="left")
    
    # Row 3: Meeting Date
    ws1['A3'].value = "Meeting Date:"
    ws1['B3'].value = format_date_ddmmmyy(meeting_date)
    ws1['A3'].font = info_font
    ws1['B3'].font = info_font
    
    # Row 4: Start Time
    ws1['A4'].value = "Start Time:"
    ws1['B4'].value = format_time_hmmss(start_time)
    ws1['A4'].font = info_font
    ws1['B4'].font = info_font
    
    # Row 5: Cutoff Time
    ws1['A5'].value = "Cutoff Time:"
    ws1['B5'].value = format_time_hmmss(cutoff_time)
    ws1['A5'].font = info_font
    ws1['B5'].font = info_font
    
    # Row 6: End Time
    ws1['A6'].value = "End Time:"
    ws1['B6'].value = format_time_hmmss(end_time)
    ws1['A6'].font = info_font
    ws1['B6'].font = info_font
    
    # Row 7: Empty row
    
    # Row 8: Table Headers - Blue background, white text (A4 Landscape optimized)
    headers1 = [
        "Sr. No.", "Member ID", "Member Name", "Status", "Date", "Timestamp",
        "Type", "Substitute Name", "Substitute Mobile"
    ]
    
    header_row = 8
    for col_idx, header in enumerate(headers1, 1):
        cell = ws1.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill  # Blue header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    ws1.row_dimensions[header_row].height = 25
    
    # Filter member and substitute attendance
    member_substitute_att = [a for a in attendance if a["type"] in ["member", "substitute"]]
    attendance_map = {a["unique_member_id"]: a for a in member_substitute_att if a.get("unique_member_id")}
    
    # Separate present and absent members
    present_members = []
    absent_members = []
    
    for member in members:
        att = attendance_map.get(member["unique_member_id"])
        if att:
            # Member is present - store with timestamp for sorting
            att_time = datetime.fromisoformat(att["timestamp"]).astimezone(IST)
            present_members.append({
                'member': member,
                'attendance': att,
                'timestamp': att_time
            })
        else:
            # Member is absent
            absent_members.append(member)
    
    # Sort present members by timestamp (earliest first)
    present_members.sort(key=lambda x: x['timestamp'])
    
    # Sort absent members alphabetically by name
    absent_members.sort(key=lambda x: (x.get("full_name") or "").lower())
    
    # Add member data with serial numbers
    current_row = header_row + 1
    sr_no = 1
    row_count = 0
    
    # First: Add present members (sorted by timestamp)
    for item in present_members:
        member = item['member']
        att = item['attendance']
        att_time = item['timestamp']
        status, type_value = get_attendance_status(att, cutoff_time)
        
        row_data = [
            sr_no,
            member["unique_member_id"],
            member["full_name"],
            "Present",
            format_date_ddmmmyy(att_time),
            format_time_hmmss(att_time),
            type_value,
            att.get("substitute_name", "") or "",
            att.get("substitute_mobile", "") or ""
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=current_row, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center" if col_idx in [1, 4, 5, 6, 7] else "left", vertical="center")
            cell.border = thin_border
            # Alternating row colors
            if row_count % 2 == 1:
                cell.fill = alt_row_fill
        
        current_row += 1
        sr_no += 1
        row_count += 1
    
    # Then: Add absent members (at the bottom)
    for member in absent_members:
        row_data = [
            sr_no,
            member["unique_member_id"],
            member["full_name"],
            "Absent",
            "", "", "Absent", "", ""
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=current_row, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center" if col_idx in [1, 4, 5, 6, 7] else "left", vertical="center")
            cell.border = thin_border
            # Alternating row colors
            if row_count % 2 == 1:
                cell.fill = alt_row_fill
        
        current_row += 1
        sr_no += 1
        row_count += 1
    
    # Summary Section - Yellow background
    summary_start_row = current_row + 1
    
    # Calculate summary values
    total_members = len(members)
    # Count members who have attendance (both direct and via substitute)
    members_with_attendance = set()
    for att in member_substitute_att:
        if att.get("unique_member_id"):
            members_with_attendance.add(att["unique_member_id"])
    
    present_count = len(members_with_attendance)  # Present includes those with substitutes
    absent_count = total_members - present_count
    if absent_count < 0:
        absent_count = 0
    total_attendance = present_count  # Total attendance = Present (substitutes are counted as present)
    
    # Summary header
    ws1.merge_cells(f'A{summary_start_row}:I{summary_start_row}')
    summary_title = ws1.cell(row=summary_start_row, column=1)
    summary_title.value = "SUMMARY"
    summary_title.font = summary_font
    summary_title.fill = yellow_fill
    summary_title.alignment = Alignment(horizontal="center")
    
    # Summary data
    summary_data = [
        ("Total Members:", total_members),
        ("Present:", present_count),
        ("Absent:", absent_count),
        ("Total Attendance:", total_attendance)
    ]
    
    for i, (label, value) in enumerate(summary_data):
        row_num = summary_start_row + 1 + i
        ws1.cell(row=row_num, column=1).value = label
        ws1.cell(row=row_num, column=1).font = summary_font
        ws1.cell(row=row_num, column=1).fill = yellow_fill
        ws1.cell(row=row_num, column=2).value = value
        ws1.cell(row=row_num, column=2).font = summary_font
        ws1.cell(row=row_num, column=2).fill = yellow_fill
        for col in range(3, 10):
            ws1.cell(row=row_num, column=col).fill = yellow_fill
    
    # Auto-adjust column widths for Sheet 1 (A4 Landscape optimized)
    # Column percentages: Sr.No(5%), MemberID(8%), Name(20%), Status(10%), Date(10%), Time(12%), Type(8%), SubName(15%), SubMobile(12%)
    column_widths = [7, 12, 30, 12, 14, 14, 12, 22, 18]
    for i, width in enumerate(column_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze header row
    ws1.freeze_panes = ws1[f'A{header_row + 1}']
    
    # Add auto filter
    ws1.auto_filter.ref = f'A{header_row}:I{current_row - 1}'
    
    # ========== SHEET 2: VISITOR REPORT ==========
    ws2 = wb.create_sheet(title="Visitor Report")
    
    # Set page orientation to Landscape
    ws2.page_setup.orientation = 'landscape'
    ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
    ws2.page_setup.fitToPage = True
    ws2.page_setup.fitToWidth = 1
    ws2.page_setup.fitToHeight = 0
    ws2.page_margins = PageMargins(left=0.59, right=0.59, top=0.39, bottom=0.39)
    
    # Row 1: Title (BNI VISITOR REPORT) - Yellow background
    ws2.merge_cells('A1:H1')
    v_title_cell = ws2['A1']
    v_title_cell.value = "BNI VISITOR REPORT"
    v_title_cell.font = title_font
    v_title_cell.fill = yellow_fill
    v_title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 25
    
    # Row 2: Chapter Name
    ws2.merge_cells('A2:H2')
    ws2['A2'].value = f"Chapter: {chapter_name}" if chapter_name else "Chapter: "
    ws2['A2'].font = info_font
    ws2['A2'].alignment = Alignment(horizontal="left")
    
    # Row 3: Date
    ws2['A3'].value = "Date:"
    ws2['B3'].value = format_date_ddmmmyy(meeting_date)
    ws2['A3'].font = info_font
    ws2['B3'].font = info_font
    
    # Row 4: Empty row
    
    # Row 5: Table Headers (Blue header)
    headers2 = [
        "Sr. No.", "Visitor Name", "Visitor Mobile Number", "Visitor Company",
        "Invited By Member Id", "BNI Member Name", "Date", "Timestamp"
    ]
    
    v_header_row = 5
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=v_header_row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill  # Blue header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    ws2.row_dimensions[v_header_row].height = 25
    
    # Filter visitor attendance
    visitor_att = [a for a in attendance if a["type"] == "visitor"]
    
    # Create a map of member IDs to names
    member_name_map = {m["unique_member_id"]: m["full_name"] for m in members}
    
    # Add visitor data
    v_current_row = v_header_row + 1
    v_sr_no = 1
    v_row_count = 0
    
    for visitor in visitor_att:
        att_time = datetime.fromisoformat(visitor["timestamp"]).astimezone(IST)
        invited_by_name = visitor.get("invited_by_member_name", "") or member_name_map.get(visitor.get("invited_by_member_id", ""), "")
        
        row_data = [
            v_sr_no,
            visitor.get("visitor_name", "") or "",
            visitor.get("visitor_mobile", "") or "",
            visitor.get("visitor_company", "") or "",
            visitor.get("invited_by_member_id", "") or "",
            invited_by_name,
            format_date_ddmmmyy(att_time),
            format_time_hmmss(att_time)
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=v_current_row, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center" if col_idx in [1, 7, 8] else "left", vertical="center")
            cell.border = thin_border
            # Alternating row colors
            if v_row_count % 2 == 1:
                cell.fill = alt_row_fill
        
        v_current_row += 1
        v_sr_no += 1
        v_row_count += 1
    
    # Auto-adjust column widths for Sheet 2 (A4 Landscape optimized)
    # Column percentages: Sr.No(5%), VisitorName(15%), Mobile(12%), Company(20%), InvitedByID(12%), BNIMember(18%), Date(8%), Time(10%)
    v_column_widths = [7, 22, 18, 28, 18, 26, 14, 14]
    for i, width in enumerate(v_column_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze header row
    ws2.freeze_panes = ws2[f'A{v_header_row + 1}']
    
    # Add auto filter
    ws2.auto_filter.ref = f'A{v_header_row}:H{v_current_row - 1}'
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_pdf_report(meeting: dict, members: list, attendance: list, chapter_name: str = "") -> bytes:
    """Generate PDF report with 2 pages - Attendance Report and Visitor Report (A4 Landscape)"""
    from reportlab.platypus import PageBreak
    
    buffer = BytesIO()
    
    # A4 Landscape with specified margins (15mm top/bottom, 10mm left/right)
    top_margin = 15 * mm
    bottom_margin = 15 * mm
    left_margin = 10 * mm
    right_margin = 10 * mm
    
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),  # A4 Landscape (297mm x 210mm)
        topMargin=top_margin,
        bottomMargin=bottom_margin,
        leftMargin=left_margin,
        rightMargin=right_margin
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Title'],
        fontSize=14,
        textColor=colors.black,
        alignment=TA_CENTER,
        spaceAfter=6
    )
    
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.black,
        alignment=TA_LEFT,
        spaceAfter=2
    )
    
    summary_style = ParagraphStyle(
        'SummaryStyle',
        parent=styles['Normal'],
        fontSize=11,  # Larger font for summary
        textColor=colors.black,
        alignment=TA_LEFT
    )
    
    # Parse meeting times with robust parser
    meeting_date = parse_datetime(meeting.get("date"))
    if meeting_date is None:
        meeting_date = datetime.now(IST)
    
    start_time = parse_datetime(meeting.get("start_time"), meeting_date)
    if start_time is None:
        start_time = meeting_date
    
    cutoff_time = parse_datetime(meeting.get("late_cutoff_time"), meeting_date)
    if cutoff_time is None:
        cutoff_time = start_time
    
    end_time = parse_datetime(meeting.get("end_time"), meeting_date)
    if end_time is None:
        end_time = meeting_date
    
    # Check if meeting ended
    now_ist = datetime.now(IST)
    meeting_ended = now_ist > end_time
    
    # ========== PAGE 1: ATTENDANCE REPORT ==========
    
    # Title with yellow background
    title_data = [["BNI ATTENDANCE REPORT"]]
    title_table = Table(title_data, colWidths=[doc.width])
    title_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.yellow),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 14),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 4*mm))
    
    # Meeting Info
    info_data = [
        [f"Chapter: {chapter_name}" if chapter_name else "Chapter: "],
        [f"Meeting Date: {format_date_ddmmmyy(meeting_date)}"],
        [f"Start Time: {format_time_hmmss(start_time)}"],
        [f"Cutoff Time: {format_time_hmmss(cutoff_time)}"],
        [f"End Time: {format_time_hmmss(end_time)}"]
    ]
    info_table = Table(info_data, colWidths=[doc.width])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 4*mm))
    
    # Attendance Table
    member_substitute_att = [a for a in attendance if a["type"] in ["member", "substitute"]]
    attendance_map = {a["unique_member_id"]: a for a in member_substitute_att if a.get("unique_member_id")}
    
    # Separate present and absent members
    present_members = []
    absent_members = []
    
    for member in members:
        att = attendance_map.get(member["unique_member_id"])
        if att:
            att_time = datetime.fromisoformat(att["timestamp"]).astimezone(IST)
            present_members.append({
                'member': member,
                'attendance': att,
                'timestamp': att_time
            })
        else:
            absent_members.append(member)
    
    # Sort present members by timestamp (earliest first)
    present_members.sort(key=lambda x: x['timestamp'])
    
    # Sort absent members alphabetically by name
    absent_members.sort(key=lambda x: (x.get("full_name") or "").lower())
    
    # Table headers - "Type" instead of "Late Type"
    table_data = [[
        'Sr.\nNo.', 'Member\nID', 'Member Name', 'Status', 'Date', 'Timestamp',
        'Type', 'Substitute\nName', 'Substitute\nMobile'
    ]]
    
    sr_no = 1
    
    # First: Add present members (sorted by timestamp)
    for item in present_members:
        member = item['member']
        att = item['attendance']
        att_time = item['timestamp']
        status, type_value = get_attendance_status(att, cutoff_time)
        
        row = [
            sr_no,
            member["unique_member_id"],
            member["full_name"][:20],
            "Present",
            format_date_ddmmmyy(att_time),
            format_time_hmmss(att_time),
            type_value,
            (att.get("substitute_name", "") or "")[:15],
            att.get("substitute_mobile", "") or ""
        ]
        table_data.append(row)
        sr_no += 1
    
    # Then: Add absent members (at the bottom)
    for member in absent_members:
        row = [
            sr_no,
            member["unique_member_id"],
            member["full_name"][:20],
            "Absent",
            "", "", "Absent", "", ""
        ]
        table_data.append(row)
        sr_no += 1
    
    # Column widths for A4 Landscape (optimized for 297mm - margins)
    # Sr.No(5%), MemberID(8%), Name(20%), Status(10%), Date(10%), Time(12%), Type(8%), SubName(15%), SubMobile(12%)
    col_widths = [40, 60, 140, 70, 70, 80, 60, 110, 90]
    
    # Header background color (Blue)
    header_blue = colors.Color(0.267, 0.447, 0.769)  # #4472C4
    
    table1 = Table(table_data, colWidths=col_widths, repeatRows=1)
    table1.setStyle(TableStyle([
        # Header style - Blue background, white text
        ('BACKGROUND', (0, 0), (-1, 0), header_blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        
        # Data rows
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),  # Larger font for landscape
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        
        # Padding
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]))
    
    elements.append(table1)
    elements.append(Spacer(1, 6*mm))
    
    # Summary Section - Yellow background with larger font
    total_members = len(members)
    # Count members who have attendance (both direct and via substitute)
    members_with_attendance = set()
    for att in member_substitute_att:
        if att.get("unique_member_id"):
            members_with_attendance.add(att["unique_member_id"])
    
    present_count = len(members_with_attendance)  # Present includes those with substitutes
    absent_count = total_members - present_count
    if absent_count < 0:
        absent_count = 0
    total_attendance = present_count  # Total attendance = Present
    attendance_percent = round((present_count / total_members * 100), 1) if total_members > 0 else 0
    
    summary_data = [
        ["SUMMARY", "", "", "", "", ""],
        ["Total Members:", str(total_members), "Present:", str(present_count), "Attendance %:", f"{attendance_percent}%"],
        ["Absent:", str(absent_count), "Total Attendance:", str(total_attendance), "", ""]
    ]
    
    summary_table = Table(summary_data, colWidths=[100, 60, 100, 60, 100, 60])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.yellow),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),  # Larger font (11-12pt)
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('SPAN', (0, 0), (-1, 0)),  # Merge SUMMARY row
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
    ]))
    elements.append(summary_table)
    
    # ========== PAGE 2: VISITOR REPORT ==========
    elements.append(PageBreak())
    
    # Visitor Title with yellow background
    v_title_data = [["BNI VISITOR REPORT"]]
    v_title_table = Table(v_title_data, colWidths=[doc.width])
    v_title_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.yellow),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(v_title_table)
    elements.append(Spacer(1, 4*mm))
    
    # Visitor Info
    v_info_data = [
        [f"Chapter: {chapter_name}" if chapter_name else "Chapter: "],
        [f"Date: {format_date_ddmmmyy(meeting_date)}"]
    ]
    v_info_table = Table(v_info_data, colWidths=[doc.width])
    v_info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
    ]))
    elements.append(v_info_table)
    elements.append(Spacer(1, 4*mm))
    
    # Visitor Table
    visitor_att = [a for a in attendance if a["type"] == "visitor"]
    member_name_map = {m["unique_member_id"]: m["full_name"] for m in members}
    
    visitor_table_data = [[
        'Sr.\nNo.', 'Visitor Name', 'Visitor Mobile\nNumber', 'Visitor Company',
        'Invited By\nMember Id', 'BNI Member\nName', 'Date', 'Timestamp'
    ]]
    
    v_sr_no = 1
    for visitor in visitor_att:
        att_time = datetime.fromisoformat(visitor["timestamp"]).astimezone(IST)
        invited_by_name = visitor.get("invited_by_member_name", "") or member_name_map.get(visitor.get("invited_by_member_id", ""), "")
        
        row = [
            v_sr_no,
            (visitor.get("visitor_name", "") or "")[:25],  # More space in landscape
            visitor.get("visitor_mobile", "") or "",
            (visitor.get("visitor_company", "") or "")[:25],
            visitor.get("invited_by_member_id", "") or "",
            invited_by_name[:25] if invited_by_name else "",
            format_date_ddmmmyy(att_time),
            format_time_hmmss(att_time)
        ]
        visitor_table_data.append(row)
        v_sr_no += 1
    
    # Column widths for visitor table (A4 Landscape optimized)
    # Sr.No(5%), VisitorName(15%), Mobile(12%), Company(20%), InvitedByID(10%), BNIMember(18%), Date(10%), Time(10%)
    v_col_widths = [40, 110, 85, 140, 70, 130, 75, 70]
    
    if len(visitor_table_data) > 1:
        table2 = Table(visitor_table_data, colWidths=v_col_widths, repeatRows=1)
        table2.setStyle(TableStyle([
            # Header style - Blue background, white text
            ('BACKGROUND', (0, 0), (-1, 0), header_blue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),  # Larger font for landscape
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Padding
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))
        elements.append(table2)
        
        # Visitor count - Yellow background
        elements.append(Spacer(1, 6*mm))
        v_summary_data = [["Total Visitors:", str(len(visitor_att))]]
        v_summary_table = Table(v_summary_data, colWidths=[120, 60])
        v_summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.yellow),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(v_summary_table)
    else:
        no_visitors = Paragraph("<i>No visitors recorded for this meeting</i>", info_style)
        elements.append(no_visitors)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
